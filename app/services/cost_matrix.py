"""Template-specific cost-matrix discovery, mapping, and generation.

Buyer workbooks do not share a schema.  The inspector records the exact
structure of each uploaded workbook and proposes writable targets from local
labels, formats, and formulas.  No target is populated until a user-approved
mapping and a current reviewed pricing basis both exist.

Source workbooks are immutable.  Generation patches a copy at the OOXML cell
level so unrelated workbook parts (styles, hidden sheets, drawings, metadata,
and formulas) remain byte-for-byte untouched wherever possible.
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import math
import os
import re
import zipfile
from collections.abc import Iterable
from copy import deepcopy
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path, PurePosixPath
from typing import Any
from uuid import uuid4
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape as xml_escape

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.datetime import to_excel
from sqlalchemy import event, func, select
from sqlalchemy.orm import Session, selectinload

from app.config import RFP_PACKAGES_DIR
from app.core.company_profile import get_company_profile
from app.core.enums import ProposalStatus, RfpDocumentType
from app.db.session import session_scope
from app.models import (
    CostMatrixArtifact,
    CostMatrixOutput,
    PricingPackage,
    Proposal,
    RfpPackageDocument,
)
from app.services.proposal_access import (
    acquire_proposal_write_fence,
    ensure_proposal_mutable,
    proposal_write_lock,
)
from app.services.review_freshness import (
    get_cost_review_freshness,
    payment_cost_review_is_current,
    payment_market_scan_is_current,
)
from app.services.storage_safety import require_contained_file

log = logging.getLogger(__name__)

ANALYSIS_VERSION = "cost-matrix-v3"
COST_MATRIX_INTAKE_POLICY_VERSION = "cost-matrix-intake-v2"
COST_MATRIX_ROLE = "cost_matrix"
RECONCILIATION_MAPPING_KEY = "__reconciliation__"

MAX_WORKBOOK_BYTES = 30 * 1024 * 1024
MAX_ZIP_ENTRIES = 5_000
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_INSPECTED_CELLS = 250_000

STATUS_DETECTED = "detected"
STATUS_NEEDS_CONFIRMATION = "needs_confirmation"
STATUS_DISMISSED = "dismissed"
STATUS_MAPPING_REQUIRED = "mapping_required"
STATUS_WAITING_FOR_COSTS = "waiting_for_costs"
STATUS_READY = "ready"
STATUS_GENERATED = "generated"
STATUS_STALE = "stale"
STATUS_ERROR = "error"

_FINANCIAL_WORDS = re.compile(
    r"\b(cost|price|pricing|rate|fee|amount|extended|extension|subtotal|total|"
    r"labor|travel|expense|odc|unit\s*price|bid)\b",
    re.IGNORECASE,
)
_FINANCIAL_HEADER = re.compile(
    r"\b(cost|price|pricing|rate|fee|amount|extended|extension|subtotal|total|"
    r"unit|markup|discount)\b",
    re.IGNORECASE,
)
_STRONG_PRICE_WORDS = re.compile(
    r"\b(cost|price|pricing|fees?|charges?|amount|extended|extension|subtotal|"
    r"markup|discount|remuneration|compensation|quotes?|bids?|budget|ceiling|"
    r"billing\s+rate|hourly\s+rate|unit\s+(?:price|rate))\b",
    re.I,
)
_PRICE_CONTEXT = re.compile(
    r"\b(cost|price|pricing|fees?|charges?|amount|extended|extension|subtotal|"
    r"markup|discount|remuneration|compensation|quotes?|bids?|budget|ceiling|"
    r"billing\s+rate|hourly\s+rate|unit\s+(?:price|rate))\b",
    re.I,
)
_MATRIX_FILENAME = re.compile(
    r"(?:cost|price|pricing|fee|bid).*(?:matrix|schedule|sheet|template|form)|"
    r"(?:matrix|schedule|sheet|template|form).*(?:cost|price|pricing|fee|bid)",
    re.IGNORECASE,
)
_POSSIBLE_MATRIX_FILENAME = re.compile(
    r"\b(cost|price|pricing|rate|fee|bid|offer|clin)\b.*"
    r"\b(matrix|schedule|sheet|template|form|proposal|breakdown)\b|"
    r"\b(matrix|schedule|sheet|template|form|proposal|breakdown)\b.*"
    r"\b(cost|price|pricing|rate|fee|bid|offer|clin)\b",
    re.IGNORECASE,
)
_METADATA_LABELS: tuple[tuple[str, re.Pattern[str]], ...] = (
    (
        "company_legal_name",
        re.compile(
            r"^\s*(?:(offeror|vendor|bidder|company)\s+(?:legal\s+)?name|"
            r"legal\s+name\s+of\s+(?:the\s+)?(?:offeror|vendor|bidder))\s*:?[\s_]*$",
            re.I,
        ),
    ),
    (
        "matrix_date",
        re.compile(
            r"^\s*(date|submission\s+date|bid\s+date|proposal\s+date)"
            r"(?:\s*\([^)]*\))?\s*:?[\s_]*$",
            re.I,
        ),
    ),
)
_TOTAL_LABEL = re.compile(r"^\s*(grand\s+)?(sub)?total\b", re.I)
_VALUE_PLACEHOLDER = re.compile(
    r"^\s*(?:\$|usd\s*)?(?:[_-]{3,}|\$?\s*0(?:\.0+)?\s*[_-]+)\s*$",
    re.I,
)
_NUMBER_BODY = r"(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?:[eE][+-]?\d+)?"
_CURRENCY_CODE = (
    r"(?:USD|EUR|GBP|CAD|AUD|CHF|JPY|CNY|RMB|INR|KRW|MXN|BRL|SEK|NOK|DKK|"
    r"ZAR|SGD|HKD|NZD|AED|SAR)"
)
_FINANCIAL_VALUE_TOKEN = re.compile(
    rf"(?<![A-Za-z0-9])(?:"
    rf"(?:[$\u20ac\u00a3\u00a5]|US\$)\s*[+-]?{_NUMBER_BODY}\s*[KMB]?|"
    rf"{_CURRENCY_CODE}\s*[$\u20ac\u00a3\u00a5]?\s*"
    rf"[+-]?{_NUMBER_BODY}\s*[KMB]?|"
    rf"[+-]?{_NUMBER_BODY}\s*[KMB]?\s*{_CURRENCY_CODE}"
    rf")(?![A-Za-z0-9])",
    re.I,
)
_PLAIN_NUMERIC_TEXT = re.compile(r"^\s*[+-]?[\d,]+(?:\.\d+)?\s*$")
_PLAIN_FINANCIAL_NUMBER_TOKEN = re.compile(
    rf"(?<![A-Za-z0-9.])[+-]?{_NUMBER_BODY}\s*[KMB]?"
    rf"(?![A-Za-z0-9]|\.\d)",
    re.I,
)
_SAFE_REQUIREMENT_NUMBER_TOKEN = re.compile(
    r"\b(?:weights?|weighted?|scores?|points?)\s*(?::|-)?\s*"
    r"(?:(?:is|at|of)\s+)?\d+(?:\.\d+)?\b|"
    r"\b(?:FAR|DFARS|CFR|CLIN|section|clause|item)\s+"
    r"[A-Za-z0-9]+(?:[.\-/][A-Za-z0-9]+)*\b|"
    r"\b(?:volume|attachment|exhibit|appendix|schedule|tab|part|RFP|RFQ|"
    r"solicitation)\s+(?:no\.?\s*)?[A-Za-z0-9]+(?:[.\-/][A-Za-z0-9]+)*\b|"
    r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|June?|July?|"
    r"Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{1,2}(?:st|nd|rd|th)?(?:,\s*\d{4})?\b|"
    r"(?<![A-Za-z0-9])\d{1,4}[/-]\d{1,2}[/-]\d{2,4}(?![A-Za-z0-9])|"
    r"(?<![A-Za-z0-9])\d{1,2}:\d{2}\s*(?:AM|PM)?\b|"
    r"(?<![A-Za-z0-9])\d+(?:\.\d+)?\s*(?:%(?!\w)|(?:percent(?:age)?|pct|"
    r"seconds?|minutes?|hours?|days?|weeks?|months?|years?|users?|licenses?|"
    r"units?|items?|pages?|points?)\b)",
    re.I,
)
_VALUE_LIKE_TEXT = re.compile(
    rf"^\s*(?:(?:[$\u20ac\u00a3\u00a5]|US\$|{_CURRENCY_CODE})\s*)?"
    rf"[+-]?{_NUMBER_BODY}\s*(?:[KMB%]|{_CURRENCY_CODE})?\s*$",
    re.I,
)
_INSTRUCTION_SHEET = re.compile(r"\b(instruction|requirement|terms?|conditions?)\b", re.I)
_POSSIBLE_MATRIX_CONTENT = re.compile(
    r"\b(bid\s+schedule|cost\s+proposal|price\s+schedule|pricing\s+schedule|"
    r"cost\s+breakdown|pricing\s+breakdown|clin)\b",
    re.I,
)
_REQUIREMENT_TEXT = re.compile(
    r"\b(shall|must|required|requirements?|compliance|instructions?|do\s+not|"
    r"mandatory|may\s+not|(?:offerors?|vendors?|contractors?)\s+"
    r"(?:will|shall|must|is\s+responsible)|responsible\s+for|submit|provide|"
    r"complete)\b",
    re.I,
)
_SECTION_M_TEXT = re.compile(
    r"\b(evaluation\s+factors?|basis\s+for\s+award|best\s+value|trade[ -]?off|"
    r"lowest\s+price\s+technically\s+acceptable|lpta|technical\s+acceptability|"
    r"scoring|scores?|points?|adjectival|weights?|weighted?|relative\s+importance|"
    r"evaluat(?:e|ed|ion)|"
    r"assess(?:ed|ment)|award\s+will|price\s+reasonableness|technical\s+merit|"
    r"past\s+performance|more\s+important\s+than)\b",
    re.I,
)
_EVALUATION_NUMBER_TEXT = re.compile(
    r"\b(weights?|weighted?|points?|scores?|percent(?:age)?|pct)\b|%",
    re.I,
)
_COTS_TEXT = re.compile(
    r"\bCOTS\b|(?i:\b(?:commercial[\s-]+)?off[\s-]the[\s-]shelf\b)"
)
_SAFE_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL_DOC = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", _NS_MAIN)
ET.register_namespace("r", _NS_REL_DOC)


class CostMatrixError(ValueError):
    """Base error safe to surface in the Cost Matrix UI."""


class CostMatrixNotReadyError(CostMatrixError):
    def __init__(self, blockers: Iterable[str]) -> None:
        self.blockers = list(blockers)
        super().__init__("; ".join(self.blockers) or "Cost matrix is not ready.")


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _canonical_sha256(value: Any) -> str:
    payload = json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        default=str,
    ).encode("utf-8")
    return _sha256(payload)


def _safe_filename(name: str) -> str:
    base = Path(name).name
    cleaned = _SAFE_FILENAME.sub("_", base).strip("._")
    return cleaned or "cost-matrix.xlsx"


def _validate_xlsx_package(filename: str, data: bytes) -> dict[str, Any]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise CostMatrixError("Cost matrices must be standard .xlsx workbooks.")
    if not data:
        raise CostMatrixError("The uploaded workbook is empty.")
    if len(data) > MAX_WORKBOOK_BYTES:
        raise CostMatrixError(
            f"Workbook exceeds the {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB safety limit."
        )
    try:
        with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise CostMatrixError("Workbook contains too many package entries.")
            total_uncompressed = sum(max(0, info.file_size) for info in infos)
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise CostMatrixError("Workbook expands beyond the safety limit.")
            names = {info.filename for info in infos}
            if len(names) != len(infos):
                raise CostMatrixError("Workbook contains duplicate package member names.")
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise CostMatrixError("The upload is not a valid XLSX workbook package.")
            unsafe = [
                name for name in names
                if name.startswith("/") or ".." in PurePosixPath(name).parts
            ]
            if unsafe:
                raise CostMatrixError("Workbook contains unsafe package paths.")
            has_macros = any("vbaproject" in name.lower() for name in names)
            if has_macros:
                raise CostMatrixError(
                    "Macro-enabled workbooks are not supported. Provide a sanitized .xlsx copy."
                )
            if any(
                name.lower().startswith("_xmlsignatures/")
                or "digitalsignature" in name.lower()
                for name in names
            ):
                raise CostMatrixError(
                    "Digitally signed workbooks cannot be completed safely because editing "
                    "invalidates the signature. Provide an unsigned .xlsx copy."
                )
            return {
                "zip_entries": len(infos),
                "uncompressed_bytes": total_uncompressed,
                "has_macros": False,
                "has_external_links": any(
                    name.startswith("xl/externalLinks/") for name in names
                ),
                "has_drawings": any(name.startswith("xl/drawings/") for name in names),
                "has_embedded_objects": any(
                    name.startswith("xl/embeddings/") for name in names
                ),
            }
    except zipfile.BadZipFile as exc:
        raise CostMatrixError("The upload is not a readable XLSX workbook.") from exc


def _plain_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _price_context(value: str, *, header: bool = False) -> bool:
    stripped = value.strip()
    if _PRICE_CONTEXT.search(stripped):
        return True
    return bool(
        header
        and re.fullmatch(
            r"(?:grand\s+)?total|(?:unit\s+)?rate|unit",
            stripped,
            flags=re.I,
        )
    )


def _redact_financial_text(
    value: str,
    *,
    financial_context: bool = False,
) -> str:
    stripped = value.strip()
    if not stripped or _VALUE_PLACEHOLDER.fullmatch(stripped):
        return ""
    redacted = _FINANCIAL_VALUE_TOKEN.sub("[financial value redacted]", stripped)
    if financial_context:
        protected: list[tuple[str, str]] = []

        def protect_requirement_number(match: re.Match[str]) -> str:
            marker = f"\ue000{'X' * (len(protected) + 1)}\ue001"
            protected.append((marker, match.group(0)))
            return marker

        redacted = _SAFE_REQUIREMENT_NUMBER_TOKEN.sub(
            protect_requirement_number,
            redacted,
        )
        redacted = _PLAIN_FINANCIAL_NUMBER_TOKEN.sub(
            "[financial value redacted]",
            redacted,
        )
        for marker, original in protected:
            redacted = redacted.replace(marker, original)
    if redacted == "[financial value redacted]":
        return ""
    return redacted


def _value_like_text(value: str) -> bool:
    """Return whether text is a cell value rather than a column label.

    Price schedules frequently store amounts as strings. Treating the most
    recent string as a header lets one text-valued amount hide the real column
    header from every row below it, so these values must not replace header
    context while a worksheet is scanned.
    """
    stripped = value.strip()
    redacted = _redact_financial_text(stripped, financial_context=True)
    residue = redacted.replace("[financial value redacted]", "").strip()
    return bool(
        _VALUE_LIKE_TEXT.fullmatch(stripped)
        or _VALUE_PLACEHOLDER.fullmatch(stripped)
        or not residue
        or re.fullmatch(r"(?:[-\u2013\u2014/(),]|\bto\b|\s)+", residue, re.I)
        or re.fullmatch(
            r"(?:n/?a|not\s+applicable|tbd|tbc|to\s+be\s+(?:determined|provided|proposed))",
            stripped,
            flags=re.I,
        )
    )


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _safe_evaluation_scalar(value: Any) -> bool:
    """Allow bounded weight/score scalars, never arbitrary amount-sized values."""
    if isinstance(value, bool):
        return False
    try:
        numeric = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, TypeError, ValueError):
        return False
    return numeric.is_finite() and Decimal("0") <= numeric <= Decimal("100")


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _currency_format(number_format: str | None) -> bool:
    fmt = (number_format or "").lower()
    if is_date_format(fmt):
        return False
    return bool(
        any(token in fmt for token in ("$", "€", "£", "¥", "[$", "_($"))
        or re.search(_CURRENCY_CODE, fmt, flags=re.I)
    )


def _percent_format(number_format: str | None) -> bool:
    return "%" in (number_format or "")


def _target_kind(
    cell,
    *,
    metadata_semantic: str | None = None,
    header: str = "",
) -> str:
    if metadata_semantic == "company_legal_name":
        return "text"
    if metadata_semantic == "matrix_date" or is_date_format(cell.number_format or ""):
        return "date"
    if _percent_format(cell.number_format):
        return "percentage"
    if _currency_format(cell.number_format):
        return "money"
    header_lower = header.lower()
    if re.search(r"\b(percent|percentage|pct|markup)\b|%", header_lower):
        return "percentage"
    if re.search(r"\b(qty|quantity|hours?|units?|count|fte)\b", header_lower):
        return "number"
    if _price_context(header, header=True):
        return "money"
    if isinstance(cell.value, (int, float, Decimal)) and not isinstance(cell.value, bool):
        return "number"
    return "money"


def _row_hidden(ws, row: int) -> bool:
    dimension = ws.row_dimensions.get(row)
    return bool(dimension is not None and dimension.hidden)


def _column_hidden(ws, column: int) -> bool:
    direct = ws.column_dimensions.get(get_column_letter(column))
    if direct is not None and direct.hidden:
        return True
    # Grouped columns may be represented by one dimension spanning min..max.
    for dimension in ws.column_dimensions.values():
        if not dimension.hidden:
            continue
        minimum = getattr(dimension, "min", None)
        maximum = getattr(dimension, "max", None)
        if minimum is not None and maximum is not None:
            if int(minimum) <= column <= int(maximum):
                return True
    return False


def _cell_hidden(ws, row: int, column: int) -> bool:
    return _row_hidden(ws, row) or _column_hidden(ws, column)


def _nearest_text_left(ws, row: int, column: int, *, distance: int = 24) -> str:
    if _row_hidden(ws, row):
        return ""
    candidates: list[str] = []
    for col in range(column - 1, max(0, column - distance - 1), -1):
        if _column_hidden(ws, col):
            continue
        value = ws.cell(row=row, column=col).value
        if isinstance(value, str) and value.strip() and not _is_formula(value):
            candidates.append(value.strip())
    unit_or_id = re.compile(
        r"^(?:ea|each|lot|mo|mos|month|months|yr|yrs|year|years|hr|hrs|"
        r"hour|hours|unit|units|fte|clin\s*)$|^(?:clin\s*)?[a-z]*\d+[a-z]*$",
        re.I,
    )
    for candidate in candidates:
        if not unit_or_id.fullmatch(candidate) and (
            len(candidate) >= 5 or " " in candidate
        ):
            return candidate
    if candidates:
        return candidates[0]
    for col in range(column + 1, min(ws.max_column, column + 3) + 1):
        if _column_hidden(ws, col):
            continue
        value = ws.cell(row=row, column=col).value
        if isinstance(value, str) and value.strip() and not _is_formula(value):
            return value.strip()
    return ""


def _metadata_candidate_is_writable(
    *,
    semantic: str,
    label: str,
    value: Any,
) -> bool:
    """Reject neighboring labels while allowing a blank or same-label placeholder."""
    if value is None or value == 0 or isinstance(value, (date, datetime)):
        return True
    if not isinstance(value, str):
        return True
    stripped = value.strip()
    if not stripped or _VALUE_PLACEHOLDER.fullmatch(stripped):
        return True
    if semantic == "matrix_date":
        try:
            date.fromisoformat(stripped)
            return True
        except ValueError:
            return False
    def normalize(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    # Many buyer templates put "Vendor Name" in the destination cell as a
    # placeholder. A different label (Company Address, UEI, etc.) is not a
    # writable destination and must never be overwritten.
    return normalize(stripped) == normalize(label)


def _nearest_header_above(ws, row: int, column: int, *, distance: int = 12) -> str:
    if _column_hidden(ws, column):
        return ""
    for prior_row in range(row - 1, max(0, row - distance - 1), -1):
        if _row_hidden(ws, prior_row):
            continue
        value = ws.cell(row=prior_row, column=column).value
        if isinstance(value, str) and value.strip() and not _is_formula(value):
            return value.strip()
    return ""


def _merged_top_left(ws, coordinate: str) -> tuple[str, str | None]:
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            top_left = ws.cell(merged.min_row, merged.min_col).coordinate
            return top_left, str(merged)
    return coordinate, None


def _candidate_target(
    *,
    target_id: str,
    sheet: str,
    cell,
    label: str,
    header: str,
    category: str,
    semantic: str | None = None,
    confidence: str,
    evidence: list[str],
    merged_range: str | None = None,
    origin: str = "detected",
) -> dict[str, Any]:
    return {
        "id": target_id,
        "sheet": sheet,
        "cell": cell.coordinate,
        "merged_range": merged_range,
        "category": category,
        "semantic": semantic,
        "kind": _target_kind(
            cell,
            metadata_semantic=semantic,
            header=header,
        ),
        "label": label or cell.coordinate,
        "header": header,
        "existing_value": _json_safe_value(cell.value),
        "number_format": cell.number_format or "General",
        "confidence": confidence,
        "evidence": evidence,
        "origin": origin,
        "formula_owned": False,
    }


_CELL_REF = re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)$", re.I)
_RANGE_REF = re.compile(
    r"^\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)$",
    re.I,
)


def _expand_formula_reference(token: str) -> list[str] | None:
    token = token.strip()
    cell_match = _CELL_REF.fullmatch(token)
    if cell_match:
        return [f"{cell_match.group(1).upper()}{cell_match.group(2)}"]
    range_match = _RANGE_REF.fullmatch(token)
    if not range_match:
        return None
    cell_range = (
        f"{range_match.group(1)}{range_match.group(2)}:"
        f"{range_match.group(3)}{range_match.group(4)}"
    )
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return [
        f"{get_column_letter(column_number)}{row_number}"
        for row_number in range(min_row, max_row + 1)
        for column_number in range(min_col, max_col + 1)
    ]


def _simple_formula_references(formula: str) -> list[str] | None:
    """Resolve only auditable SUM/ref-addition formulas; reject other math."""
    expression = str(formula or "").strip()
    if expression.startswith("="):
        expression = expression[1:].strip()
    sum_match = re.fullmatch(r"SUM\((.*)\)", expression, flags=re.I)
    if sum_match:
        tokens = [part.strip() for part in sum_match.group(1).split(",")]
    elif "+" in expression:
        tokens = [part.strip() for part in expression.split("+")]
    else:
        return None
    if not tokens or any(not token for token in tokens):
        return None
    coordinates: list[str] = []
    for token in tokens:
        expanded = _expand_formula_reference(token)
        if expanded is None:
            return None
        coordinates.extend(expanded)
    # Preserve formula order while removing duplicates.
    return list(dict.fromkeys(coordinates))


def _is_reconciliation_total_label(value: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()
    tokens = normalized.split()
    return bool(
        "total" in tokens
        and "subtotal" not in tokens
        and not any(
            first == "sub" and second == "total"
            for first, second in zip(tokens, tokens[1:], strict=False)
        )
    )


def _build_reconciliation_manifest(
    targets: list[dict[str, Any]],
    formulas: list[dict[str, Any]],
    *,
    visible_sheets: set[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    pricing_targets = [
        target for target in targets if target.get("category") == "pricing"
    ]
    coordinate_targets = {
        (str(target["sheet"]), str(target["cell"]).upper()): target
        for target in pricing_targets
    }
    candidates: list[dict[str, Any]] = []
    for formula in formulas:
        if str(formula.get("sheet") or "") not in visible_sheets:
            continue
        if formula.get("visible") is False:
            continue
        if not _is_reconciliation_total_label(
            str(formula.get("label") or "")
        ):
            continue
        coordinates = _simple_formula_references(str(formula.get("formula") or ""))
        formula_supported = bool(coordinates)
        coordinates = coordinates or []
        member_ids = [
            str(coordinate_targets[(str(formula["sheet"]), coordinate)]["id"])
            for coordinate in coordinates
            if (str(formula["sheet"]), coordinate) in coordinate_targets
        ]
        member_targets = [
            coordinate_targets[(str(formula["sheet"]), coordinate)]
            for coordinate in coordinates
            if (str(formula["sheet"]), coordinate) in coordinate_targets
        ]
        # A total of hours, quantities, percentages, or other non-money units
        # is useful workbook logic but must never be reconciled to a dollar
        # proposal total.
        if member_targets and not any(
            target.get("kind") == "money" for target in member_targets
        ):
            continue
        candidates.append({
            "sheet": formula["sheet"],
            "formula_cell": formula["cell"],
            "formula": formula["formula"],
            "formula_supported": formula_supported,
            "member_coordinates": coordinates,
            "member_target_ids": list(dict.fromkeys(member_ids)),
        })

    all_money_ids = {
        str(target["id"])
        for target in pricing_targets
        if target.get("kind") == "money"
    }
    automatic: list[dict[str, Any]] = []
    if len(candidates) == 1:
        candidate_ids = set(candidates[0]["member_target_ids"])
        all_formula_inputs_are_targets = (
            len(candidate_ids) == len(candidates[0]["member_coordinates"])
        )
        if (
            candidates[0]["formula_supported"]
            and len(candidate_ids) >= 2
            and candidate_ids == all_money_ids
            and all_formula_inputs_are_targets
        ):
            automatic = [{
                "type": "sum_to_proposed_price",
                **candidates[0],
                "tolerance_usd": 0.01,
            }]

    needs_review = bool(candidates and not automatic)
    review = {
        "review_required": needs_review,
        "reason": (
            "multiple or partial buyer total formulas require an operator to define "
            "how the workbook reconciles to the proposal total"
            if needs_review
            else None
        ),
        "candidates": candidates,
    }
    return automatic, review


def inspect_cost_matrix(filename: str, data: bytes) -> dict[str, Any]:
    """Return a deterministic, template-specific workbook manifest.

    ``classification.is_cost_matrix`` is conservative. A detected workbook may
    still have zero proposed targets; that is a supported ``mapping_required``
    state and the UI lets the operator add exact cells manually.
    """
    package = _validate_xlsx_package(filename, data)
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise CostMatrixError(f"Workbook could not be inspected: {exc}") from exc

    total_cells = sum(ws.max_row * ws.max_column for ws in workbook.worksheets)
    if total_cells > MAX_INSPECTED_CELLS:
        workbook.close()
        raise CostMatrixError(
            "Workbook used ranges are too large for safe automatic inspection."
        )

    targets: list[dict[str, Any]] = []
    target_coordinates: set[tuple[str, str]] = set()
    formulas: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    visible_financial_terms = 0
    visible_strong_price_terms = 0
    visible_currency_cells = 0
    visible_value_placeholders = 0
    visible_possible_matrix_cues = 0
    formula_total_count = 0

    def add_target(target: dict[str, Any]) -> None:
        key = (target["sheet"], target["cell"])
        if key in target_coordinates:
            return
        target["id"] = f"target-{len(targets) + 1}"
        target_coordinates.add(key)
        targets.append(target)

    for ws in workbook.worksheets:
        sheet_formula_count = 0
        sheet_currency_count = 0
        sheet_visible_currency_count = 0
        sheet_comment_count = 0
        sheet_hyperlink_count = 0
        sheet_text: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                visibly_exposed = bool(
                    ws.sheet_state == "visible"
                    and not _cell_hidden(ws, cell.row, cell.column)
                )
                if visibly_exposed and isinstance(value, str) and value.strip():
                    sheet_text.append(value.strip())
                if _is_formula(value):
                    sheet_formula_count += 1
                    label = _nearest_text_left(ws, cell.row, cell.column)
                    formulas.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "formula": value,
                        "label": label,
                        "visible": visibly_exposed,
                    })
                    if visibly_exposed and _is_reconciliation_total_label(label or ""):
                        formula_total_count += 1
                if _currency_format(cell.number_format):
                    sheet_currency_count += 1
                    if visibly_exposed:
                        sheet_visible_currency_count += 1
                if cell.comment is not None:
                    sheet_comment_count += 1
                if cell.hyperlink is not None:
                    sheet_hyperlink_count += 1

        sheet_blob = "\n".join(sheet_text)
        term_count = len(_FINANCIAL_WORDS.findall(sheet_blob))
        strong_price_term_count = len(_STRONG_PRICE_WORDS.findall(sheet_blob))
        placeholder_count = sum(
            1 for value in sheet_text if _VALUE_PLACEHOLDER.fullmatch(value)
        )
        sheet_financial_context = bool(
            _FINANCIAL_WORDS.search(ws.title)
            or _POSSIBLE_MATRIX_FILENAME.search(ws.title)
            or _POSSIBLE_MATRIX_CONTENT.search(sheet_blob)
            or term_count >= 2
        )
        if ws.sheet_state == "visible":
            visible_financial_terms += term_count
            visible_strong_price_terms += strong_price_term_count
            visible_currency_cells += sheet_visible_currency_count
            visible_value_placeholders += placeholder_count
            if _POSSIBLE_MATRIX_CONTENT.search(sheet_blob):
                visible_possible_matrix_cues += 1

        sheets.append({
            "name": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimension": ws.calculate_dimension(),
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "formula_count": sheet_formula_count,
            "currency_formatted_cells": sheet_currency_count,
            "data_validation_count": len(
                getattr(ws.data_validations, "dataValidation", []) or []
            ),
            "tables": sorted(str(name) for name in ws.tables.keys()),
            "chart_count": len(getattr(ws, "_charts", []) or []),
            "image_count": len(getattr(ws, "_images", []) or []),
            "comment_count": sheet_comment_count,
            "hyperlink_count": sheet_hyperlink_count,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "protected": bool(ws.protection.sheet),
            "hidden_row_count": sum(
                1 for dimension in ws.row_dimensions.values() if dimension.hidden
            ),
            "hidden_column_count": sum(
                1 for dimension in ws.column_dimensions.values() if dimension.hidden
            ),
        })

        # Hidden sheets are inventoried and preserved but never proposed as
        # write targets automatically.
        if ws.sheet_state != "visible":
            continue

        # Metadata fields are label-driven and independent from the financial
        # grid. Date remains semantically ambiguous until mapped by a person.
        for row in ws.iter_rows():
            for label_cell in row:
                if (
                    isinstance(label_cell, MergedCell)
                    or _cell_hidden(ws, label_cell.row, label_cell.column)
                ):
                    continue
                label_value = _plain_cell_value(label_cell.value)
                if not label_value or _is_formula(label_value):
                    continue
                for semantic, pattern in _METADATA_LABELS:
                    if not pattern.search(label_value):
                        continue
                    for offset in range(1, 5):
                        col = label_cell.column + offset
                        if col > ws.max_column:
                            break
                        possible = ws.cell(label_cell.row, col)
                        top_left, merged_range = _merged_top_left(ws, possible.coordinate)
                        possible = ws[top_left]
                        if possible.coordinate == label_cell.coordinate:
                            continue
                        if _cell_hidden(ws, possible.row, possible.column):
                            continue
                        if _is_formula(possible.value):
                            continue
                        if ws.protection.sheet and possible.protection.locked:
                            continue
                        if not _metadata_candidate_is_writable(
                            semantic=semantic,
                            label=label_value,
                            value=possible.value,
                        ):
                            continue
                        add_target(_candidate_target(
                            target_id="",
                            sheet=ws.title,
                            cell=possible,
                            label=label_value,
                            header="",
                            category="metadata",
                            semantic=semantic,
                            confidence="high",
                            evidence=["adjacent to recognized metadata label"],
                            merged_range=merged_range,
                        ))
                        break
                    break

        # Financial inputs are inferred from a combination of local header,
        # row label, value type, and number format. No one cue is authoritative.
        for row in ws.iter_rows():
            for cell in row:
                if (
                    isinstance(cell, MergedCell)
                    or _is_formula(cell.value)
                    or _cell_hidden(ws, cell.row, cell.column)
                ):
                    continue
                key = (ws.title, cell.coordinate)
                if key in target_coordinates:
                    continue
                if ws.protection.sheet and cell.protection.locked:
                    continue
                row_label = _nearest_text_left(ws, cell.row, cell.column)
                header = _nearest_header_above(ws, cell.row, cell.column)
                has_financial_header = bool(_FINANCIAL_HEADER.search(header or ""))
                has_currency = _currency_format(cell.number_format)
                has_percent = _percent_format(cell.number_format)
                placeholder = bool(
                    isinstance(cell.value, str)
                    and _VALUE_PLACEHOLDER.fullmatch(cell.value.strip())
                )
                numeric_or_blank = cell.value is None or placeholder or (
                    isinstance(cell.value, (int, float, Decimal))
                    and not isinstance(cell.value, bool)
                    and Decimal(str(cell.value)) == 0
                )
                if not numeric_or_blank or not row_label:
                    continue
                # Avoid guessing at ordinary blank grid cells. A financial
                # format is strong evidence; otherwise require an explicit
                # financial column header and a numeric/blank input cell.
                if not (
                    has_currency
                    or has_percent
                    or has_financial_header
                    or (sheet_financial_context and placeholder)
                    or (
                        sheet_financial_context
                        and isinstance(cell.value, (int, float, Decimal))
                        and Decimal(str(cell.value)) == 0
                    )
                ):
                    continue
                evidence: list[str] = []
                if has_currency:
                    evidence.append("currency/accounting number format")
                if has_percent:
                    evidence.append("percentage number format")
                if has_financial_header:
                    evidence.append(f"financial column header: {header}")
                if placeholder:
                    evidence.append("buyer value placeholder")
                if sheet_financial_context and not evidence:
                    evidence.append("financial worksheet context")
                add_target(_candidate_target(
                    target_id="",
                    sheet=ws.title,
                    cell=cell,
                    label=row_label,
                    header=header,
                    category="pricing",
                    confidence="high" if len(evidence) >= 2 else "medium",
                    evidence=evidence,
                ))

    defined_names: list[dict[str, Any]] = []
    try:
        for named in workbook.defined_names.values():
            defined_names.append({
                "name": str(getattr(named, "name", "")),
                "local_sheet_id": getattr(named, "localSheetId", None),
                "hidden": bool(getattr(named, "hidden", False)),
                "reference": str(getattr(named, "attr_text", "") or ""),
            })
    except Exception:
        log.warning("could not inventory workbook defined names", exc_info=True)
    workbook_properties = {
        "creator_present": bool(getattr(workbook.properties, "creator", None)),
        "last_modified_by_present": bool(
            getattr(workbook.properties, "lastModifiedBy", None)
        ),
        "workbook_protected": bool(
            getattr(workbook.security, "lockStructure", False)
            or getattr(workbook.security, "lockWindows", False)
        ),
        "defined_names": defined_names,
    }
    workbook.close()

    pricing_targets = [t for t in targets if t["category"] == "pricing"]
    reconciliations, reconciliation_review = _build_reconciliation_manifest(
        targets,
        formulas,
        visible_sheets={
            str(sheet["name"])
            for sheet in sheets
            if sheet.get("state") == "visible"
        },
    )
    filename_signal = bool(_MATRIX_FILENAME.search(Path(filename).stem))
    possible_filename_signal = bool(
        filename_signal or _POSSIBLE_MATRIX_FILENAME.search(Path(filename).stem)
    )
    score = 0
    evidence: list[str] = []
    if filename_signal:
        score += 4
        evidence.append("filename identifies a cost/pricing matrix or schedule")
    if pricing_targets:
        score += min(4, 1 + len(pricing_targets))
        evidence.append(f"{len(pricing_targets)} candidate financial input cell(s)")
    if visible_financial_terms >= 2:
        score += 2
        evidence.append("visible workbook labels contain financial terminology")
    if formula_total_count:
        score += 1
        evidence.append("total/subtotal formula detected")
    if visible_value_placeholders:
        score += 1
        evidence.append(
            f"{visible_value_placeholders} visible buyer value placeholder(s)"
        )
    if visible_possible_matrix_cues:
        score += 1
        evidence.append("visible workbook title/labels indicate a bid or CLIN schedule")
    # A filename alone is awareness evidence, never enough to remove the
    # workbook from ordinary requirements intake. Automatic confirmation
    # requires at least one concrete writable financial target.
    strong_price_evidence = bool(
        filename_signal
        or visible_strong_price_terms
        or visible_currency_cells
    )
    is_matrix = bool(
        pricing_targets
        and strong_price_evidence
        and (
            (filename_signal and score >= 5)
            or (len(pricing_targets) >= 2 and score >= 5)
        )
    )
    possible_matrix = bool(
        not is_matrix
        and (
            possible_filename_signal
            or bool(pricing_targets)
            or visible_financial_terms >= 2
            or visible_value_placeholders > 0
            or visible_possible_matrix_cues > 0
            or formula_total_count > 0
        )
    )
    confidence = "high" if score >= 7 else "medium" if score >= 4 else "low"

    warnings: list[str] = []
    hidden = [s["name"] for s in sheets if s["state"] != "visible"]
    if hidden:
        warnings.append(
            f"Preserve and review {len(hidden)} hidden sheet(s) before submission."
        )
    hidden_axes = sum(
        int(sheet.get("hidden_row_count") or 0)
        + int(sheet.get("hidden_column_count") or 0)
        for sheet in sheets
    )
    if hidden_axes:
        warnings.append(
            "Hidden rows and columns are preserved but excluded from automatic "
            "mapping and agent intake."
        )
    if package["has_external_links"]:
        warnings.append("Workbook contains external links; verify them before submission.")
    if package["has_embedded_objects"]:
        warnings.append("Workbook contains embedded objects that require human review.")
    if any(s["protected"] for s in sheets):
        warnings.append("One or more worksheets are protected; mapped cells may be locked.")
    if workbook_properties["workbook_protected"]:
        warnings.append("The workbook structure is protected and requires human review.")
    if any(s["data_validation_count"] for s in sheets):
        warnings.append("Workbook data-validation rules are retained in generated copies.")
    if formulas:
        warnings.append(
            "Formula cells will not be overwritten; Excel-compatible software must recalculate them."
        )
    warnings.append(
        "Workbook metadata and any hidden content are retained in generated copies; review before submission."
    )

    return {
        "analysis_version": ANALYSIS_VERSION,
        "template_sha256": _sha256(data),
        "filename": filename,
        "classification": {
            "is_cost_matrix": is_matrix,
            "possible_cost_matrix": possible_matrix,
            "decision": (
                "confirmed" if is_matrix else "needs_confirmation" if possible_matrix else "not_matrix"
            ),
            "confidence": confidence,
            "score": score,
            "evidence": evidence,
        },
        "package": package,
        "workbook": workbook_properties,
        "sheets": sheets,
        "targets": targets,
        "formulas": formulas,
        "reconciliations": reconciliations,
        "reconciliation_review": reconciliation_review,
        "warnings": warnings,
    }


def try_inspect_cost_matrix(filename: str, data: bytes) -> dict[str, Any] | None:
    """Best-effort classifier used during general package intake."""
    if Path(filename).suffix.lower() != ".xlsx":
        return None
    try:
        manifest = inspect_cost_matrix(filename, data)
    except CostMatrixError:
        log.warning("could not inspect possible cost matrix %s", filename, exc_info=True)
        return None
    classification = manifest.get("classification") or {}
    return manifest if (
        classification.get("is_cost_matrix")
        or classification.get("possible_cost_matrix")
    ) else None


def _source(
    key: str,
    label: str,
    value: Any,
    kind: str,
    *,
    group: str,
) -> dict[str, Any]:
    return {
        "key": key,
        "label": label,
        "value": value,
        "kind": kind,
        "group": group,
    }


def _base_source_catalog(proposal: Proposal) -> list[dict[str, Any]]:
    company = get_company_profile().get("company") or {}
    return [
        _source(
            "company.legal_name",
            "Company — legal name",
            company.get("legal_name"),
            "text",
            group="Proposal metadata",
        ),
        _source(
            "proposal.title",
            "Proposal — title",
            proposal.title,
            "text",
            group="Proposal metadata",
        ),
        _source(
            "proposal.agency",
            "Proposal — agency",
            proposal.agency,
            "text",
            group="Proposal metadata",
        ),
        _source(
            "proposal.due_date",
            "Proposal — due date",
            proposal.due_date.isoformat() if proposal.due_date else None,
            "date",
            group="Proposal metadata",
        ),
        _source(
            "system.generation_date",
            "Workbook generation date",
            date.today().isoformat(),
            "date",
            group="Proposal metadata",
        ),
    ]


def _number_kind(key: str) -> str:
    lowered = key.lower()
    if any(token in lowered for token in ("pct", "percent", "margin", "rate_pct")):
        return "percentage"
    if any(token in lowered for token in ("usd", "cost", "price", "fee", "revenue", "profit", "amount")):
        return "money"
    return "number"


def _append_numeric_json_sources(
    catalog: list[dict[str, Any]],
    value: Any,
    *,
    key_prefix: str,
    label_prefix: str,
    group: str,
) -> None:
    if isinstance(value, dict):
        for key in sorted(value):
            child = value[key]
            friendly = str(key).replace("_", " ").strip().title()
            _append_numeric_json_sources(
                catalog,
                child,
                key_prefix=f"{key_prefix}.{key}",
                label_prefix=f"{label_prefix} — {friendly}" if label_prefix else friendly,
                group=group,
            )
    elif isinstance(value, list):
        for index, child in enumerate(value):
            _append_numeric_json_sources(
                catalog,
                child,
                key_prefix=f"{key_prefix}.{index}",
                label_prefix=f"{label_prefix} #{index + 1}",
                group=group,
            )
    elif isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        kind = _number_kind(key_prefix)
        # Pricing models store percentages as Excel-style decimals (0.24 =
        # 24%). A larger pct-suffixed value has an ambiguous convention, so
        # expose it as a plain number and require an explicit manual conversion.
        if kind == "percentage" and abs(float(value)) > 1:
            kind = "number"
        catalog.append(_source(
            key_prefix,
            label_prefix,
            float(value),
            kind,
            group=group,
        ))


def _it_source_catalog(
    db: Session,
    proposal: Proposal,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    catalog = _base_source_catalog(proposal)
    scenario = (proposal.proposed_scenario or "").upper().strip()
    package = None
    if scenario:
        package = db.execute(
            select(PricingPackage)
            .where(
                PricingPackage.proposal_id == proposal.id,
                PricingPackage.scenario == scenario,
            )
            .options(selectinload(PricingPackage.lines))
        ).scalar_one_or_none()
    if package is None:
        return catalog, {
            "service_line": "it_services",
            "scenario": scenario or None,
            "pricing_exists": False,
        }

    aggregate = [
        ("pricing.total_proposed_price", "Selected scenario — total proposed price", package.total_proposed_price, "money"),
        ("pricing.loaded_labor_cost", "Selected scenario — loaded labor cost", package.loaded_labor_cost, "money"),
        ("pricing.subcontractor_costs", "Selected scenario — subcontractor cost", package.subcontractor_costs, "money"),
    ]
    odcs = list(package.odcs_json or [])
    def odc_extended_amount(row: dict[str, Any]) -> float:
        extended = row.get("extended_amount_usd")
        if isinstance(extended, (int, float, Decimal)) and not isinstance(extended, bool):
            return float(extended)
        amount = row.get("amount_usd", row.get("amount"))
        if not isinstance(amount, (int, float, Decimal)) or isinstance(amount, bool):
            return 0.0
        years = row.get("year_count", 1)
        if not isinstance(years, (int, float, Decimal)) or isinstance(years, bool):
            years = 1
        return float(amount) * float(years)

    odc_total = sum(
        odc_extended_amount(row)
        for row in odcs
        if isinstance(row, dict)
    )
    aggregate.append(("pricing.odcs_total", "Selected scenario — total other direct costs", odc_total, "money"))
    for key, label, value, kind in aggregate:
        catalog.append(_source(key, label, float(value) if value is not None else None, kind, group=f"{scenario} scenario totals"))

    _append_numeric_json_sources(
        catalog,
        dict(package.indirect_costs_json or {}),
        key_prefix="pricing.indirect",
        label_prefix="Selected scenario — indirect costs",
        group=f"{scenario} scenario totals",
    )
    _append_numeric_json_sources(
        catalog,
        dict(package.pnl_projection_json or {}),
        key_prefix="pricing.pnl",
        label_prefix="Selected scenario — P&L",
        group=f"{scenario} scenario totals",
    )

    phases = list(package.phase_breakdown_json or [])
    for index, phase in enumerate(phases):
        if not isinstance(phase, dict):
            continue
        name = str(phase.get("name") or f"Phase {index + 1}").strip()
        for field, field_label in (
            ("phase_price_usd", "price"),
            ("phase_subtotal_cost_usd", "subtotal cost"),
            ("phase_loaded_cost_usd", "loaded labor cost"),
            ("phase_ga_usd", "G&A"),
            ("phase_contingency_cost_usd", "contingency cost"),
            ("phase_profit_usd", "profit"),
        ):
            value = phase.get(field)
            if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                catalog.append(_source(
                    f"pricing.phase.{index}.{field}",
                    f"Phase: {name} — {field_label}",
                    float(value),
                    "money",
                    group="Lifecycle phases",
                ))
        for field, field_label in (
            ("phase_total_hours", "total hours"),
            ("phase_contingency_hours", "contingency hours"),
            ("start_month", "start month"),
            ("duration_months", "duration in months"),
        ):
            value = phase.get(field)
            if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                catalog.append(_source(
                    f"pricing.phase.{index}.{field}",
                    f"Phase: {name} â€” {field_label}",
                    float(value),
                    "number",
                    group="Lifecycle phases",
                ))

    for line in sorted(package.lines, key=lambda item: item.id):
        name = line.labor_category
        fields = (
            ("hours", "hours", "number"),
            ("loaded_hourly_rate_usd", "loaded hourly rate", "money"),
            ("loaded_cost_usd", "loaded cost", "money"),
            ("proposed_billing_rate_usd", "proposed billing rate", "money"),
            ("billed_total_usd", "billed total", "money"),
        )
        for field, field_label, kind in fields:
            value = getattr(line, field)
            catalog.append(_source(
                f"pricing.labor.{line.id}.{field}",
                f"Labor: {name} — {field_label}",
                float(value) if value is not None else None,
                kind,
                group="Labor categories",
            ))

    for index, odc in enumerate(odcs):
        if not isinstance(odc, dict):
            continue
        item = str(odc.get("item") or f"ODC {index + 1}").strip()
        value = odc_extended_amount(odc)
        if value or any(
            key in odc for key in ("amount", "amount_usd", "extended_amount_usd")
        ):
            catalog.append(_source(
                f"pricing.odc.{index}.extended_amount_usd",
                f"Other direct cost: {item} â€” extended amount",
                float(value),
                "money",
                group="Other direct costs",
            ))
        amount = odc.get("amount_usd", odc.get("amount"))
        if isinstance(amount, (int, float, Decimal)) and not isinstance(amount, bool):
            catalog.append(_source(
                f"pricing.odc.{index}.amount_usd",
                f"Other direct cost: {item} â€” unit/annual amount",
                float(amount),
                "money",
                group="Other direct costs",
            ))
        years = odc.get("year_count")
        if isinstance(years, (int, float, Decimal)) and not isinstance(years, bool):
            catalog.append(_source(
                f"pricing.odc.{index}.year_count",
                f"Other direct cost: {item} â€” year count",
                float(years),
                "number",
                group="Other direct costs",
            ))

    review = get_cost_review_freshness(db, proposal.id, scenario=scenario)
    return catalog, {
        "service_line": "it_services",
        "scenario": scenario,
        "pricing_exists": package.total_proposed_price is not None,
        "pricing_sha256": _canonical_sha256({
            "scenario": scenario,
            "loaded_labor_cost": package.loaded_labor_cost,
            "odcs": package.odcs_json or [],
            "subcontractor_costs": package.subcontractor_costs,
            "indirect_costs": package.indirect_costs_json or {},
            "total_proposed_price": package.total_proposed_price,
            "pnl": package.pnl_projection_json or {},
            "phases": package.phase_breakdown_json or [],
            "lines": [
                {
                    "id": line.id,
                    "labor_category": line.labor_category,
                    "hours": line.hours,
                    "loaded_rate": line.loaded_hourly_rate_usd,
                    "loaded_cost": line.loaded_cost_usd,
                    "billing_rate": line.proposed_billing_rate_usd,
                    "billed_total": line.billed_total_usd,
                }
                for line in sorted(package.lines, key=lambda row: row.id)
            ],
        }),
        "cost_review": review,
    }


def _payment_source_catalog(
    db: Session,
    proposal: Proposal,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    catalog = _base_source_catalog(proposal)
    scan: dict[str, Any] = {}
    review: dict[str, Any] = {}
    try:
        decoded = json.loads(proposal.payment_market_scan_json or "{}")
        if isinstance(decoded, dict):
            scan = decoded
    except (TypeError, json.JSONDecodeError):
        pass
    try:
        decoded = json.loads(proposal.payment_cost_review_findings_json or "{}")
        if isinstance(decoded, dict):
            review = decoded
    except (TypeError, json.JSONDecodeError):
        pass
    # Expose numeric leaves from the approved scan without assuming which fee
    # schedule shape a buyer workbook uses.
    _append_numeric_json_sources(
        catalog,
        scan,
        key_prefix="payment.scan",
        label_prefix="Payment pricing",
        group="Payment pricing and profit math",
    )
    scan_current = bool(scan) and payment_market_scan_is_current(scan)
    review_current = bool(review) and payment_cost_review_is_current(
        proposal.id,
        review,
        db=db,
    )
    return catalog, {
        "service_line": "payment_systems",
        "scenario": proposal.selected_pricing_model,
        "pricing_exists": bool(scan),
        "pricing_sha256": _canonical_sha256(scan) if scan else None,
        "payment_scan_current": scan_current,
        "payment_review_current": review_current,
    }


def _source_catalog(
    db: Session,
    proposal: Proposal,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if (proposal.service_line or "it_services").strip() == "payment_systems":
        return _payment_source_catalog(db, proposal)
    return _it_source_catalog(db, proposal)


def get_cost_source_catalog(proposal_id: int) -> dict[str, Any]:
    with session_scope() as db:
        proposal = db.get(Proposal, proposal_id)
        if proposal is None:
            raise CostMatrixError(f"Proposal {proposal_id} was not found.")
        sources, basis = _source_catalog(db, proposal)
        return {"sources": sources, "basis": basis}


def _normalized_label(value: str) -> str:
    value = re.sub(r"\b(i+|[a-z])-?\d+(?:\.[a-z0-9]+)*\b", " ", value.lower())
    value = re.sub(r"[^a-z0-9]+", " ", value)
    stop = {"the", "and", "of", "for", "to", "total", "cost", "price", "pricing"}
    return " ".join(word for word in value.split() if word not in stop)


def _mapping_semantics(value: str) -> set[str]:
    lowered = value.lower()
    semantics: set[str] = set()
    patterns = {
        "hours": r"\b(hours?|labor\s+hours?|fte)\b",
        "quantity": r"\b(qty|quantity|units?|count)\b",
        "rate": r"\b(hourly|rate|unit\s+price)\b",
        "total": r"\b(total|extended|extension|amount|cost)\b",
        "price": r"\b(price|pricing)\b",
        "fee": r"\bfees?\b",
        "percentage": r"\b(percent|percentage|pct|markup|margin)\b|%",
    }
    for name, pattern in patterns.items():
        if re.search(pattern, lowered):
            semantics.add(name)
    return semantics


def _mapping_suggestion(
    target: dict[str, Any],
    sources: list[dict[str, Any]],
) -> dict[str, Any] | None:
    semantic = target.get("semantic")
    if semantic == "company_legal_name":
        return {"mode": "source", "source_key": "company.legal_name", "confidence": "high"}
    # A generic "Date" label is intentionally never auto-filled: the RFP may
    # mean preparation date, bid date, or submission date.
    if semantic == "matrix_date":
        return None
    raw_label = str(target.get("label") or "")
    raw_header = str(target.get("header") or "")
    label = _normalized_label(raw_label)
    if not label:
        return None
    target_kind = str(target.get("kind") or "")
    target_semantics = _mapping_semantics(f"{raw_label} {raw_header}")
    ranked: list[tuple[float, dict[str, Any]]] = []
    for source in sources:
        if source.get("value") is None:
            continue
        # Units are deliberately strict. A value can be converted only by a
        # human entering an explicit manual value; the system never treats
        # hours, dollars, percentages, or basis points as interchangeable.
        if str(source.get("kind") or "") != target_kind:
            continue
        raw_source_label = str(source.get("label") or "")
        source_semantics = _mapping_semantics(raw_source_label)
        if target_semantics and source_semantics and not (
            target_semantics & source_semantics
        ):
            continue
        source_label = _normalized_label(raw_source_label)
        if not source_label:
            continue
        score = SequenceMatcher(None, label, source_label).ratio()
        label_tokens = set(label.split())
        source_tokens = set(source_label.split())
        if label_tokens and label_tokens.issubset(source_tokens):
            score = max(score, 0.9)
        if target_semantics & source_semantics:
            score = min(1.0, score + 0.04)
        ranked.append((score, source))
    ranked.sort(key=lambda item: item[0], reverse=True)
    best = ranked[0] if ranked else None
    # Equal-looking money fields (loaded cost vs billed total, for example)
    # require a person to choose; do not preselect one based on list order.
    if best and len(ranked) > 1 and best[0] - ranked[1][0] < 0.06:
        return None
    if best and best[0] >= 0.88:
        return {
            "mode": "source",
            "source_key": best[1]["key"],
            "confidence": "high" if best[0] >= 0.95 else "medium",
        }
    return None


def _register_artifact(
    db: Session,
    *,
    proposal: Proposal,
    document: RfpPackageDocument,
    content: bytes,
    analysis: dict[str, Any],
    status: str = STATUS_MAPPING_REQUIRED,
    registration_source: str = "original_auto",
) -> CostMatrixArtifact:
    existing = db.scalar(
        select(CostMatrixArtifact).where(
            CostMatrixArtifact.source_document_id == document.id
        )
    )
    if existing is not None:
        return existing
    analysis = deepcopy(analysis)
    analysis["registration"] = {
        "source": registration_source,
        "prior_document_role": document.document_role,
        "prior_document_type": getattr(
            document.document_type,
            "value",
            str(document.document_type or RfpDocumentType.UNKNOWN.value),
        ),
    }
    confirmed = status != STATUS_NEEDS_CONFIRMATION
    if confirmed:
        document.document_type = RfpDocumentType.FORM_TEMPLATE
        document.document_role = COST_MATRIX_ROLE
    structure = dict(document.structure_json or {})
    structure.update({
        "content_sha256": _sha256(content),
        "cost_matrix": confirmed,
        "possible_cost_matrix": not confirmed,
        "cost_matrix_analysis_version": ANALYSIS_VERSION,
        "cost_matrix_confidence": analysis["classification"]["confidence"],
    })
    document.structure_json = structure
    artifact = CostMatrixArtifact(
        proposal_id=proposal.id,
        source_document_id=document.id,
        status=status,
        template_sha256=_sha256(content),
        analysis_version=ANALYSIS_VERSION,
        analysis_json=analysis,
        mapping_json={},
    )
    db.add(artifact)
    db.flush()
    return artifact


def register_original_cost_matrices(
    db: Session,
    *,
    proposal: Proposal,
    documents_and_content: Iterable[tuple[RfpPackageDocument, bytes]],
) -> list[CostMatrixArtifact]:
    """Register matrices discovered in the original package transaction."""
    artifacts: list[CostMatrixArtifact] = []
    for document, content in documents_and_content:
        analysis = try_inspect_cost_matrix(document.filename, content)
        if analysis is None:
            continue
        status = (
            STATUS_MAPPING_REQUIRED
            if analysis["classification"].get("is_cost_matrix")
            else STATUS_NEEDS_CONFIRMATION
        )
        artifacts.append(_register_artifact(
            db,
            proposal=proposal,
            document=document,
            content=content,
            analysis=analysis,
            status=status,
            registration_source="original_auto",
        ))
    return artifacts


def _register_file_rollback_cleanup(
    db: Session,
    path: Path,
    *,
    package_id: int,
) -> None:
    state = {"committed": False}

    def after_commit(_session: Session) -> None:
        state["committed"] = True

    def after_rollback(_session: Session) -> None:
        if state["committed"]:
            return
        try:
            safe = require_contained_file(
                path,
                root=RFP_PACKAGES_DIR,
                expected_parent_name=str(package_id),
                description="rolled-back cost matrix",
            )
            if safe.exists():
                safe.unlink()
        except Exception:
            log.exception("failed to clean up rolled-back cost matrix %s", path)

    event.listen(db, "after_commit", after_commit, once=True)
    event.listen(db, "after_rollback", after_rollback, once=True)


def _review_late_attached_cost_matrix(
    proposal_id: int,
    document_id: int,
) -> int:
    """Run the normal source-aware review after the upload transaction commits."""

    # Local import avoids coupling the intake orchestrator to this service at
    # module-import time. The review itself owns durable failure handling.
    from app.jobs.intake import review_late_attached_requirements

    return review_late_attached_requirements(proposal_id, document_id)


def _mark_late_matrix_review_failed(document_id: int) -> None:
    """Fail closed when review dispatch itself crashes unexpectedly."""

    with session_scope() as db:
        document = db.get(RfpPackageDocument, document_id)
        if document is None:
            return
        structure = dict(document.structure_json or {})
        review = dict(structure.get("requirements_review") or {})
        review.update({
            "schema_version": 1,
            "status": "failed",
            "source_document_id": document_id,
            "requires_manual_review": True,
            "reason": (
                "Late-added cost matrix requirements review failed. "
                "Retry the review before scope sign-off."
            ),
            "updated_at": datetime.now(UTC).isoformat(),
        })
        structure["requirements_review"] = review
        document.structure_json = structure


def attach_cost_matrix(
    proposal_id: int,
    *,
    filename: str,
    content: bytes,
) -> int:
    """Attach a workbook later and return its artifact id.

    An upload through this dedicated action is explicit operator intent, so a
    structurally valid workbook is accepted even when the conservative
    classifier has low confidence. It still remains mapping-required.
    """
    analysis = inspect_cost_matrix(filename, content)
    if not analysis["classification"]["is_cost_matrix"]:
        analysis = deepcopy(analysis)
        analysis["classification"].update({
            "is_cost_matrix": True,
            "confidence": "operator",
            "evidence": [
                *analysis["classification"].get("evidence", []),
                "operator attached through the Cost Matrix workflow",
            ],
        })
    digest = _sha256(content)
    artifact_id = 0
    review_document_id = 0
    has_instruction_text = False
    dispatch_late_review = False
    with proposal_write_lock(proposal_id):
        with session_scope() as db:
            acquire_proposal_write_fence(db, proposal_id)
            proposal = ensure_proposal_mutable(
                db,
                proposal_id,
                operation="attach cost matrix",
            )
            if proposal is None:
                raise CostMatrixError(f"Proposal {proposal_id} was not found.")
            duplicate = db.scalar(
                select(CostMatrixArtifact.id)
                .join(RfpPackageDocument)
                .where(
                    CostMatrixArtifact.proposal_id == proposal_id,
                    RfpPackageDocument.structure_json.is_not(None),
                )
            )
            # SQLite JSON querying differs across versions; verify hashes in
            # Python so this works consistently on SQLite and server databases.
            if duplicate is not None:
                rows = db.execute(
                    select(CostMatrixArtifact, RfpPackageDocument)
                    .join(
                        RfpPackageDocument,
                        RfpPackageDocument.id == CostMatrixArtifact.source_document_id,
                    )
                    .where(CostMatrixArtifact.proposal_id == proposal_id)
                ).all()
                if any((doc.structure_json or {}).get("content_sha256") == digest for _, doc in rows):
                    raise CostMatrixError("This cost matrix is already attached to the proposal.")

            package_dir = RFP_PACKAGES_DIR / str(proposal.rfp_package_id)
            package_dir.mkdir(parents=True, exist_ok=True)
            base = _safe_filename(filename)
            safe = base
            bump = 0
            while (package_dir / safe).exists():
                bump += 1
                safe = f"{Path(base).stem}_{bump}{Path(base).suffix}"
            path = package_dir / safe
            temp_path = package_dir / f".{uuid4().hex}.uploading"
            try:
                temp_path.write_bytes(content)
                while True:
                    try:
                        os.link(temp_path, path)
                        break
                    except FileExistsError:
                        bump += 1
                        safe = f"{Path(base).stem}_{bump}{Path(base).suffix}"
                        path = package_dir / safe
                temp_path.unlink()
            finally:
                if temp_path.exists():
                    temp_path.unlink()
            _register_file_rollback_cleanup(
                db,
                path,
                package_id=proposal.rfp_package_id,
            )
            document = RfpPackageDocument(
                rfp_package_id=proposal.rfp_package_id,
                filename=filename,
                storage_path=str(path),
                document_type=RfpDocumentType.FORM_TEMPLATE,
                structure_json={"content_sha256": digest},
                document_role="supplemental",
            )
            db.add(document)
            db.flush()
            artifact = _register_artifact(
                db,
                proposal=proposal,
                document=document,
                content=content,
                analysis=analysis,
                status=STATUS_MAPPING_REQUIRED,
                registration_source="late_operator",
            )
            instruction_text, page_count = extract_cost_matrix_instruction_text(
                path,
                dict(artifact.analysis_json or analysis),
            )
            document.extracted_text_md = instruction_text or None
            document.page_count = page_count
            structure = dict(document.structure_json or {})
            structure.update({
                "cost_matrix_intake_analysis_version": ANALYSIS_VERSION,
                "cost_matrix_intake_policy_version": (
                    COST_MATRIX_INTAKE_POLICY_VERSION
                ),
                "cost_matrix_intake_mode": "instructions",
            })
            has_instruction_text = bool((instruction_text or "").strip())
            if has_instruction_text:
                current_status = (
                    proposal.status.value
                    if hasattr(proposal.status, "value")
                    else str(proposal.status)
                )
                scope_gate_reopened = current_status not in {
                    ProposalStatus.INTAKING.value,
                    ProposalStatus.AWAITING_SCOPE_SIGNOFF.value,
                }
                if scope_gate_reopened:
                    proposal.status = ProposalStatus.AWAITING_SCOPE_SIGNOFF
                # Active intake will discover this committed document in its
                # package-wide snapshot. Starting a second scoped extraction
                # here would race that run and could duplicate rows.
                dispatch_late_review = (
                    current_status != ProposalStatus.INTAKING.value
                )
                structure["requirements_review"] = {
                    "schema_version": 1,
                    "status": "pending",
                    "source_document_id": document.id,
                    "requires_manual_review": False,
                    "reason": (
                        "Late-added cost matrix instructions are queued for "
                        "source extraction and independent review."
                        + (
                            " Scope sign-off was reopened for these new instructions."
                            if scope_gate_reopened
                            else ""
                        )
                    ),
                    "scope_gate_reopened": scope_gate_reopened,
                    "extraction": {
                        "initial_item_count": 0,
                        "final_item_count": 0,
                    },
                    "updated_at": datetime.now(UTC).isoformat(),
                }
            else:
                structure["requirements_review"] = {
                    "schema_version": 1,
                    "status": "not_applicable",
                    "source_document_id": document.id,
                    "requires_manual_review": False,
                    "reason": (
                        "The cost matrix contains no separate written "
                        "requirements to review."
                    ),
                    "extraction": {
                        "initial_item_count": 0,
                        "final_item_count": 0,
                    },
                    "updated_at": datetime.now(UTC).isoformat(),
                }
            document.structure_json = structure
            _recompute_cots_from_current_intake_text(db, proposal)
            artifact_id = int(artifact.id)
            review_document_id = int(document.id)

    if dispatch_late_review:
        try:
            _review_late_attached_cost_matrix(
                proposal_id,
                review_document_id,
            )
        except Exception:
            log.exception(
                "late cost-matrix requirements review dispatch failed for document %d",
                review_document_id,
            )
            _mark_late_matrix_review_failed(review_document_id)
    return artifact_id


def _recompute_cots_from_current_intake_text(
    db: Session,
    proposal: Proposal,
) -> None:
    """Keep the deterministic COTS flag aligned after a role transition."""
    texts = db.execute(
        select(RfpPackageDocument.extracted_text_md)
        .where(RfpPackageDocument.rfp_package_id == proposal.rfp_package_id)
    ).scalars().all()
    proposal.cots_orientation = any(
        bool(_COTS_TEXT.search(text or "")) for text in texts
    )


def confirm_cost_matrix(proposal_id: int, artifact_id: int) -> None:
    """Confirm an ambiguous original workbook as a required cost matrix."""
    with proposal_write_lock(proposal_id):
        with session_scope() as db:
            acquire_proposal_write_fence(db, proposal_id)
            proposal = ensure_proposal_mutable(
                db,
                proposal_id,
                operation="confirm cost matrix",
            )
            artifact = db.get(CostMatrixArtifact, artifact_id)
            if proposal is None or artifact is None or artifact.proposal_id != proposal_id:
                raise CostMatrixError("Cost matrix candidate was not found for this proposal.")
            document = db.get(RfpPackageDocument, artifact.source_document_id)
            if document is None:
                raise CostMatrixError("Source workbook record is missing.")
            document.document_role = COST_MATRIX_ROLE
            document.document_type = RfpDocumentType.FORM_TEMPLATE
            structure = dict(document.structure_json or {})
            structure.update({
                "cost_matrix": True,
                "possible_cost_matrix": False,
                "cost_matrix_dismissed": False,
                "cost_matrix_intake_analysis_version": ANALYSIS_VERSION,
                "cost_matrix_intake_policy_version": (
                    COST_MATRIX_INTAKE_POLICY_VERSION
                ),
                "cost_matrix_intake_mode": "instructions",
            })
            document.structure_json = structure
            analysis = deepcopy(artifact.analysis_json or {})
            classification = dict(analysis.get("classification") or {})
            classification.update({
                "is_cost_matrix": True,
                "possible_cost_matrix": False,
                "decision": "confirmed",
                "confidence": "operator",
                "evidence": [
                    *classification.get("evidence", []),
                    "operator confirmed this workbook as a required cost matrix",
                ],
            })
            analysis["classification"] = classification
            source_path = require_contained_file(
                document.storage_path,
                root=RFP_PACKAGES_DIR,
                expected_parent_name=str(proposal.rfp_package_id),
                description="cost matrix source workbook",
            )
            if not source_path.is_file() or _sha256(source_path.read_bytes()) != artifact.template_sha256:
                raise CostMatrixError(
                    "Source workbook is missing or changed after inspection; reattach it."
                )
            structured_text, page_count = extract_cost_matrix_instruction_text(
                source_path,
                analysis,
            )
            # Replace any earlier broad candidate extraction immediately so
            # every later consumer sees instructions, never the pricing grid.
            document.extracted_text_md = structured_text or None
            document.page_count = page_count
            _recompute_cots_from_current_intake_text(db, proposal)
            analysis["role_transition"] = {
                "decision": "confirmed",
                "decided_at": datetime.now(UTC).isoformat(),
                "derived_context_policy": (
                    "visible requirement, evaluation, and COTS language retained; "
                    "pricing values and hidden sheets excluded"
                ),
            }
            artifact.analysis_json = analysis
            artifact.status = STATUS_MAPPING_REQUIRED
            artifact.human_reviewed_at = datetime.now(UTC)
            artifact.last_error = None


def dismiss_cost_matrix(
    proposal_id: int,
    artifact_id: int,
    *,
    reason: str | None = None,
) -> None:
    """Dismiss a false-positive matrix while retaining its audit record."""
    with proposal_write_lock(proposal_id):
        with session_scope() as db:
            acquire_proposal_write_fence(db, proposal_id)
            proposal = ensure_proposal_mutable(
                db,
                proposal_id,
                operation="dismiss cost matrix",
            )
            artifact = db.get(CostMatrixArtifact, artifact_id)
            if proposal is None or artifact is None or artifact.proposal_id != proposal_id:
                raise CostMatrixError("Cost matrix candidate was not found for this proposal.")
            document = db.get(RfpPackageDocument, artifact.source_document_id)
            if document is None:
                raise CostMatrixError("Source workbook record is missing.")
            analysis = deepcopy(artifact.analysis_json or {})
            registration = dict(analysis.get("registration") or {})
            document.document_role = registration.get("prior_document_role")
            prior_type = registration.get("prior_document_type")
            if prior_type:
                document.document_type = prior_type
            source_path = require_contained_file(
                document.storage_path,
                root=RFP_PACKAGES_DIR,
                expected_parent_name=str(proposal.rfp_package_id),
                description="ordinary workbook attachment",
            )
            if not source_path.is_file() or _sha256(source_path.read_bytes()) != artifact.template_sha256:
                raise CostMatrixError(
                    "Source workbook is missing or changed after inspection; reattach it."
                )
            ordinary_text, page_count = extract_cost_matrix_instruction_text(
                source_path,
                analysis,
                include_visible_context=True,
            )
            # Restore the broad visible-text representation immediately; a
            # retry must not reuse the reduced instruction-only matrix text.
            document.extracted_text_md = ordinary_text or None
            document.page_count = page_count
            _recompute_cots_from_current_intake_text(db, proposal)
            structure = dict(document.structure_json or {})
            structure.update({
                "cost_matrix": False,
                "possible_cost_matrix": False,
                "cost_matrix_dismissed": True,
                "cost_matrix_intake_analysis_version": ANALYSIS_VERSION,
                "cost_matrix_intake_policy_version": (
                    COST_MATRIX_INTAKE_POLICY_VERSION
                ),
                "cost_matrix_intake_mode": "visible_context",
            })
            document.structure_json = structure
            decision = dict(analysis.get("operator_decision") or {})
            decision.update({
                "decision": "dismissed",
                "reason": (reason or "Operator determined this workbook is not a fillable cost matrix").strip(),
                "decided_at": datetime.now(UTC).isoformat(),
            })
            analysis["operator_decision"] = decision
            artifact.analysis_json = analysis
            artifact.status = STATUS_DISMISSED
            artifact.human_reviewed_at = datetime.now(UTC)
            artifact.last_error = None


def _target_map(artifact: CostMatrixArtifact) -> dict[str, dict[str, Any]]:
    return {
        str(target["id"]): target
        for target in (artifact.analysis_json or {}).get("targets", [])
        if isinstance(target, dict) and target.get("id")
    }


def _coerce_manual_value(value: Any, kind: str) -> Any:
    if kind in {"money", "number", "percentage"}:
        if isinstance(value, bool) or value is None or str(value).strip() == "":
            raise CostMatrixError("A numeric manual value is required.")
        rendered = str(value).replace(",", "").replace("$", "").strip()
        has_percent_sign = rendered.endswith("%")
        if has_percent_sign:
            rendered = rendered[:-1].strip()
        try:
            number = Decimal(rendered)
        except (InvalidOperation, ValueError) as exc:
            raise CostMatrixError(f"{value!r} is not a valid numeric value.") from exc
        if not number.is_finite():
            raise CostMatrixError("Manual values must be finite numbers.")
        if kind == "percentage":
            if has_percent_sign:
                number /= Decimal("100")
            elif abs(number) > 1:
                raise CostMatrixError(
                    "Percentage values use workbook decimals: enter 24% or 0.24, not 24."
                )
        return float(number)
    if kind == "date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        try:
            return date.fromisoformat(str(value).strip()).isoformat()
        except ValueError as exc:
            raise CostMatrixError("Dates must use YYYY-MM-DD format.") from exc
    text = str(value or "").strip()
    if not text:
        raise CostMatrixError("A manual text value is required.")
    return text


def _validate_mappings(
    artifact: CostMatrixArtifact,
    mappings: dict[str, Any],
    sources: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    targets = _target_map(artifact)
    source_map = {str(source["key"]): source for source in sources}
    allowed = set(targets) | {RECONCILIATION_MAPPING_KEY}
    unknown = sorted(set(mappings) - allowed)
    if unknown:
        raise CostMatrixError(f"Unknown matrix target(s): {', '.join(unknown)}")
    normalized: dict[str, dict[str, Any]] = {}
    for target_id, raw in mappings.items():
        if not isinstance(raw, dict):
            raise CostMatrixError(f"Mapping for {target_id} is invalid.")
        if target_id == RECONCILIATION_MAPPING_KEY:
            mode = str(raw.get("mode") or "").strip()
            if mode == "aggregate_to_proposal_total":
                normalized[target_id] = {"mode": mode}
            elif mode == "independent_totals":
                reason = str(raw.get("reason") or "").strip()
                if not reason:
                    raise CostMatrixError(
                        "Explain why the buyer totals are independent of the overall proposal total."
                    )
                normalized[target_id] = {"mode": mode, "reason": reason}
            else:
                raise CostMatrixError(
                    "Choose how the workbook's multiple or partial totals reconcile."
                )
            continue
        target = targets[target_id]
        mode = str(raw.get("mode") or "").strip().lower()
        if mode == "source":
            source_key = str(raw.get("source_key") or "").strip()
            source = source_map.get(source_key)
            if source is None:
                raise CostMatrixError(
                    f"Mapped source {source_key!r} is not available for {target['label']}."
                )
            if source.get("value") is None:
                raise CostMatrixError(
                    f"Mapped source {source['label']} has no value yet."
                )
            if str(target.get("kind") or "text") != str(source.get("kind") or ""):
                raise CostMatrixError(
                    f"{source['label']} uses {source.get('kind') or 'an unknown unit'} and "
                    f"cannot populate the {target.get('kind') or 'text'} field "
                    f"{target['label']}. Enter an explicitly reviewed manual value if a "
                    "conversion is required."
                )
            normalized[target_id] = {
                "mode": "source",
                "source_key": source_key,
            }
        elif mode == "manual":
            normalized[target_id] = {
                "mode": "manual",
                "value": _coerce_manual_value(raw.get("value"), str(target.get("kind") or "text")),
                "note": str(raw.get("note") or "").strip() or None,
            }
        elif mode == "skip":
            reason = str(raw.get("reason") or "").strip()
            if not reason:
                raise CostMatrixError(
                    f"Explain why {target['label']} should remain unchanged."
                )
            normalized[target_id] = {"mode": "skip", "reason": reason}
        else:
            raise CostMatrixError(
                f"Choose a source, manual approved value, or explicit skip for {target['label']}."
            )
    return normalized


def save_cost_matrix_mapping(
    proposal_id: int,
    artifact_id: int,
    mappings: dict[str, Any],
) -> None:
    with proposal_write_lock(proposal_id):
        with session_scope() as db:
            acquire_proposal_write_fence(db, proposal_id)
            proposal = ensure_proposal_mutable(
                db,
                proposal_id,
                operation="map cost matrix",
            )
            artifact = db.get(CostMatrixArtifact, artifact_id)
            if proposal is None or artifact is None or artifact.proposal_id != proposal_id:
                raise CostMatrixError("Cost matrix was not found for this proposal.")
            if artifact.status in {STATUS_NEEDS_CONFIRMATION, STATUS_DISMISSED}:
                raise CostMatrixError(
                    "Confirm this workbook as a cost matrix before mapping it."
                )
            sources, _basis = _source_catalog(db, proposal)
            artifact.mapping_json = _validate_mappings(artifact, mappings, sources)
            artifact.last_error = None
            target_ids = set(_target_map(artifact))
            artifact.status = (
                STATUS_WAITING_FOR_COSTS
                if target_ids and target_ids.issubset(artifact.mapping_json)
                else STATUS_MAPPING_REQUIRED
            )


def add_cost_matrix_target(
    proposal_id: int,
    artifact_id: int,
    *,
    sheet: str,
    cell_coordinate: str,
    label: str | None = None,
    kind: str = "money",
) -> str:
    """Add an exact template cell when conservative inspection missed it."""
    from openpyxl.utils.cell import coordinate_to_tuple

    allowed_kinds = {"money", "number", "percentage", "text", "date"}
    if kind not in allowed_kinds:
        raise CostMatrixError(f"Unsupported target kind: {kind}")
    try:
        row, column = coordinate_to_tuple(cell_coordinate.upper().strip())
    except Exception as exc:
        raise CostMatrixError("Cell must be an Excel coordinate such as C9.") from exc
    if not (1 <= row <= 1_048_576 and 1 <= column <= 16_384):
        raise CostMatrixError("Cell is outside Excel's supported worksheet bounds.")

    with proposal_write_lock(proposal_id):
        with session_scope() as db:
            acquire_proposal_write_fence(db, proposal_id)
            proposal = ensure_proposal_mutable(
                db,
                proposal_id,
                operation="add cost matrix target",
            )
            artifact = db.get(CostMatrixArtifact, artifact_id)
            if proposal is None or artifact is None or artifact.proposal_id != proposal_id:
                raise CostMatrixError("Cost matrix was not found for this proposal.")
            if artifact.status in {STATUS_NEEDS_CONFIRMATION, STATUS_DISMISSED}:
                raise CostMatrixError(
                    "Confirm this workbook as a cost matrix before adding targets."
                )
            document = db.get(RfpPackageDocument, artifact.source_document_id)
            if document is None:
                raise CostMatrixError("Source workbook record is missing.")
            source_path = require_contained_file(
                document.storage_path,
                root=RFP_PACKAGES_DIR,
                expected_parent_name=str(proposal.rfp_package_id),
                description="cost matrix source workbook",
            )
            data = source_path.read_bytes()
            if _sha256(data) != artifact.template_sha256:
                raise CostMatrixError("Source workbook changed after inspection; reattach it.")
            workbook = load_workbook(io.BytesIO(data), data_only=False, keep_links=False)
            try:
                if sheet not in workbook.sheetnames:
                    raise CostMatrixError(f"Workbook has no sheet named {sheet!r}.")
                ws = workbook[sheet]
                if ws.sheet_state != "visible":
                    raise CostMatrixError(
                        "Hidden worksheets are preserved as source material and cannot be mapped."
                    )
                if _cell_hidden(ws, row, column):
                    raise CostMatrixError(
                        "Cells in hidden rows or columns are preserved as source material "
                        "and cannot be mapped."
                    )
                raw_cell = ws.cell(row=row, column=column)
                if ws.protection.sheet and raw_cell.protection.locked:
                    raise CostMatrixError(
                        "This cell is locked on a protected worksheet and cannot be mapped."
                    )
                top_left, merged_range = _merged_top_left(ws, raw_cell.coordinate)
                if top_left != raw_cell.coordinate:
                    raise CostMatrixError(
                        f"{raw_cell.coordinate} is inside merged range {merged_range}; "
                        f"map its top-left cell {top_left} instead."
                    )
                if _cell_hidden(ws, raw_cell.row, raw_cell.column):
                    raise CostMatrixError(
                        "Cells in hidden rows or columns cannot be mapped."
                    )
                if _is_formula(raw_cell.value):
                    raise CostMatrixError("Formula cells are read-only and cannot be mapped.")
                analysis = deepcopy(artifact.analysis_json or {})
                targets = list(analysis.get("targets") or [])
                if any(
                    target.get("sheet") == sheet
                    and str(target.get("cell")).upper() == raw_cell.coordinate.upper()
                    for target in targets
                ):
                    raise CostMatrixError("That workbook cell is already a mapping target.")
                existing_ids = {
                    str(target.get("id")) for target in targets if target.get("id")
                }
                next_number = 1
                while f"manual-{next_number}" in existing_ids:
                    next_number += 1
                target_id = f"manual-{next_number}"
                target = _candidate_target(
                    target_id=target_id,
                    sheet=sheet,
                    cell=raw_cell,
                    label=(label or "").strip() or _nearest_text_left(ws, row, column) or raw_cell.coordinate,
                    header=_nearest_header_above(ws, row, column),
                    category=(
                        "metadata" if kind in {"text", "date"} else "pricing"
                    ),
                    confidence="operator",
                    evidence=["operator-added exact workbook cell"],
                    merged_range=merged_range,
                    origin="operator",
                )
                target["id"] = target_id
                target["kind"] = kind
                targets.append(target)
                analysis["targets"] = targets
                reconciliations, review = _build_reconciliation_manifest(
                    targets,
                    list(analysis.get("formulas") or []),
                    visible_sheets={
                        str(sheet_info["name"])
                        for sheet_info in (analysis.get("sheets") or [])
                        if sheet_info.get("state") == "visible"
                    },
                )
                analysis["reconciliations"] = reconciliations
                analysis["reconciliation_review"] = review
                artifact.analysis_json = analysis
                artifact.status = STATUS_MAPPING_REQUIRED
                return target_id
            finally:
                workbook.close()


def _resolve_mapping_values(
    artifact: CostMatrixArtifact,
    sources: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[str]]:
    targets = _target_map(artifact)
    source_map = {str(source["key"]): source for source in sources}
    mappings = artifact.mapping_json or {}
    resolved: dict[str, Any] = {}
    blockers: list[str] = []
    if not any(target.get("category") == "pricing" for target in targets.values()):
        blockers.append(
            "No financial input cells are defined; add the workbook cells that require pricing."
        )
    for target_id, target in targets.items():
        mapping = mappings.get(target_id)
        if not isinstance(mapping, dict):
            blockers.append(f"Map or explicitly skip {target['label']} ({target['sheet']}!{target['cell']}).")
            continue
        mode = mapping.get("mode")
        if mode == "skip":
            if not str(mapping.get("reason") or "").strip():
                blockers.append(f"Provide a skip reason for {target['label']}.")
            continue
        if mode == "source":
            source = source_map.get(str(mapping.get("source_key") or ""))
            if source is None or source.get("value") is None:
                blockers.append(f"Mapped value for {target['label']} is no longer available.")
                continue
            target_kind = str(target.get("kind") or "text")
            source_kind = str(source.get("kind") or "")
            if target_kind != source_kind:
                blockers.append(
                    f"Mapped source units changed for {target['label']}: "
                    f"expected {target_kind}, now {source_kind or 'unknown'}. Remap it."
                )
                continue
            try:
                resolved[target_id] = _coerce_manual_value(
                    source["value"],
                    target_kind,
                )
            except CostMatrixError as exc:
                blockers.append(
                    f"Current mapped value for {target['label']} is invalid: {exc}"
                )
        elif mode == "manual":
            try:
                resolved[target_id] = _coerce_manual_value(
                    mapping.get("value"),
                    str(target.get("kind") or "text"),
                )
            except CostMatrixError as exc:
                blockers.append(f"{target['label']}: {exc}")
        else:
            blockers.append(f"Choose how to populate {target['label']}.")
    numeric_pricing_ids = [
        target_id
        for target_id, target in targets.items()
        if (
            target.get("category") == "pricing"
            and target.get("kind") in {"money", "number", "percentage"}
        )
    ]
    money_ids = [
        target_id
        for target_id in numeric_pricing_ids
        if targets[target_id].get("kind") == "money"
    ]
    # If the workbook exposes monetary cells, populating only quantities or
    # percentages is not a completed cost matrix. For unit-only templates,
    # fall back to their numeric pricing fields.
    financial_ids = money_ids or numeric_pricing_ids
    if financial_ids and all(
        (mappings.get(target_id) or {}).get("mode") == "skip"
        for target_id in financial_ids
    ):
        blockers.append(
            "All financial inputs are skipped. Populate at least one financial cell, "
            "or dismiss/reclassify this workbook if it is not a fillable cost matrix."
        )
    return resolved, blockers


def _workflow_blockers(basis: dict[str, Any]) -> list[str]:
    blockers: list[str] = []
    if basis.get("service_line") == "payment_systems":
        if not basis.get("pricing_exists"):
            blockers.append("Run Payment Market Research before completing the matrix.")
        elif not basis.get("payment_scan_current"):
            blockers.append("Refresh the payment pricing basis before completing the matrix.")
        if not basis.get("payment_review_current"):
            blockers.append("Run the Payment Cost Reviewer before completing the matrix.")
    else:
        if not basis.get("scenario"):
            blockers.append("Select a pricing scenario explicitly on the Cost tab.")
        if not basis.get("pricing_exists"):
            blockers.append("Run the Cost Analyst for the selected scenario first.")
        review = basis.get("cost_review") or {}
        if not review.get("verified"):
            blockers.append(str(review.get("detail") or "Run the Cost Reviewer first."))
    return blockers


def _reconciliation_blockers(
    artifact: CostMatrixArtifact,
    *,
    sources: list[dict[str, Any]],
    resolved: dict[str, Any],
) -> list[str]:
    source_map = {str(source["key"]): source for source in sources}
    expected = source_map.get("pricing.total_proposed_price")
    targets = _target_map(artifact)
    mappings = artifact.mapping_json or {}
    blockers: list[str] = []
    review = (artifact.analysis_json or {}).get("reconciliation_review") or {}
    if review.get("review_required"):
        choice = mappings.get(RECONCILIATION_MAPPING_KEY) or {}
        mode = choice.get("mode")
        if mode not in {"aggregate_to_proposal_total", "independent_totals"}:
            return [
                "Review the workbook's multiple or partial buyer totals and choose "
                "whether all unique inputs aggregate to the proposal total or the "
                "totals are independent."
            ]
        if mode == "independent_totals":
            if not str(choice.get("reason") or "").strip():
                return ["Explain why the buyer totals are independent."]
            return []
        if expected is None or expected.get("value") is None:
            return [
                "The selected pricing basis does not expose one overall proposal total; "
                "choose independent totals with an explanation or finish the price build."
            ]
        financial_ids = [
            target_id
            for target_id, target in targets.items()
            if (
                target.get("category") == "pricing"
                and target.get("kind") == "money"
            )
        ]
        values: list[Decimal] = []
        for target_id in financial_ids:
            if target_id in resolved:
                raw = resolved[target_id]
            elif (mappings.get(target_id) or {}).get("mode") == "skip":
                raw = (targets.get(target_id) or {}).get("existing_value")
                if not isinstance(raw, (int, float, Decimal)) or isinstance(raw, bool):
                    values.append(Decimal("0"))
                    continue
            else:
                return []  # ordinary mapping blockers are more actionable
            try:
                values.append(Decimal(str(raw)))
            except (InvalidOperation, TypeError, ValueError):
                return [
                    "A mapped financial value cannot be converted for reconciliation; "
                    "review the affected matrix mapping."
                ]
        actual = sum(values, Decimal("0"))
        expected_total = Decimal(str(expected["value"]))
        if abs(actual - expected_total) > Decimal("0.01"):
            blockers.append(
                f"The unique mapped financial inputs total ${actual:,.2f}, but the "
                f"selected proposed price is ${expected_total:,.2f}; reconcile them "
                "to the cent or choose independent totals with an explanation."
            )
        return blockers

    if expected is None or expected.get("value") is None:
        return []
    for rule in (artifact.analysis_json or {}).get("reconciliations", []):
        if rule.get("type") != "sum_to_proposed_price":
            continue
        values: list[Decimal] = []
        missing = False
        for target_id in rule.get("member_target_ids") or []:
            if target_id in resolved:
                raw = resolved[target_id]
            elif (mappings.get(target_id) or {}).get("mode") == "skip":
                raw = (targets.get(target_id) or {}).get("existing_value")
                if not isinstance(raw, (int, float, Decimal)) or isinstance(raw, bool):
                    values.append(Decimal("0"))
                    continue
            else:
                missing = True
                break
            try:
                values.append(Decimal(str(raw)))
            except (InvalidOperation, TypeError, ValueError):
                blockers.append(
                    f"Buyer total {rule['sheet']}!{rule['formula_cell']} cannot be "
                    "reconciled because a mapped value is not numeric."
                )
                missing = True
                break
        if missing:
            continue  # the ordinary mapping blockers are more actionable
        actual = sum(values, Decimal("0"))
        expected_total = Decimal(str(expected["value"]))
        tolerance = Decimal(str(rule.get("tolerance_usd") or 0.01))
        if abs(actual - expected_total) > tolerance:
            blockers.append(
                f"Buyer total {rule['sheet']}!{rule['formula_cell']} would be "
                f"${actual:,.2f}, but the selected proposed price is "
                f"${expected_total:,.2f}; reconcile the mapped line items to the cent."
            )
    return blockers


def _generation_provenance(
    artifact: CostMatrixArtifact,
    *,
    sources: list[dict[str, Any]],
    basis: dict[str, Any],
    resolved: dict[str, Any],
) -> dict[str, Any]:
    provenance = {
        "version": "cost-matrix-generation-v1",
        "template_sha256": artifact.template_sha256,
        "analysis_version": artifact.analysis_version,
        "mapping": artifact.mapping_json or {},
        "resolved_values": resolved,
        "workflow_basis": basis,
        "source_values": {
            source["key"]: source.get("value")
            for source in sources
            if source.get("key") in {
                mapping.get("source_key")
                for mapping in (artifact.mapping_json or {}).values()
                if isinstance(mapping, dict) and mapping.get("mode") == "source"
            }
        },
    }
    provenance["sha256"] = _canonical_sha256(provenance)
    return provenance


def _latest_output(artifact: CostMatrixArtifact) -> CostMatrixOutput | None:
    return max(artifact.outputs, key=lambda output: output.version, default=None)


def _output_integrity(
    proposal: Proposal,
    output: CostMatrixOutput | None,
) -> tuple[bool, str | None]:
    if output is None:
        return False, None
    try:
        path = require_contained_file(
            output.output_storage_path,
            root=RFP_PACKAGES_DIR,
            description="generated cost matrix",
        )
        package_root = (RFP_PACKAGES_DIR / str(proposal.rfp_package_id)).resolve()
        path.resolve().relative_to(package_root)
        if not path.is_file():
            return False, "The latest generated output file is missing; regenerate it."
        if _sha256(path.read_bytes()) != output.output_sha256:
            return False, "The latest generated output failed its integrity check; regenerate it."
        return True, None
    except Exception:
        return False, (
            "The latest generated output is outside this proposal's managed package "
            "or cannot be verified; regenerate it."
        )


def _readiness(
    db: Session,
    proposal: Proposal,
    artifact: CostMatrixArtifact,
) -> dict[str, Any]:
    sources, basis = _source_catalog(db, proposal)
    resolved, generation_blockers = _resolve_mapping_values(artifact, sources)
    generation_blockers.extend(_reconciliation_blockers(
        artifact,
        sources=sources,
        resolved=resolved,
    ))
    generation_blockers.extend(_workflow_blockers(basis))

    document = db.get(RfpPackageDocument, artifact.source_document_id)
    source_ok = False
    if document is None:
        generation_blockers.append("Source workbook record is missing.")
    else:
        try:
            source_path = require_contained_file(
                document.storage_path,
                root=RFP_PACKAGES_DIR,
                expected_parent_name=str(proposal.rfp_package_id),
                description="cost matrix source workbook",
            )
            source_ok = source_path.is_file() and _sha256(source_path.read_bytes()) == artifact.template_sha256
        except Exception:
            source_ok = False
        if not source_ok:
            generation_blockers.append("Source workbook is missing or changed after inspection.")

    provenance = _generation_provenance(
        artifact,
        sources=sources,
        basis=basis,
        resolved=resolved,
    )
    latest = _latest_output(artifact)
    output_intact, output_issue = _output_integrity(proposal, latest)
    blockers = list(generation_blockers)
    if latest is not None and output_issue:
        blockers.append(output_issue)
    current_output = bool(
        latest
        and latest.pricing_basis_sha256 == provenance["sha256"]
        and not generation_blockers
        and source_ok
        and output_intact
    )
    suggestions = {
        target_id: suggestion
        for target_id, target in _target_map(artifact).items()
        if target_id not in (artifact.mapping_json or {})
        if (suggestion := _mapping_suggestion(target, sources)) is not None
    }
    return {
        "ready": not generation_blockers,
        "blockers": blockers,
        "resolved_values": resolved,
        "basis": basis,
        "sources": sources,
        "provenance": provenance,
        "has_output": latest is not None,
        "output_current": current_output,
        "suggestions": suggestions,
    }


def _xlsx_sheet_parts(data: bytes) -> dict[str, str]:
    with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall(f"{{{_NS_REL_PKG}}}Relationship")
        if rel.attrib.get("Id") and rel.attrib.get("Target")
    }
    parts: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{_NS_REL_DOC}}}id")
        target = rels.get(rel_id or "")
        if not name or not target:
            continue
        target_path = PurePosixPath(target.lstrip("/"))
        if target.startswith("/"):
            part = str(target_path)
        else:
            part = str(PurePosixPath("xl") / target_path)
        # Resolve occasional '../' relationship targets safely.
        normalized: list[str] = []
        for token in PurePosixPath(part).parts:
            if token == "..":
                if normalized:
                    normalized.pop()
            elif token not in {".", "/"}:
                normalized.append(token)
        parts[name] = "/".join(normalized)
    return parts


def _excel_number(value: Any) -> str:
    number = Decimal(str(value))
    if not number.is_finite():
        raise CostMatrixError("Workbook values must be finite numbers.")
    rendered = format(number, "f")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered or "0"


def _cell_payload(
    value: Any,
    kind: str,
    *,
    date_1904: bool,
    number_format: str = "General",
    prefix: str = "",
) -> tuple[str, str]:
    if kind == "date":
        parsed = value
        if not isinstance(parsed, (date, datetime)):
            parsed = date.fromisoformat(str(value))
        if isinstance(parsed, datetime):
            parsed = parsed.date()
        # A date serial in a General-formatted cell renders as an unexplained
        # integer. Preserve the buyer's style: date-formatted destinations get
        # an Excel serial; General destinations get human-readable ISO text.
        if not is_date_format(number_format or ""):
            text = parsed.isoformat()
            return (
                "inlineStr",
                f"<{prefix}is><{prefix}t>{xml_escape(text)}</{prefix}t></{prefix}is>",
            )
        from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900
        serial = to_excel(
            parsed,
            epoch=CALENDAR_MAC_1904 if date_1904 else CALENDAR_WINDOWS_1900,
        )
        return "n", f"<{prefix}v>{_excel_number(serial)}</{prefix}v>"
    if kind in {"money", "number", "percentage"}:
        return "n", f"<{prefix}v>{_excel_number(value)}</{prefix}v>"
    if isinstance(value, bool):
        return "b", f"<{prefix}v>{1 if value else 0}</{prefix}v>"
    text = str(value)
    space = ' xml:space="preserve"' if text != text.strip() or "\n" in text else ""
    return (
        "inlineStr",
        f"<{prefix}is><{prefix}t{space}>{xml_escape(text)}</{prefix}t></{prefix}is>",
    )


def _column_index(coordinate: str) -> int:
    letters = re.match(r"[A-Z]+", coordinate.upper())
    if not letters:
        raise CostMatrixError(f"Invalid cell coordinate {coordinate!r}.")
    value = 0
    for char in letters.group(0):
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def _replace_or_insert_cell_xml(
    sheet_xml: bytes,
    *,
    coordinate: str,
    value: Any,
    kind: str,
    date_1904: bool,
    number_format: str = "General",
) -> bytes:
    text = sheet_xml.decode("utf-8")
    coord = coordinate.upper()
    root_match = re.search(
        r"<(?P<prefix>[A-Za-z_][A-Za-z0-9_.-]*:)?worksheet\b",
        text,
    )
    if not root_match:
        raise CostMatrixError("Worksheet XML has no worksheet root element.")
    prefix = root_match.group("prefix") or ""
    tag_prefix = re.escape(prefix)
    cell_type, payload = _cell_payload(
        value,
        kind,
        date_1904=date_1904,
        number_format=number_format,
        prefix=prefix,
    )
    coord_re = re.escape(coord)
    opening_cell = re.compile(
        rf"<{tag_prefix}c\b(?=[^>]*\br=[\"']{coord_re}[\"'])(?P<attrs>[^>]*?)(?P<self>/?)>",
        re.DOTALL,
    )
    opening_match = opening_cell.search(text)
    if opening_match:
        attrs = opening_match.group("attrs")
        if opening_match.group("self") == "/":
            match_start = opening_match.start()
            match_end = opening_match.end()
            body = ""
        else:
            closing_tag = f"</{prefix}c>"
            closing_at = text.find(closing_tag, opening_match.end())
            if closing_at < 0:
                raise CostMatrixError(f"Malformed worksheet cell {coord}.")
            match_start = opening_match.start()
            match_end = closing_at + len(closing_tag)
            body = text[opening_match.end():closing_at]
        if re.search(rf"<{tag_prefix}f(?:\s|>)", body):
            raise CostMatrixError(f"Refusing to overwrite formula cell {coord}.")
        attrs = re.sub(r"\s+t=(?:\"[^\"]*\"|'[^']*')", "", attrs)
        attrs = attrs.rstrip() + f' t="{cell_type}"'
        body = re.sub(
            rf"<{tag_prefix}v(?:\s[^>]*)?>.*?</{tag_prefix}v>",
            "",
            body,
            flags=re.DOTALL,
        )
        body = re.sub(
            rf"<{tag_prefix}is(?:\s[^>]*)?>.*?</{tag_prefix}is>",
            "",
            body,
            flags=re.DOTALL,
        )
        replacement = f"<{prefix}c{attrs}>{body}{payload}</{prefix}c>"
        updated = text[:match_start] + replacement + text[match_end:]
        return _expand_worksheet_dimension(updated, coord, prefix=prefix).encode("utf-8")

    # Explicit operator targets can point at an unmaterialized blank cell.
    # Insert it into the existing row without normalizing the rest of the XML.
    row_number_match = re.search(r"\d+", coord)
    if not row_number_match:
        raise CostMatrixError(f"Invalid cell coordinate {coord!r}.")
    row_number = int(row_number_match.group(0))
    new_cell = (
        f'<{prefix}c r="{coord}" t="{cell_type}">{payload}</{prefix}c>'
    )
    row_re = re.compile(
        rf"<{tag_prefix}row\b(?P<attrs>[^>]*\br=[\"']{row_number}[\"'][^>]*)>"
        rf"(?P<body>.*?)</{tag_prefix}row>",
        re.DOTALL,
    )
    row_match = row_re.search(text)
    if row_match:
        body = row_match.group("body")
        target_column = _column_index(coord)
        insert_at = len(body)
        for cell_match in re.finditer(
            rf"<{tag_prefix}c\b[^>]*\br=[\"']([A-Z]+\d+)[\"'][^>]*"
            rf"(?:/>|>.*?</{tag_prefix}c>)",
            body,
            re.DOTALL,
        ):
            if _column_index(cell_match.group(1)) > target_column:
                insert_at = cell_match.start()
                break
        new_body = body[:insert_at] + new_cell + body[insert_at:]
        replacement = (
            f"<{prefix}row{row_match.group('attrs')}>{new_body}</{prefix}row>"
        )
        updated = text[:row_match.start()] + replacement + text[row_match.end():]
        return _expand_worksheet_dimension(updated, coord, prefix=prefix).encode("utf-8")

    self_row_re = re.compile(
        rf"<{tag_prefix}row\b(?P<attrs>[^>]*\br=[\"']{row_number}[\"'][^>]*)/>",
        re.DOTALL,
    )
    self_row_match = self_row_re.search(text)
    if self_row_match:
        replacement = (
            f"<{prefix}row{self_row_match.group('attrs')}>{new_cell}</{prefix}row>"
        )
        updated = (
            text[:self_row_match.start()] + replacement + text[self_row_match.end():]
        )
        return _expand_worksheet_dimension(updated, coord, prefix=prefix).encode("utf-8")

    new_row = f'<{prefix}row r="{row_number}">{new_cell}</{prefix}row>'
    empty_sheet_data = re.compile(rf"<{tag_prefix}sheetData\b(?P<attrs>[^>]*)/>")
    empty_match = empty_sheet_data.search(text)
    if empty_match:
        replacement = (
            f"<{prefix}sheetData{empty_match.group('attrs')}>"
            f"{new_row}</{prefix}sheetData>"
        )
        updated = text[:empty_match.start()] + replacement + text[empty_match.end():]
        return _expand_worksheet_dimension(updated, coord, prefix=prefix).encode("utf-8")

    sheet_data_re = re.compile(
        rf"<{tag_prefix}sheetData\b[^>]*>(?P<body>.*?)</{tag_prefix}sheetData>",
        re.DOTALL,
    )
    sheet_data_match = sheet_data_re.search(text)
    if not sheet_data_match:
        raise CostMatrixError("Worksheet has no writable sheetData section.")
    body = sheet_data_match.group("body")
    insert_at = len(body)
    any_row_re = re.compile(
        rf"<{tag_prefix}row\b[^>]*\br=[\"'](\d+)[\"'][^>]*"
        rf"(?:/>|>.*?</{tag_prefix}row>)",
        re.DOTALL,
    )
    for match in any_row_re.finditer(body):
        if int(match.group(1)) > row_number:
            insert_at = match.start()
            break
    new_body = body[:insert_at] + new_row + body[insert_at:]
    updated = (
        text[:sheet_data_match.start("body")]
        + new_body
        + text[sheet_data_match.end("body"):]
    )
    return _expand_worksheet_dimension(updated, coord, prefix=prefix).encode("utf-8")


def _expand_worksheet_dimension(text: str, coordinate: str, *, prefix: str) -> str:
    tag_prefix = re.escape(prefix)
    dimension_re = re.compile(
        rf"<{tag_prefix}dimension\b(?P<before>[^>]*?\bref=)"
        rf"(?P<quote>[\"'])(?P<ref>[^\"']+)(?P=quote)(?P<after>[^>]*)/?>",
        re.DOTALL,
    )
    match = dimension_re.search(text)
    if not match:
        return text
    try:
        min_col, min_row, max_col, max_row = range_boundaries(match.group("ref"))
        cell_col, cell_row, _, _ = range_boundaries(coordinate)
    except ValueError:
        return text
    min_col = min(min_col, cell_col)
    max_col = max(max_col, cell_col)
    min_row = min(min_row, cell_row)
    max_row = max(max_row, cell_row)
    expanded = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    trailing_attrs = re.sub(r"/\s*$", "", match.group("after"))
    replacement = (
        f"<{prefix}dimension{match.group('before')}"
        f"{match.group('quote')}{expanded}{match.group('quote')}"
        f"{trailing_attrs}/>"
    )
    return text[:match.start()] + replacement + text[match.end():]


def _force_formula_recalculation(workbook_xml: bytes) -> bytes:
    text = workbook_xml.decode("utf-8")
    root_match = re.search(
        r"<(?P<prefix>[A-Za-z_][A-Za-z0-9_.-]*:)?workbook\b",
        text,
    )
    if not root_match:
        raise CostMatrixError("Workbook XML has no workbook root element.")
    prefix = root_match.group("prefix") or ""
    tag_prefix = re.escape(prefix)
    calc_re = re.compile(
        rf"<{tag_prefix}calcPr\b(?P<attrs>[^>]*?)"
        rf"(?:/\s*>|>.*?</{tag_prefix}calcPr\s*>)",
        re.DOTALL,
    )
    match = calc_re.search(text)
    required = {
        "calcMode": "auto",
        "fullCalcOnLoad": "1",
        "forceFullCalc": "1",
    }
    if match:
        attrs = match.group("attrs")
        for key, value in required.items():
            attr_re = re.compile(rf"\s+{key}=(?:\"[^\"]*\"|'[^']*')")
            if attr_re.search(attrs):
                attrs = attr_re.sub(f' {key}="{value}"', attrs)
            else:
                attrs += f' {key}="{value}"'
        replacement = f"<{prefix}calcPr{attrs}/>"
        return (text[:match.start()] + replacement + text[match.end():]).encode("utf-8")
    marker = f"</{prefix}workbook>"
    if marker not in text:
        raise CostMatrixError("Workbook XML is missing its closing element.")
    calc = (
        f'<{prefix}calcPr calcMode="auto" fullCalcOnLoad="1" '
        'forceFullCalc="1"/>'
    )
    return text.replace(marker, calc + marker, 1).encode("utf-8")


def _patch_workbook(
    source: bytes,
    *,
    targets: dict[str, dict[str, Any]],
    resolved: dict[str, Any],
) -> tuple[bytes, list[str]]:
    sheet_parts = _xlsx_sheet_parts(source)
    with zipfile.ZipFile(io.BytesIO(source), "r") as archive:
        originals = {info.filename: archive.read(info.filename) for info in archive.infolist()}
        infos = archive.infolist()
    workbook_xml = originals["xl/workbook.xml"]
    date_1904 = bool(re.search(rb"\bdate1904=(?:\"1\"|'1'|\"true\"|'true')", workbook_xml))
    changed: set[str] = set()
    grouped: dict[str, list[tuple[dict[str, Any], Any]]] = {}
    for target_id, value in resolved.items():
        target = targets[target_id]
        grouped.setdefault(str(target["sheet"]), []).append((target, value))
    for sheet, writes in grouped.items():
        part = sheet_parts.get(sheet)
        if not part or part not in originals:
            raise CostMatrixError(f"Could not locate worksheet part for {sheet!r}.")
        updated = originals[part]
        for target, value in writes:
            updated = _replace_or_insert_cell_xml(
                updated,
                coordinate=str(target["cell"]),
                value=value,
                kind=str(target.get("kind") or "text"),
                date_1904=date_1904,
                number_format=str(target.get("number_format") or "General"),
            )
        originals[part] = updated
        changed.add(part)
    originals["xl/workbook.xml"] = _force_formula_recalculation(workbook_xml)
    if originals["xl/workbook.xml"] != workbook_xml:
        changed.add("xl/workbook.xml")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        for info in infos:
            archive.writestr(info, originals[info.filename])
    return output.getvalue(), sorted(changed)


def _comparable_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _validate_generated_workbook(
    source: bytes,
    output: bytes,
    *,
    artifact: CostMatrixArtifact,
    resolved: dict[str, Any],
    changed_parts: list[str],
) -> dict[str, Any]:
    source_wb = load_workbook(io.BytesIO(source), data_only=False, keep_links=False)
    output_wb = load_workbook(io.BytesIO(output), data_only=False, keep_links=False)
    try:
        if source_wb.sheetnames != output_wb.sheetnames:
            raise CostMatrixError("Generated workbook changed sheet names or order.")
        for source_ws, output_ws in zip(
            source_wb.worksheets,
            output_wb.worksheets,
            strict=True,
        ):
            if source_ws.sheet_state != output_ws.sheet_state:
                raise CostMatrixError(f"Generated workbook changed visibility of {source_ws.title!r}.")
            source_merges = sorted(str(item) for item in source_ws.merged_cells.ranges)
            output_merges = sorted(str(item) for item in output_ws.merged_cells.ranges)
            if source_merges != output_merges:
                raise CostMatrixError(f"Generated workbook changed merges on {source_ws.title!r}.")
        for formula in (artifact.analysis_json or {}).get("formulas", []):
            sheet = formula["sheet"]
            cell = formula["cell"]
            if output_wb[sheet][cell].value != formula["formula"]:
                raise CostMatrixError(f"Generated workbook changed formula {sheet}!{cell}.")
        targets = _target_map(artifact)
        written: list[dict[str, Any]] = []
        for target_id, expected in resolved.items():
            target = targets[target_id]
            actual = _comparable_value(output_wb[target["sheet"]][target["cell"]].value)
            expected_value = _comparable_value(expected)
            if isinstance(expected_value, (int, float)) and isinstance(actual, (int, float)):
                equal = math.isclose(float(actual), float(expected_value), rel_tol=0, abs_tol=1e-8)
            else:
                equal = actual == expected_value
            if not equal:
                raise CostMatrixError(
                    f"Validation failed for {target['sheet']}!{target['cell']}: "
                    f"expected {expected_value!r}, read {actual!r}."
                )
            source_style = source_wb[target["sheet"]][target["cell"]].style_id
            output_style = output_wb[target["sheet"]][target["cell"]].style_id
            if source_style != output_style:
                raise CostMatrixError(
                    f"Generated workbook changed the style of {target['sheet']}!{target['cell']}."
                )
            written.append({
                "target_id": target_id,
                "sheet": target["sheet"],
                "cell": target["cell"],
                "value": expected_value,
            })
        return {
            "valid": True,
            "written_cells": written,
            "formula_count_preserved": len((artifact.analysis_json or {}).get("formulas", [])),
            "sheet_count_preserved": len(source_wb.sheetnames),
            "changed_package_parts": changed_parts,
            "formula_recalculation_requested": True,
        }
    finally:
        source_wb.close()
        output_wb.close()


def generate_cost_matrix(proposal_id: int, artifact_id: int) -> int:
    """Generate and atomically publish a new immutable workbook revision."""
    published_path: Path | None = None
    try:
        with proposal_write_lock(proposal_id):
            with session_scope() as db:
                acquire_proposal_write_fence(db, proposal_id)
                proposal = ensure_proposal_mutable(
                    db,
                    proposal_id,
                    operation="generate cost matrix",
                )
                artifact = db.execute(
                    select(CostMatrixArtifact)
                    .where(CostMatrixArtifact.id == artifact_id)
                    .options(selectinload(CostMatrixArtifact.outputs))
                ).scalar_one_or_none()
                if proposal is None or artifact is None or artifact.proposal_id != proposal_id:
                    raise CostMatrixError("Cost matrix was not found for this proposal.")
                if artifact.status in {STATUS_NEEDS_CONFIRMATION, STATUS_DISMISSED}:
                    raise CostMatrixError(
                        "Confirm this workbook as a required cost matrix before generating it."
                    )
                readiness = _readiness(db, proposal, artifact)
                if not readiness["ready"]:
                    raise CostMatrixNotReadyError(readiness["blockers"])

                document = db.get(RfpPackageDocument, artifact.source_document_id)
                if document is None:
                    raise CostMatrixError("Source workbook record is missing.")
                source_path = require_contained_file(
                    document.storage_path,
                    root=RFP_PACKAGES_DIR,
                    expected_parent_name=str(proposal.rfp_package_id),
                    description="cost matrix source workbook",
                )
                source = source_path.read_bytes()
                if _sha256(source) != artifact.template_sha256:
                    raise CostMatrixError("Source workbook changed after inspection; generation stopped.")
                generated, changed_parts = _patch_workbook(
                    source,
                    targets=_target_map(artifact),
                    resolved=readiness["resolved_values"],
                )
                validation = _validate_generated_workbook(
                    source,
                    generated,
                    artifact=artifact,
                    resolved=readiness["resolved_values"],
                    changed_parts=changed_parts,
                )

                # Re-read all current sources at the persistence boundary. This
                # prevents a stale mapping/output from being certified if a cost
                # input changed during workbook construction.
                current_sources, current_basis = _source_catalog(db, proposal)
                current_resolved, current_blockers = _resolve_mapping_values(
                    artifact,
                    current_sources,
                )
                current_blockers.extend(_reconciliation_blockers(
                    artifact,
                    sources=current_sources,
                    resolved=current_resolved,
                ))
                current_blockers.extend(_workflow_blockers(current_basis))
                current_provenance = _generation_provenance(
                    artifact,
                    sources=current_sources,
                    basis=current_basis,
                    resolved=current_resolved,
                )
                if current_blockers or current_provenance["sha256"] != readiness["provenance"]["sha256"]:
                    raise CostMatrixError(
                        "Pricing or mapping changed during generation; no output was published."
                    )

                version = int(db.scalar(
                    select(func.max(CostMatrixOutput.version)).where(
                        CostMatrixOutput.artifact_id == artifact.id
                    )
                ) or 0) + 1
                output_dir = RFP_PACKAGES_DIR / str(proposal.rfp_package_id) / "cost_matrix_outputs"
                output_dir.mkdir(parents=True, exist_ok=True)
                base = _safe_filename(document.filename)
                filename = (
                    f"{Path(base).stem}_matrix{artifact.id}_completed_v{version}.xlsx"
                )
                destination = output_dir / filename
                if destination.exists():
                    raise CostMatrixError(
                        "A generated output already exists at the next immutable revision "
                        "path; generation stopped without overwriting it."
                    )
                temp_path = output_dir / f".{uuid4().hex}.generating"
                try:
                    temp_path.write_bytes(generated)
                    # A hard-link publication is atomic and fails if a file
                    # appeared at the destination. This preserves immutability
                    # even when another process writes into the package folder.
                    os.link(temp_path, destination)
                    published_path = destination
                    temp_path.unlink()
                finally:
                    if temp_path.exists():
                        temp_path.unlink()

                output = CostMatrixOutput(
                    artifact_id=artifact.id,
                    version=version,
                    pricing_scenario=current_basis.get("scenario"),
                    pricing_basis_sha256=current_provenance["sha256"],
                    generation_provenance_json=current_provenance,
                    mapping_snapshot_json=deepcopy(artifact.mapping_json or {}),
                    output_filename=filename,
                    output_storage_path=str(destination),
                    output_sha256=_sha256(generated),
                    validation_json=validation,
                    generated_at=datetime.now(UTC),
                )
                db.add(output)
                db.flush()
                artifact.status = STATUS_GENERATED
                artifact.last_error = None
                output_id = output.id
            published_path = None  # committed successfully
            return output_id
    except Exception:
        if published_path is not None:
            try:
                safe = require_contained_file(
                    published_path,
                    root=RFP_PACKAGES_DIR,
                    description="uncommitted generated cost matrix",
                )
                if safe.exists():
                    safe.unlink()
            except Exception:
                log.exception("failed to clean up uncommitted output %s", published_path)
        raise


def _artifact_snapshot(
    db: Session,
    proposal: Proposal,
    artifact: CostMatrixArtifact,
) -> dict[str, Any]:
    latest = _latest_output(artifact)
    if artifact.status in {STATUS_NEEDS_CONFIRMATION, STATUS_DISMISSED}:
        status = artifact.status
        readiness = {
            "ready": False,
            "blockers": (
                [
                    "Confirm this workbook as a required cost matrix or dismiss it "
                    "as an ordinary attachment."
                ]
                if artifact.status == STATUS_NEEDS_CONFIRMATION
                else []
            ),
            "basis": {},
            "sources": [],
            "has_output": latest is not None,
            "output_current": False,
            "suggestions": {},
        }
    else:
        readiness = _readiness(db, proposal, artifact)
        if latest is not None:
            status = STATUS_GENERATED if readiness["output_current"] else STATUS_STALE
        elif readiness["ready"]:
            status = STATUS_READY
        elif any("Map or explicitly skip" in blocker or "Choose how" in blocker for blocker in readiness["blockers"]):
            status = STATUS_MAPPING_REQUIRED
        else:
            status = STATUS_WAITING_FOR_COSTS
    return {
        "id": artifact.id,
        "proposal_id": artifact.proposal_id,
        "source_document_id": artifact.source_document_id,
        "filename": artifact.source_document.filename,
        "template_sha256": artifact.template_sha256,
        "analysis_version": artifact.analysis_version,
        "analysis": deepcopy(artifact.analysis_json or {}),
        "mapping": deepcopy(artifact.mapping_json or {}),
        "status": status,
        "stored_status": artifact.status,
        "last_error": artifact.last_error,
        "readiness": {
            key: deepcopy(value)
            for key, value in readiness.items()
            if key not in {"sources", "provenance", "resolved_values"}
        },
        "sources": deepcopy(readiness["sources"]),
        "latest_output": (
            {
                "id": latest.id,
                "version": latest.version,
                "filename": latest.output_filename,
                "sha256": latest.output_sha256,
                "pricing_scenario": latest.pricing_scenario,
                "generated_at": latest.generated_at.isoformat(),
                "current": readiness["output_current"],
                "validation": deepcopy(latest.validation_json or {}),
            }
            if latest is not None
            else None
        ),
        "output_history": [
            {
                "id": output.id,
                "version": output.version,
                "filename": output.output_filename,
                "sha256": output.output_sha256,
                "pricing_scenario": output.pricing_scenario,
                "generated_at": output.generated_at.isoformat(),
                "current": bool(
                    latest is not None
                    and output.id == latest.id
                    and readiness["output_current"]
                ),
            }
            for output in sorted(artifact.outputs, key=lambda row: row.version, reverse=True)
        ],
    }


def get_cost_matrix_snapshots(proposal_id: int) -> list[dict[str, Any]]:
    with session_scope() as db:
        proposal = db.get(Proposal, proposal_id)
        if proposal is None:
            return []
        artifacts = db.execute(
            select(CostMatrixArtifact)
            .where(CostMatrixArtifact.proposal_id == proposal_id)
            .options(
                selectinload(CostMatrixArtifact.source_document),
                selectinload(CostMatrixArtifact.outputs),
            )
            .order_by(CostMatrixArtifact.id)
        ).scalars().all()
        return [_artifact_snapshot(db, proposal, artifact) for artifact in artifacts]


def get_cost_matrix_download(output_id: int) -> tuple[bytes, str]:
    """Read an existing revision; allowed for archived proposals."""
    with session_scope() as db:
        output = db.get(CostMatrixOutput, output_id)
        if output is None:
            raise CostMatrixError("Generated cost matrix was not found.")
        path = require_contained_file(
            output.output_storage_path,
            root=RFP_PACKAGES_DIR,
            description="generated cost matrix",
        )
        if not path.is_file():
            raise CostMatrixError("Generated cost matrix file is missing.")
        data = path.read_bytes()
        if _sha256(data) != output.output_sha256:
            raise CostMatrixError("Generated cost matrix failed its integrity check.")
        return data, output.output_filename


def extract_cost_matrix_instruction_text(
    storage_path: str | Path,
    analysis: dict[str, Any],
    *,
    include_visible_context: bool = False,
) -> tuple[str, int]:
    """Extract visible buyer instructions without flattening price-entry rows.

    A workbook can be both a deliverable template and a source of binding
    instructions. Confirmed matrices therefore stay in intake, but only
    instruction-oriented visible text is supplied to requirements agents.
    Hidden sheets, mapped input rows, formulas, and numeric pricing values are
    excluded.
    """
    excluded_rows: dict[str, set[int]] = {}
    for target in analysis.get("targets") or []:
        if target.get("category") != "pricing":
            continue
        coordinate = str(target.get("cell") or "")
        row_match = re.search(r"\d+", coordinate)
        if row_match:
            excluded_rows.setdefault(str(target.get("sheet") or ""), set()).add(
                int(row_match.group(0))
            )
    for formula in analysis.get("formulas") or []:
        coordinate = str(formula.get("cell") or "")
        row_match = re.search(r"\d+", coordinate)
        if row_match and _TOTAL_LABEL.search(str(formula.get("label") or "")):
            excluded_rows.setdefault(str(formula.get("sheet") or ""), set()).add(
                int(row_match.group(0))
            )

    workbook = load_workbook(
        filename=str(storage_path),
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    pages: list[tuple[str, list[str]]] = []
    try:
        for ws in workbook.worksheets:
            if ws.sheet_state != "visible":
                continue
            instruction_sheet = bool(_INSTRUCTION_SHEET.search(ws.title))
            column_headers: dict[int, str] = {}
            lines: list[str] = []
            for row in ws.iter_rows():
                if row and _row_hidden(ws, row[0].row):
                    continue
                prior_column_headers = dict(column_headers)
                # Keep the latest label-like text for each column. Exact
                # numeric/currency strings do not replace it, so a sequence of
                # text-formatted amounts still inherits its real header.
                for cell in row:
                    if _column_hidden(ws, cell.column):
                        continue
                    value = cell.value
                    if (
                        isinstance(value, str)
                        and value.strip()
                        and not _is_formula(value)
                        and not _value_like_text(value)
                    ):
                        column_headers[cell.column] = value.strip()
                excluded_pricing_row = bool(
                    row and row[0].row in excluded_rows.get(ws.title, set())
                )
                raw_string_values = [
                    value.strip()
                    for cell in row
                    if not _column_hidden(ws, cell.column)
                    if isinstance((value := cell.value), str)
                    and value.strip()
                    and not _is_formula(value)
                ]
                if not raw_string_values:
                    continue
                raw_rendered = " | ".join(raw_string_values)
                financial_row_context = _price_context(raw_rendered)
                string_values = [
                    redacted
                    for value in raw_string_values
                    if (
                        redacted := _redact_financial_text(
                            value,
                            financial_context=financial_row_context,
                        )
                    )
                    and not (
                        financial_row_context
                        and _PLAIN_NUMERIC_TEXT.fullmatch(redacted)
                    )
                ]
                if not string_values:
                    continue
                rendered = " | ".join(string_values)
                derived_context = bool(
                    _REQUIREMENT_TEXT.search(rendered)
                    or _SECTION_M_TEXT.search(rendered)
                    or _COTS_TEXT.search(rendered)
                )
                evaluation_numeric_context = bool(
                    _EVALUATION_NUMBER_TEXT.search(raw_rendered)
                )
                # Build a single safe representation for both confirmed and
                # candidate workbooks. This preserves meaningful numeric
                # requirements (for example, SLA percentages and evaluation
                # weights) while suppressing currency-formatted cells and
                # values that sit in an actual price-entry context.
                include_numbers = instruction_sheet or derived_context
                safe_values: list[str] = []
                for cell in row:
                    if _column_hidden(ws, cell.column):
                        continue
                    value = cell.value
                    if value is None or _is_formula(value):
                        continue
                    if isinstance(value, str):
                        header = prior_column_headers.get(cell.column, "")
                        header_financial_context = _price_context(
                            header or "",
                            header=True,
                        )
                        safe_evaluation_scalar = bool(
                            evaluation_numeric_context
                            and _PLAIN_NUMERIC_TEXT.fullmatch(value)
                            and _safe_evaluation_scalar(value)
                            and not header_financial_context
                        )
                        cell_financial_context = bool(
                            header_financial_context
                            or (
                                financial_row_context
                                and not safe_evaluation_scalar
                            )
                        )
                        redacted = _redact_financial_text(
                            value,
                            financial_context=cell_financial_context,
                        )
                        if redacted and not (
                            cell_financial_context
                            and _PLAIN_NUMERIC_TEXT.fullmatch(redacted)
                        ):
                            safe_values.append(redacted)
                        continue
                    if not include_numbers or _currency_format(cell.number_format):
                        continue
                    if isinstance(value, (date, datetime, bool)):
                        safe_values.append(str(_json_safe_value(value)))
                        continue
                    header = prior_column_headers.get(cell.column, "")
                    header_financial_context = _price_context(
                        header or "",
                        header=True,
                    )
                    if header_financial_context or (
                        financial_row_context
                        and not (
                            evaluation_numeric_context
                            and _safe_evaluation_scalar(value)
                        )
                    ):
                        continue
                    safe_values.append(str(_json_safe_value(value)))
                if include_visible_context:
                    # Possible/dismissed matrices remain useful intake sources,
                    # but never expose hidden sheets or uncontextualized numeric
                    # price grids. Numeric values are retained only on rows that
                    # contain requirement/evaluation/COTS language (or on an
                    # explicit Instructions sheet), preserving SLA thresholds
                    # and stated weights without feeding raw prices to agents.
                    if safe_values:
                        lines.append(" | ".join(safe_values))
                    continue
                if (
                    (instruction_sheet and not excluded_pricing_row)
                    or derived_context
                ) and safe_values:
                    lines.append(" | ".join(safe_values))
            if lines:
                pages.append((ws.title, lines))
    finally:
        workbook.close()

    rendered_pages: list[str] = []
    for page_number, (sheet_name, lines) in enumerate(pages, start=1):
        rendered_pages.extend([
            f"--- Page {page_number} ---",
            f"[Sheet: {sheet_name}]",
            *lines,
        ])
    return "\n".join(rendered_pages), len(pages)


def get_cost_matrix_requirements_context(proposal_id: int) -> list[dict[str, Any]]:
    """Buyer cost-line labels supplied to the Cost Analyst as scope context."""
    with session_scope() as db:
        artifacts = db.execute(
            select(CostMatrixArtifact)
            .where(
                CostMatrixArtifact.proposal_id == proposal_id,
                CostMatrixArtifact.status.notin_([
                    STATUS_NEEDS_CONFIRMATION,
                    STATUS_DISMISSED,
                ]),
            )
            .options(selectinload(CostMatrixArtifact.source_document))
            .order_by(CostMatrixArtifact.id)
        ).scalars().all()
        return [
            {
                "artifact_id": artifact.id,
                "filename": artifact.source_document.filename,
                "targets": [
                    {
                        "sheet": target.get("sheet"),
                        "cell": target.get("cell"),
                        "label": target.get("label"),
                        "header": target.get("header"),
                        "kind": target.get("kind"),
                    }
                    for target in (artifact.analysis_json or {}).get("targets", [])
                    if target.get("category") == "pricing"
                ],
            }
            for artifact in artifacts
        ]


def is_cost_matrix_document(db: Session, document_id: int) -> bool:
    return bool(db.scalar(
        select(CostMatrixArtifact.id).where(
            CostMatrixArtifact.source_document_id == document_id,
            CostMatrixArtifact.status.notin_([
                STATUS_NEEDS_CONFIRMATION,
                STATUS_DISMISSED,
            ]),
        )
    ))


def cost_matrix_submission_check(
    proposal_id: int,
    *,
    db: Session | None = None,
) -> dict[str, Any]:
    """Return whether every attached matrix has a current generated revision."""
    if db is None:
        with session_scope() as owned_db:
            return cost_matrix_submission_check(proposal_id, db=owned_db)
    proposal = db.get(Proposal, proposal_id)
    if proposal is None:
        return {"verified": False, "count": 0, "current": 0, "detail": "Proposal not found"}
    artifacts = db.execute(
        select(CostMatrixArtifact)
        .where(CostMatrixArtifact.proposal_id == proposal_id)
        .options(
            selectinload(CostMatrixArtifact.source_document),
            selectinload(CostMatrixArtifact.outputs),
        )
    ).scalars().all()
    pending = [
        artifact for artifact in artifacts
        if artifact.status == STATUS_NEEDS_CONFIRMATION
    ]
    active = [
        artifact for artifact in artifacts
        if artifact.status not in {STATUS_NEEDS_CONFIRMATION, STATUS_DISMISSED}
    ]
    if not active and not pending:
        return {
            "verified": True,
            "count": 0,
            "current": 0,
            "detail": "No cost matrix was supplied",
        }
    current = sum(
        1 for artifact in active
        if _readiness(db, proposal, artifact)["output_current"]
    )
    if pending:
        detail = (
            f"{len(pending)} possible cost matrix workbook(s) need confirmation or "
            "dismissal before submission"
        )
    else:
        detail = (
            f"{current} of {len(active)} cost matrix workbook(s) generated and current"
        )
    return {
        "verified": not pending and current == len(active),
        "count": len(active),
        "pending": len(pending),
        "current": current,
        "detail": detail,
    }


__all__ = [
    "ANALYSIS_VERSION",
    "COST_MATRIX_INTAKE_POLICY_VERSION",
    "COST_MATRIX_ROLE",
    "RECONCILIATION_MAPPING_KEY",
    "CostMatrixError",
    "CostMatrixNotReadyError",
    "add_cost_matrix_target",
    "attach_cost_matrix",
    "confirm_cost_matrix",
    "cost_matrix_submission_check",
    "dismiss_cost_matrix",
    "extract_cost_matrix_instruction_text",
    "generate_cost_matrix",
    "get_cost_matrix_download",
    "get_cost_matrix_requirements_context",
    "get_cost_matrix_snapshots",
    "get_cost_source_catalog",
    "inspect_cost_matrix",
    "is_cost_matrix_document",
    "register_original_cost_matrices",
    "save_cost_matrix_mapping",
    "try_inspect_cost_matrix",
]
