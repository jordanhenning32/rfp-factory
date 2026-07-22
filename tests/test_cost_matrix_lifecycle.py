from __future__ import annotations

import io
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _strong_matrix_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buyer Pricing"
    sheet["A1"] = "Work item"
    sheet["B1"] = "Total cost"
    sheet["A2"] = "Workstream Alpha"
    sheet["B2"] = 0
    sheet["B2"].number_format = "$#,##0.00"
    sheet["A3"] = "Workstream Beta"
    sheet["B3"] = 0
    sheet["B3"].number_format = "$#,##0.00"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _ambiguous_pricing_instructions_bytes() -> bytes:
    """A price-related workbook that must be noticed, not auto-confirmed."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Instructions"
    sheet["A1"] = (
        "Offerors shall explain all pricing assumptions and include travel "
        "in the proposed total."
    )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _ambiguous_pricing_context_bytes() -> bytes:
    """Possible matrix whose pricing row also carries derived context."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing Schedule"
    sheet["A1"] = "Pricing requirement"
    sheet["B1"] = "Amount"
    sheet["A2"] = (
        "Offeror shall complete this line for a COTS solution; the technical "
        "approach is an evaluation factor."
    )
    sheet["B2"] = 123456
    sheet["B2"].number_format = "General"
    sheet["A3"] = "Offeror must provide price"
    sheet["B3"] = "USD 123,456.00"
    sheet["A4"] = "Project Manager"
    sheet["B4"] = "123456.00"
    sheet["A5"] = "Offeror shall enter the proposed value"
    sheet["B5"] = "123456.00"
    sheet["A6"] = "Offeror shall enter the proposed euro value"
    sheet["B6"] = "\u20ac123,456"
    sheet["A7"] = "Price will be evaluated for reasonableness."
    sheet["B7"] = "$100K-$200K"
    sheet["A8"] = "Award will be made to the responsible offeror."
    sheet["B8"] = "1.5M USD"
    sheet["A9"] = "Technical merit is more important than price."
    sheet["B9"] = "1.23456E5"
    sheet["A10"] = "Past performance will be assessed."
    sheet["B10"] = "$1.2M"
    sheet["A11"] = "The contractor will maintain all records."
    sheet["A12"] = "It is mandatory that all resumes are signed."
    sheet["A13"] = "Offerors may not alter this worksheet."
    sheet["A14"] = "Review and update minimum cybersecurity requirements"
    sheet["B14"] = 125000
    sheet["A15"] = "Establish Compliance Framework"
    sheet["B15"] = 75000

    requirements = workbook.create_sheet("Requirements")
    requirements.append(["Requirement", "Value"])
    requirements.append(["Uptime rate must be achieved", 99.9])
    requirements.append(["Response rate shall be achieved", 95])
    requirements.append(["Support unit is required", 15])
    requirements.append(["Total system availability is required", 99.9])
    requirements.append(["Technical approach weighted", 25])
    requirements.append(["Offeror must respond within minutes", 15])
    hidden = workbook.create_sheet("Internal Hidden")
    hidden["A1"] = "Offeror shall disclose the hidden COTS evaluation note."
    hidden.sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _mixed_matrix_with_instructions_bytes() -> bytes:
    """A confirmed matrix that also contains binding buyer instructions."""
    workbook = Workbook()
    pricing = workbook.active
    pricing.title = "Buyer Pricing"
    pricing.append(["Labor category", "Hours", "Hourly rate", "Total price"])
    pricing.append(["Project Manager", 0, 0, "=B2*C2"])
    pricing["C2"].number_format = "$#,##0.00"
    pricing["D2"].number_format = "$#,##0.00"
    pricing.append(["Grand Total", None, None, "=SUM(D2:D2)"])
    pricing["D3"].number_format = "$#,##0.00"

    instructions = workbook.create_sheet("Instructions and Terms")
    instructions["A1"] = (
        "Offerors shall propose a commercial off-the-shelf (COTS) solution."
    )
    instructions["A2"] = (
        "The technical approach will be evaluated for implementation risk."
    )
    instructions["A3"] = "Do not alter formulas or worksheet structure."

    hidden = workbook.create_sheet("Hidden Instructions")
    hidden["A1"] = "Offerors shall disclose the hidden internal note."
    hidden.sheet_state = "hidden"

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _percentage_matrix_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buyer Pricing"
    sheet["A1"] = "Pricing factor"
    sheet["B1"] = "Percentage"
    sheet["A2"] = "Markup %"
    sheet["B2"] = 0
    sheet["B2"].number_format = "0.00%"
    sheet["A3"] = "Target margin %"
    sheet["B3"] = 0
    sheet["B3"].number_format = "0.00%"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.fixture()
def lifecycle_env(inmemory_db, monkeypatch, tmp_path):
    import app.db.session as db_session
    import app.jobs.intake as intake
    import app.services.cost_matrix as cost_matrix
    import app.services.proposals as proposals

    package_root = tmp_path / "rfp_packages"
    package_root.mkdir()
    monkeypatch.setattr(cost_matrix, "RFP_PACKAGES_DIR", package_root)
    monkeypatch.setattr(proposals, "RFP_PACKAGES_DIR", package_root)
    monkeypatch.setattr(cost_matrix, "session_scope", db_session.session_scope)
    monkeypatch.setattr(intake, "session_scope", db_session.session_scope)
    return {
        "db": db_session,
        "root": package_root,
        "service": cost_matrix,
        "proposals": proposals,
        "intake": intake,
    }


def _mock_successful_late_requirements_review(env, monkeypatch) -> None:
    """Exercise the real late-attach orchestration without paid model calls."""

    from app.agents.compliance_completeness import ComplianceCompletenessReport
    from app.agents.compliance_matrix import (
        ComplianceExtractionResult,
        ExtractedComplianceItem,
    )
    from app.agents.compliance_validator import ComplianceValidationReport

    monkeypatch.setattr(
        env["intake"],
        "extract_compliance_items",
        lambda **_kwargs: ComplianceExtractionResult(
            items=[
                ExtractedComplianceItem(
                    requirement_id="REQ-001",
                    requirement_text=(
                        "Offeror shall complete this line for a COTS solution; "
                        "the technical approach is an evaluation factor."
                    ),
                    requirement_type="shall",
                    category="pricing",
                    source_section="Pricing Schedule",
                    source_page=1,
                )
            ],
            coverage_state="complete",
            source_chunks_total=1,
            source_chunks_completed=1,
        ),
    )
    monkeypatch.setattr(
        env["intake"],
        "audit_compliance_completeness",
        lambda **_kwargs: ComplianceCompletenessReport(
            source_units_total=1,
            primary_model="gemini-review-test",
            fallback_model="haiku-fallback-test",
            source_sha256="source",
            matrix_sha256="matrix",
            reviewed_unit_labels=["page 1"],
        ),
    )

    def clean_validation(items, _proposal_id, **_kwargs):
        return ComplianceValidationReport(
            total_count=len(items),
            primary_model="gemini-review-test",
            fallback_model="haiku-fallback-test",
            reviewed_requirement_ids=[item.requirement_id for item in items],
        )

    monkeypatch.setattr(
        env["intake"],
        "_validate_and_apply_corrections",
        clean_validation,
    )


def _create_proposal(env, *, matrix_name: str, matrix_content: bytes) -> tuple[int, int]:
    from app.models import CostMatrixArtifact

    with env["db"].SessionLocal() as db:
        proposal = env["proposals"].create_proposal_with_files(
            db,
            title="Template lifecycle regression",
            files=[
                env["proposals"].UploadedFile("solicitation.pdf", b"%PDF fake"),
                env["proposals"].UploadedFile(matrix_name, matrix_content),
            ],
        )
        db.commit()
        proposal_id = proposal.id
        artifact_id = (
            db.query(CostMatrixArtifact)
            .filter_by(proposal_id=proposal_id)
            .one()
            .id
        )
    return proposal_id, artifact_id


def _create_reviewed_it_matrix(
    env,
    monkeypatch,
    *,
    matrix_content: bytes | None = None,
    matrix_name: str = "Buyer Cost Matrix.xlsx",
) -> tuple[int, int]:
    from app.models import PricingPackage, Proposal

    proposal_id, artifact_id = _create_proposal(
        env,
        matrix_name=matrix_name,
        matrix_content=matrix_content or _strong_matrix_bytes(),
    )
    monkeypatch.setattr(
        env["service"],
        "get_cost_review_freshness",
        lambda db, proposal_id, scenario: {
            "verified": True,
            "legacy": False,
            "review_count": 1,
            "detail": "Current reviewed cost build",
        },
    )
    with env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        proposal.proposed_scenario = "MEDIUM"
        proposal.service_line = "it_services"
        db.add(
            PricingPackage(
                proposal_id=proposal_id,
                scenario="MEDIUM",
                loaded_labor_cost=20,
                odcs_json=[],
                subcontractor_costs=0,
                indirect_costs_json={},
                total_proposed_price=30,
                pnl_projection_json={},
                phase_breakdown_json=[],
            )
        )
        db.commit()
    return proposal_id, artifact_id


def _map_two_financial_targets(env, proposal_id: int, artifact_id: int) -> None:
    matrix = env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    target_ids = [
        target["id"]
        for target in matrix["analysis"]["targets"]
        if target["category"] == "pricing"
    ]
    assert len(target_ids) == 2
    env["service"].save_cost_matrix_mapping(
        proposal_id,
        artifact_id,
        {
            target_ids[0]: {"mode": "manual", "value": 10},
            target_ids[1]: {"mode": "manual", "value": 20},
        },
    )


def test_possible_matrix_stays_in_intake_until_confirmed(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Instructions.xlsx",
        matrix_content=_ambiguous_pricing_instructions_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert artifact.status == "needs_confirmation"
        assert document.document_role != "cost_matrix"
        assert not lifecycle_env["service"].is_cost_matrix_document(
            db, document.id
        )

    candidate_check = lifecycle_env["service"].cost_matrix_submission_check(
        proposal_id
    )
    assert candidate_check["verified"] is False
    assert "confirmation" in candidate_check["detail"].lower()
    from app.services.submission_commitments import (
        compute_system_verified_items,
        evaluate_submission_readiness,
    )

    with lifecycle_env["db"].SessionLocal() as db:
        system_items = compute_system_verified_items(proposal_id, db=db)
    candidate_items = [
        item for item in system_items
        if item["key"] == "cost_matrices_current"
    ]
    assert len(candidate_items) == 1
    assert candidate_items[0]["verified"] is False
    with lifecycle_env["db"].SessionLocal() as db:
        submission = evaluate_submission_readiness(proposal_id, db=db)
    assert submission["ready"] is False
    assert any(
        "confirmation" in blocker.lower()
        for blocker in submission["blockers"]
    )
    assert (
        lifecycle_env["service"].get_cost_matrix_requirements_context(
            proposal_id
        )
        == []
    )

    extracted: list[str] = []

    def fake_extract(path: str, filename: str):
        extracted.append(filename)
        return "--- Page 1 ---\nBuyer-authored source content", 1

    monkeypatch.setattr(
        lifecycle_env["intake"], "_extract_text_for_intake", fake_extract
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    assert set(extracted) == {
        "solicitation.pdf",
    }
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert "Offerors shall explain all pricing assumptions" in (
            document.extracted_text_md or ""
        )

    lifecycle_env["service"].confirm_cost_matrix(proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert artifact.status == "mapping_required"
        assert document.document_role == "cost_matrix"
        assert lifecycle_env["service"].is_cost_matrix_document(
            db, document.id
        )
    check = lifecycle_env["service"].cost_matrix_submission_check(proposal_id)
    assert check["count"] == 1
    assert check["verified"] is False
    assert len(
        lifecycle_env["service"].get_cost_matrix_requirements_context(
            proposal_id
        )
    ) == 1


def test_confirmed_matrix_instructions_reach_intake_consumers_without_price_rows(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Buyer Cost Matrix with Instructions.xlsx",
        matrix_content=_mixed_matrix_with_instructions_bytes(),
    )
    ordinary_extracts: list[str] = []

    def fake_extract(path: str, filename: str):
        ordinary_extracts.append(filename)
        return "--- Page 1 ---\nThe contractor shall deliver the base scope.", 1

    monkeypatch.setattr(
        lifecycle_env["intake"], "_extract_text_for_intake", fake_extract
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    assert ordinary_extracts == ["solicitation.pdf"]

    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        matrix_document = db.get(
            RfpPackageDocument, artifact.source_document_id
        )
        assert artifact.status == "mapping_required"
        assert matrix_document.document_role == "cost_matrix"
        assert lifecycle_env["service"].is_cost_matrix_document(
            db, matrix_document.id
        )
        instruction_text = matrix_document.extracted_text_md or ""

    assert "commercial off-the-shelf (COTS)" in instruction_text
    assert "technical approach will be evaluated" in instruction_text
    assert "Do not alter formulas" in instruction_text
    assert "Project Manager" not in instruction_text
    assert "Grand Total" not in instruction_text
    assert "hidden internal note" not in instruction_text

    lifecycle_env["intake"]._detect_cots_orientation(proposal_id)
    with lifecycle_env["db"].SessionLocal() as db:
        assert db.get(Proposal, proposal_id).cots_orientation is True

    compliance_inputs: dict[str, str] = {}

    def fake_extract_one_doc(_proposal_id: int, doc: dict):
        compliance_inputs[doc["filename"]] = doc["text"]
        return doc["filename"], []

    monkeypatch.setattr(
        lifecycle_env["intake"],
        "_extract_one_doc_for_matrix",
        fake_extract_one_doc,
    )
    assert lifecycle_env["intake"]._run_compliance_matrix(proposal_id) == 0
    assert (
        "commercial off-the-shelf (COTS)"
        in compliance_inputs["Buyer Cost Matrix with Instructions.xlsx"]
    )

    section_m_inputs: dict[str, str] = {}

    class FakeEvaluationResult:
        factors = ["technical"]

        def as_dict(self):
            return {"factors": []}

    def fake_extract_evaluation_criteria(**kwargs):
        section_m_inputs["text"] = kwargs["document_text"]
        return FakeEvaluationResult()

    monkeypatch.setattr(
        lifecycle_env["intake"],
        "extract_evaluation_criteria",
        fake_extract_evaluation_criteria,
    )
    assert lifecycle_env["intake"]._run_section_m_extractor(proposal_id) == 1
    assert "technical approach will be evaluated" in section_m_inputs["text"]

    context = lifecycle_env["service"].get_cost_matrix_requirements_context(
        proposal_id
    )
    assert len(context) == 1
    assert context[0]["filename"] == "Buyer Cost Matrix with Instructions.xlsx"


def test_confirming_after_ordinary_intake_replaces_full_workbook_extraction(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Instructions.xlsx",
        matrix_content=_ambiguous_pricing_instructions_bytes(),
    )

    def fake_extract(path: str, filename: str):
        if filename.endswith(".xlsx"):
            return (
                "--- Page 1 ---\nRAW PRICING GRID | Project Manager | $123,456\n"
                "Offerors shall explain all pricing assumptions.",
                1,
            )
        return "--- Page 1 ---\nThe contractor shall deliver the base scope.", 1

    monkeypatch.setattr(
        lifecycle_env["intake"], "_extract_text_for_intake", fake_extract
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2

    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        candidate_text = document.extracted_text_md or ""
        assert "Offerors shall explain all pricing assumptions" in candidate_text
        assert "RAW PRICING GRID" not in candidate_text

    lifecycle_env["service"].confirm_cost_matrix(proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        structured_text = document.extracted_text_md or ""
    assert "Offerors shall explain all pricing assumptions" in structured_text
    assert "RAW PRICING GRID" not in structured_text
    assert "Project Manager" not in structured_text
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2


def test_confirm_preserves_derived_context_and_recomputes_cots(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Context.xlsx",
        matrix_content=_ambiguous_pricing_context_bytes(),
    )
    monkeypatch.setattr(
        lifecycle_env["intake"],
        "_extract_text_for_intake",
        lambda path, filename: (
            "--- Page 1 ---\nThe contractor shall deliver the base scope.",
            1,
        ),
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        assert artifact.status == "needs_confirmation"
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        candidate_text = document.extracted_text_md or ""
        assert "Offeror shall complete this line" in candidate_text
        assert "COTS solution" in candidate_text
        assert "evaluation factor" in candidate_text
        assert "minimum cybersecurity requirements" in candidate_text
        assert "Establish Compliance Framework" in candidate_text
        assert "123456" not in candidate_text
        assert "123,456" not in candidate_text
        assert "125000" not in candidate_text
        assert "75000" not in candidate_text
        assert "100K" not in candidate_text
        assert "200K" not in candidate_text
        assert "1.5M" not in candidate_text
        assert "1.23456E5" not in candidate_text
        assert "1.2M" not in candidate_text
        assert "Uptime rate must be achieved | 99.9" in candidate_text
        assert "Response rate shall be achieved | 95" in candidate_text
        assert "Support unit is required | 15" in candidate_text
        assert "Total system availability is required | 99.9" in candidate_text
        assert "Technical approach weighted | 25" in candidate_text
        assert "Offeror must respond within minutes | 15" in candidate_text
        assert "hidden COTS evaluation note" not in candidate_text
        proposal = db.get(Proposal, proposal_id)
        proposal.cots_orientation = False
        db.commit()

    lifecycle_env["service"].confirm_cost_matrix(proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        proposal = db.get(Proposal, proposal_id)
        structured_text = document.extracted_text_md or ""
        assert proposal.cots_orientation is True
        assert (
            artifact.analysis_json["role_transition"]["decision"]
            == "confirmed"
        )

    assert "Offeror shall complete this line" in structured_text
    assert "COTS solution" in structured_text
    assert "evaluation factor" in structured_text
    assert "evaluated for reasonableness" in structured_text
    assert "Award will be made" in structured_text
    assert "Technical merit is more important" in structured_text
    assert "Past performance will be assessed" in structured_text
    assert "contractor will maintain" in structured_text
    assert "mandatory that all resumes" in structured_text
    assert "Offerors may not alter" in structured_text
    assert "minimum cybersecurity requirements" in structured_text
    assert "Establish Compliance Framework" in structured_text
    assert "123456" not in structured_text
    assert "123,456" not in structured_text
    assert "125000" not in structured_text
    assert "75000" not in structured_text
    assert "100K" not in structured_text
    assert "200K" not in structured_text
    assert "1.5M" not in structured_text
    assert "1.23456E5" not in structured_text
    assert "1.2M" not in structured_text
    assert "Uptime rate must be achieved | 99.9" in structured_text
    assert "Response rate shall be achieved | 95" in structured_text
    assert "Support unit is required | 15" in structured_text
    assert "Total system availability is required | 99.9" in structured_text
    assert "Technical approach weighted | 25" in structured_text
    assert "Offeror must respond within minutes | 15" in structured_text
    assert "hidden COTS evaluation note" not in structured_text


def test_cached_candidate_text_is_upgraded_to_current_safe_policy(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Context.xlsx",
        matrix_content=_ambiguous_pricing_context_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        document.extracted_text_md = (
            "--- Page 1 ---\nRAW PRICE 123456\n"
            "Offeror shall disclose the hidden COTS evaluation note."
        )
        document.page_count = 1
        structure = dict(document.structure_json or {})
        # Simulate a cache stamped by the same workbook analyzer before the
        # intake redaction policy was versioned.
        structure["cost_matrix_intake_analysis_version"] = "cost-matrix-v3"
        structure["cost_matrix_intake_mode"] = "visible_context"
        structure.pop("cost_matrix_intake_policy_version", None)
        document.structure_json = structure
        db.commit()

    monkeypatch.setattr(
        lifecycle_env["intake"],
        "_extract_text_for_intake",
        lambda path, filename: (
            "--- Page 1 ---\nThe contractor shall deliver the base scope.",
            1,
        ),
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        refreshed = document.extracted_text_md or ""
        structure = document.structure_json or {}
    assert "RAW PRICE" not in refreshed
    assert "123456" not in refreshed
    assert "hidden COTS evaluation note" not in refreshed
    assert "Offeror shall complete this line" in refreshed
    assert structure["cost_matrix_intake_analysis_version"] == "cost-matrix-v3"
    assert structure["cost_matrix_intake_policy_version"] == "cost-matrix-intake-v2"
    assert structure["cost_matrix_intake_mode"] == "visible_context"


def test_late_attachment_immediately_extracts_safe_context_and_recomputes_cots(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.core.enums import (
        ComplianceStatus,
        ProposalStatus,
        RequirementCategory,
        RequirementType,
    )
    from app.models import (
        ComplianceMatrixItem,
        CostMatrixArtifact,
        Proposal,
        RfpPackageDocument,
    )

    _mock_successful_late_requirements_review(lifecycle_env, monkeypatch)

    with lifecycle_env["db"].SessionLocal() as db:
        proposal = lifecycle_env["proposals"].create_proposal_with_files(
            db,
            title="Late price schedule context",
            files=[
                lifecycle_env["proposals"].UploadedFile(
                    "solicitation.pdf",
                    b"%PDF fake",
                ),
            ],
        )
        proposal.cots_orientation = False
        proposal.status = ProposalStatus.DRAFTING
        solicitation = next(
            document
            for document in proposal.rfp_package.documents
            if document.filename == "solicitation.pdf"
        )
        db.add(
            ComplianceMatrixItem(
                proposal_id=proposal.id,
                requirement_id="REQ-007",
                requirement_text="The contractor must perform the base scope.",
                source_doc=solicitation.filename,
                source_document_id=solicitation.id,
                source_page=1,
                requirement_type=RequirementType.MUST,
                category=RequirementCategory.TECHNICAL,
                compliance_status=ComplianceStatus.TO_BE_DRAFTED,
            )
        )
        db.commit()
        proposal_id = proposal.id

    artifact_id = lifecycle_env["service"].attach_cost_matrix(
        proposal_id,
        filename="Late Pricing Schedule Context.xlsx",
        content=_ambiguous_pricing_context_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        proposal = db.get(Proposal, proposal_id)
        extracted = document.extracted_text_md or ""
        structure = document.structure_json or {}
        late_items = (
            db.query(ComplianceMatrixItem)
            .filter(ComplianceMatrixItem.source_document_id == document.id)
            .all()
        )
        assert artifact.status == "mapping_required"
        assert document.document_role == "cost_matrix"
        assert proposal.cots_orientation is True
        assert [item.requirement_id for item in late_items] == ["REQ-008"]
        review = structure["requirements_review"]
        assert review["status"] == "complete"
        assert review["requires_manual_review"] is False
        assert review["extraction"]["final_item_count"] == 1
        assert review["classification"]["state"] == "complete"
        assert review["completeness"]["state"] == "complete"
        late_document_id = document.id
    assert "Offeror shall complete this line" in extracted
    assert "evaluation factor" in extracted
    assert "123456" not in extracted
    assert "123,456" not in extracted
    assert "hidden COTS evaluation note" not in extracted
    assert structure["cost_matrix_intake_policy_version"] == "cost-matrix-intake-v2"
    assert structure["cost_matrix_intake_mode"] == "instructions"

    # Retrying an unsafe durable state re-runs review but reconciles the same
    # source text back to its stable row instead of appending a duplicate.
    with lifecycle_env["db"].SessionLocal() as db:
        document = db.get(RfpPackageDocument, late_document_id)
        retry_structure = dict(document.structure_json or {})
        retry_review = dict(retry_structure["requirements_review"])
        retry_review.update({"status": "failed", "requires_manual_review": True})
        retry_structure["requirements_review"] = retry_review
        document.structure_json = retry_structure
        db.add(
            ComplianceMatrixItem(
                proposal_id=proposal_id,
                requirement_id="REQ-009",
                requirement_text="Stale requirement from an incomplete extraction.",
                source_doc=document.filename,
                source_document_id=late_document_id,
                source_page=1,
                requirement_type=RequirementType.SHOULD,
                category=RequirementCategory.ADMINISTRATIVE,
                compliance_status=ComplianceStatus.TO_BE_DRAFTED,
            )
        )
        db.commit()

    assert (
        lifecycle_env["intake"].review_late_attached_requirements(
            proposal_id,
            late_document_id,
        )
        == 1
    )
    with lifecycle_env["db"].SessionLocal() as db:
        late_items = (
            db.query(ComplianceMatrixItem)
            .filter(ComplianceMatrixItem.source_document_id == late_document_id)
            .all()
        )
        document = db.get(RfpPackageDocument, late_document_id)
        assert [item.requirement_id for item in late_items] == ["REQ-008"]
        assert [item.requirement_text for item in late_items] == [
            "Offeror shall complete this line for a COTS solution; "
            "the technical approach is an evaluation factor."
        ]
        assert document.structure_json["requirements_review"]["status"] == "complete"


@pytest.mark.parametrize(
    ("initial_status", "expected_status", "gate_reopened"),
    [
        ("intaking", "intaking", False),
        ("drafting", "awaiting_scope_signoff", True),
    ],
)
def test_late_matrix_instructions_reopen_scope_only_after_intake(
    lifecycle_env,
    monkeypatch,
    initial_status,
    expected_status,
    gate_reopened,
) -> None:
    from app.core.enums import ProposalStatus
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    # Leave the committed queued state in place so this test observes the
    # upload boundary before a background/provider review can advance it.
    dispatch_calls: list[tuple[int, int]] = []
    monkeypatch.setattr(
        lifecycle_env["service"],
        "_review_late_attached_cost_matrix",
        lambda pid, did: dispatch_calls.append((pid, did)),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        proposal = lifecycle_env["proposals"].create_proposal_with_files(
            db,
            title="Late matrix scope gate",
            files=[
                lifecycle_env["proposals"].UploadedFile(
                    "solicitation.pdf",
                    b"%PDF fake",
                ),
            ],
        )
        proposal.status = ProposalStatus(initial_status)
        db.commit()
        proposal_id = proposal.id

    artifact_id = lifecycle_env["service"].attach_cost_matrix(
        proposal_id,
        filename="Late Pricing Instructions.xlsx",
        content=_ambiguous_pricing_context_bytes(),
    )

    with lifecycle_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        review = document.structure_json["requirements_review"]
        status = (
            proposal.status.value
            if hasattr(proposal.status, "value")
            else str(proposal.status)
        )
        assert status == expected_status
        assert review["status"] == "pending"
        assert review["scope_gate_reopened"] is gate_reopened
        assert len(dispatch_calls) == (0 if initial_status == "intaking" else 1)


def test_late_matrix_review_failure_is_durable_and_blocks_scope_signoff(
    lifecycle_env,
    monkeypatch,
) -> None:
    import app.services.workflow as workflow
    from app.core.enums import (
        ComplianceStatus,
        ProposalStatus,
        RequirementCategory,
        RequirementType,
    )
    from app.models import (
        ComplianceMatrixItem,
        CostMatrixArtifact,
        Proposal,
        RfpPackageDocument,
    )

    monkeypatch.setattr(workflow, "session_scope", lifecycle_env["db"].session_scope)

    def failed_extraction(**_kwargs):
        raise RuntimeError("provider unavailable")

    monkeypatch.setattr(
        lifecycle_env["intake"],
        "extract_compliance_items",
        failed_extraction,
    )

    with lifecycle_env["db"].SessionLocal() as db:
        proposal = lifecycle_env["proposals"].create_proposal_with_files(
            db,
            title="Late matrix review failure",
            files=[
                lifecycle_env["proposals"].UploadedFile(
                    "solicitation.pdf",
                    b"%PDF fake",
                ),
            ],
        )
        proposal.status = ProposalStatus.DRAFTING
        solicitation = next(
            document
            for document in proposal.rfp_package.documents
            if document.filename == "solicitation.pdf"
        )
        db.add(
            ComplianceMatrixItem(
                proposal_id=proposal.id,
                requirement_id="REQ-001",
                requirement_text="The contractor must perform the base scope.",
                source_doc=solicitation.filename,
                source_document_id=solicitation.id,
                source_page=1,
                requirement_type=RequirementType.MUST,
                category=RequirementCategory.TECHNICAL,
                compliance_status=ComplianceStatus.TO_BE_DRAFTED,
            )
        )
        db.commit()
        proposal_id = proposal.id

    artifact_id = lifecycle_env["service"].attach_cost_matrix(
        proposal_id,
        filename="Late Pricing Instructions.xlsx",
        content=_ambiguous_pricing_context_bytes(),
    )

    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        review = document.structure_json["requirements_review"]
        assert review["status"] == "failed"
        assert review["requires_manual_review"] is True
        assert "Retry the review before scope sign-off" in review["reason"]
        assert (
            db.query(ComplianceMatrixItem)
            .filter(ComplianceMatrixItem.source_document_id == document.id)
            .count()
            == 0
        )
        db.get(Proposal, proposal_id).status = ProposalStatus.AWAITING_SCOPE_SIGNOFF
        db.commit()

    result = workflow.sign_off_scope(proposal_id)
    assert result["ok"] is False
    assert any(
        "Late Pricing Instructions.xlsx (failed)" in blocker
        for blocker in result["blockers"]
    )


def test_dismissing_after_structured_intake_restores_ordinary_workbook_extraction(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Buyer Cost Matrix with Instructions.xlsx",
        matrix_content=_mixed_matrix_with_instructions_bytes(),
    )

    def fake_extract(path: str, filename: str):
        if filename.endswith(".xlsx"):
            return "--- Page 1 ---\nRAW FULL WORKBOOK CONTENT", 1
        return "--- Page 1 ---\nThe contractor shall deliver the base scope.", 1

    monkeypatch.setattr(
        lifecycle_env["intake"], "_extract_text_for_intake", fake_extract
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert "commercial off-the-shelf" in (document.extracted_text_md or "")
        assert "RAW FULL WORKBOOK CONTENT" not in (
            document.extracted_text_md or ""
        )

    lifecycle_env["service"].dismiss_cost_matrix(proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        ordinary_text = document.extracted_text_md or ""
    assert "Project Manager" in ordinary_text
    assert "Grand Total" in ordinary_text
    assert "hidden internal note" not in ordinary_text
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2


def test_possible_matrix_can_be_dismissed_without_submission_block(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Instructions.xlsx",
        matrix_content=_ambiguous_pricing_instructions_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        source_document_id = artifact.source_document_id
        prior_role = db.get(RfpPackageDocument, source_document_id).document_role
        db.get(Proposal, proposal_id).cots_orientation = True
        db.commit()

    with pytest.raises(
        lifecycle_env["service"].CostMatrixError,
        match="Confirm this workbook",
    ):
        lifecycle_env["service"].generate_cost_matrix(proposal_id, artifact_id)

    lifecycle_env["service"].dismiss_cost_matrix(proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, source_document_id)
        assert artifact.status == "dismissed"
        assert document.document_role == prior_role
        assert db.get(Proposal, proposal_id).cots_orientation is False
        assert "Offerors shall explain all pricing assumptions" in (
            document.extracted_text_md or ""
        )
        assert document.page_count == 1
        assert not lifecycle_env["service"].is_cost_matrix_document(
            db, source_document_id
        )
    with pytest.raises(
        lifecycle_env["service"].CostMatrixError,
        match="Confirm this workbook",
    ):
        lifecycle_env["service"].generate_cost_matrix(proposal_id, artifact_id)
    assert lifecycle_env["service"].cost_matrix_submission_check(proposal_id) == {
        "verified": True,
        "count": 0,
        "current": 0,
        "detail": "No cost matrix was supplied",
    }
    assert (
        lifecycle_env["service"].get_cost_matrix_requirements_context(
            proposal_id
        )
        == []
    )

    ordinary_extracts: list[str] = []

    def fake_extract(path: str, filename: str):
        ordinary_extracts.append(filename)
        return "--- Page 1 ---\nBuyer-authored source content", 1

    monkeypatch.setattr(
        lifecycle_env["intake"], "_extract_text_for_intake", fake_extract
    )
    assert lifecycle_env["intake"]._parse_documents(proposal_id) == 2
    assert ordinary_extracts == ["solicitation.pdf"]


@pytest.mark.parametrize("action", ["confirm_cost_matrix", "dismiss_cost_matrix"])
def test_archived_candidate_cannot_be_confirmed_or_dismissed(
    lifecycle_env,
    action: str,
) -> None:
    from app.core.enums import ProposalStatus
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Pricing Schedule Instructions.xlsx",
        matrix_content=_ambiguous_pricing_instructions_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        proposal.status = ProposalStatus.ARCHIVED
        db.commit()

    with pytest.raises(PermissionError, match="archived"):
        getattr(lifecycle_env["service"], action)(proposal_id, artifact_id)

    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert artifact.status == "needs_confirmation"
        assert document.document_role != "cost_matrix"


def test_all_financial_targets_skipped_is_not_ready(
    lifecycle_env,
    monkeypatch,
) -> None:
    proposal_id, artifact_id = _create_reviewed_it_matrix(
        lifecycle_env, monkeypatch
    )
    matrix = lifecycle_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    target_ids = [
        target["id"]
        for target in matrix["analysis"]["targets"]
        if target["category"] == "pricing"
    ]
    lifecycle_env["service"].save_cost_matrix_mapping(
        proposal_id,
        artifact_id,
        {
            target_id: {
                "mode": "skip",
                "reason": "Operator marked this line not applicable",
            }
            for target_id in target_ids
        },
    )
    matrix = lifecycle_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert matrix["readiness"]["ready"] is False
    assert any(
        "financial" in blocker.lower() and "skip" in blocker.lower()
        for blocker in matrix["readiness"]["blockers"]
    )
    with pytest.raises(lifecycle_env["service"].CostMatrixNotReadyError):
        lifecycle_env["service"].generate_cost_matrix(
            proposal_id, artifact_id
        )


def test_mapping_suggestions_never_cross_numeric_units() -> None:
    from app.services.cost_matrix import _mapping_suggestion

    money_target = {
        "label": "Project Manager",
        "header": "Total Cost",
        "kind": "money",
        "semantic": None,
    }
    hours_only = [
        {
            "key": "pricing.labor.7.hours",
            "label": "Labor: Project Manager — hours",
            "value": 500,
            "kind": "number",
        }
    ]
    assert _mapping_suggestion(money_target, hours_only) is None

    percentage_target = {
        "label": "Markup",
        "header": "Markup %",
        "kind": "percentage",
        "semantic": None,
    }
    basis_points_source = [
        {
            "key": "payment.scan.proposed_credit_card_markup_bps",
            "label": "Proposed credit card markup bps",
            "value": 24,
            "kind": "number",
        }
    ]
    assert _mapping_suggestion(percentage_target, basis_points_source) is None

    phase_target = {
        "label": "Build phase",
        "header": "Price",
        "kind": "money",
        "semantic": None,
    }
    exact_money_source = [
        {
            "key": "pricing.phase.0.phase_price_usd",
            "label": "Phase: Build phase — price",
            "value": 100_000,
            "kind": "money",
        }
    ]
    suggestion = _mapping_suggestion(phase_target, exact_money_source)
    assert suggestion is not None
    assert suggestion["source_key"] == "pricing.phase.0.phase_price_usd"


def test_percentage_values_require_explicit_decimal_semantics(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import PricingPackage

    proposal_id, artifact_id = _create_reviewed_it_matrix(
        lifecycle_env,
        monkeypatch,
        matrix_content=_percentage_matrix_bytes(),
        matrix_name="Buyer Percentage Cost Matrix.xlsx",
    )
    matrix = lifecycle_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    percentage_targets = [
        target
        for target in matrix["analysis"]["targets"]
        if target["category"] == "pricing" and target["kind"] == "percentage"
    ]
    assert len(percentage_targets) == 2

    with pytest.raises(
        lifecycle_env["service"].CostMatrixError,
        match="enter 24% or 0.24, not 24",
    ):
        lifecycle_env["service"].save_cost_matrix_mapping(
            proposal_id,
            artifact_id,
            {
                percentage_targets[0]["id"]: {
                    "mode": "manual",
                    "value": 24,
                }
            },
        )

    lifecycle_env["service"].save_cost_matrix_mapping(
        proposal_id,
        artifact_id,
        {
            percentage_targets[0]["id"]: {
                "mode": "manual",
                "value": "24%",
            },
            percentage_targets[1]["id"]: {
                "mode": "manual",
                "value": "0.10",
            },
        },
    )
    matrix = lifecycle_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert matrix["mapping"][percentage_targets[0]["id"]]["value"] == 0.24

    output_id = lifecycle_env["service"].generate_cost_matrix(
        proposal_id, artifact_id
    )
    generated, _filename = lifecycle_env["service"].get_cost_matrix_download(
        output_id
    )
    workbook = load_workbook(io.BytesIO(generated), data_only=False)
    try:
        assert (
            workbook[percentage_targets[0]["sheet"]][
                percentage_targets[0]["cell"]
            ].value
            == 0.24
        )
    finally:
        workbook.close()

    with lifecycle_env["db"].SessionLocal() as db:
        package = (
            db.query(PricingPackage)
            .filter_by(proposal_id=proposal_id, scenario="MEDIUM")
            .one()
        )
        package.indirect_costs_json = {
            "target_margin_pct": 24,
            "reviewed_margin_pct": 0.24,
        }
        db.commit()
    sources = {
        source["key"]: source
        for source in lifecycle_env["service"].get_cost_source_catalog(
            proposal_id
        )["sources"]
    }
    assert sources["pricing.indirect.target_margin_pct"]["kind"] == "number"
    assert sources["pricing.indirect.reviewed_margin_pct"]["kind"] == "percentage"


def test_it_source_catalog_uses_persisted_odc_and_phase_shapes(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import PricingPackage

    proposal_id, _artifact_id = _create_reviewed_it_matrix(
        lifecycle_env, monkeypatch
    )
    with lifecycle_env["db"].SessionLocal() as db:
        package = (
            db.query(PricingPackage)
            .filter_by(proposal_id=proposal_id, scenario="MEDIUM")
            .one()
        )
        package.odcs_json = [
            {
                "item": "Travel",
                "amount_usd": 100,
                "year_count": 3,
                "extended_amount_usd": 300,
            },
            {
                "item": "Equipment",
                "amount_usd": 50,
                "year_count": 1,
            },
        ]
        package.phase_breakdown_json = [
            {
                "name": "Build",
                "start_month": 2,
                "duration_months": 4,
                "phase_total_hours": 10,
                "phase_contingency_hours": 1,
                "phase_contingency_cost_usd": 25,
                "phase_price_usd": 200,
                "phase_subtotal_cost_usd": 150,
                "phase_loaded_cost_usd": 100,
                "phase_ga_usd": 25,
                "phase_profit_usd": 50,
            }
        ]
        db.commit()

    catalog = lifecycle_env["service"].get_cost_source_catalog(proposal_id)
    sources = {source["key"]: source for source in catalog["sources"]}
    assert sources["pricing.odcs_total"]["value"] == 350
    assert sources["pricing.odc.0.extended_amount_usd"]["value"] == 300
    assert sources["pricing.odc.0.amount_usd"]["value"] == 100
    assert sources["pricing.odc.0.year_count"]["value"] == 3
    assert sources[
        "pricing.phase.0.phase_contingency_cost_usd"
    ]["value"] == 25
    assert sources["pricing.phase.0.phase_total_hours"]["value"] == 10
    assert sources["pricing.phase.0.phase_contingency_hours"]["value"] == 1
    assert sources["pricing.phase.0.start_month"]["value"] == 2
    assert sources["pricing.phase.0.duration_months"]["value"] == 4


def test_generation_never_overwrites_colliding_output(
    lifecycle_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixOutput, Proposal

    proposal_id, artifact_id = _create_reviewed_it_matrix(
        lifecycle_env, monkeypatch
    )
    _map_two_financial_targets(lifecycle_env, proposal_id, artifact_id)
    with lifecycle_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        output_dir = (
            lifecycle_env["root"]
            / str(proposal.rfp_package_id)
            / "cost_matrix_outputs"
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    collision = (
        output_dir
        / f"Buyer_Cost_Matrix_matrix{artifact_id}_completed_v1.xlsx"
    )
    sentinel = b"existing file that does not belong to this generation"
    collision.write_bytes(sentinel)

    output_id: int | None = None
    try:
        output_id = lifecycle_env["service"].generate_cost_matrix(
            proposal_id, artifact_id
        )
    except lifecycle_env["service"].CostMatrixError:
        # Failing closed is acceptable; silently replacing the file is not.
        pass

    assert collision.read_bytes() == sentinel
    if output_id is not None:
        with lifecycle_env["db"].SessionLocal() as db:
            output = db.get(CostMatrixOutput, output_id)
            assert Path(output.output_storage_path) != collision
            assert Path(output.output_storage_path).is_file()
    else:
        with lifecycle_env["db"].SessionLocal() as db:
            assert db.query(CostMatrixOutput).count() == 0
        assert not list(output_dir.glob("*.generating"))


@pytest.mark.parametrize("damage", ["missing", "tampered"])
def test_submission_readiness_fails_closed_when_output_is_not_intact(
    lifecycle_env,
    monkeypatch,
    damage: str,
) -> None:
    from app.models import CostMatrixOutput

    proposal_id, artifact_id = _create_reviewed_it_matrix(
        lifecycle_env, monkeypatch
    )
    _map_two_financial_targets(lifecycle_env, proposal_id, artifact_id)
    output_id = lifecycle_env["service"].generate_cost_matrix(
        proposal_id, artifact_id
    )
    with lifecycle_env["db"].SessionLocal() as db:
        output = db.get(CostMatrixOutput, output_id)
        output_path = Path(output.output_storage_path)
    if damage == "missing":
        output_path.unlink()
    else:
        output_path.write_bytes(b"tampered workbook bytes")

    check = lifecycle_env["service"].cost_matrix_submission_check(proposal_id)
    assert check["verified"] is False
    assert check["current"] == 0
    snapshot = lifecycle_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert snapshot["latest_output"]["current"] is False
    assert any(
        "output" in blocker.lower()
        for blocker in snapshot["readiness"]["blockers"]
    )
    with pytest.raises(lifecycle_env["service"].CostMatrixError):
        lifecycle_env["service"].get_cost_matrix_download(output_id)


def test_deleting_source_document_cascades_artifact_and_outputs(
    lifecycle_env,
) -> None:
    from app.models import (
        CostMatrixArtifact,
        CostMatrixOutput,
        RfpPackageDocument,
    )

    _proposal_id, artifact_id = _create_proposal(
        lifecycle_env,
        matrix_name="Buyer Cost Matrix.xlsx",
        matrix_content=_strong_matrix_bytes(),
    )
    with lifecycle_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        source_document_id = artifact.source_document_id
        output = CostMatrixOutput(
            artifact_id=artifact_id,
            version=1,
            pricing_scenario="MEDIUM",
            pricing_basis_sha256="a" * 64,
            generation_provenance_json={},
            mapping_snapshot_json={},
            output_filename="completed.xlsx",
            output_storage_path=str(lifecycle_env["root"] / "completed.xlsx"),
            output_sha256="b" * 64,
            validation_json={"valid": True},
            generated_at=datetime.now(UTC),
        )
        db.add(output)
        db.commit()
        output_id = output.id

    with lifecycle_env["db"].SessionLocal() as db:
        document = db.get(RfpPackageDocument, source_document_id)
        db.delete(document)
        db.commit()

    with lifecycle_env["db"].SessionLocal() as db:
        assert db.get(CostMatrixArtifact, artifact_id) is None
        assert db.get(CostMatrixOutput, output_id) is None
