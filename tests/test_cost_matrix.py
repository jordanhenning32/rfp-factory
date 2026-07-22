from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _matrix_bytes(*, alternate: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fee Schedule" if alternate else "Buyer Pricing "
    ws["B2"] = "Vendor Name:"
    ws.merge_cells("C2:E2")
    ws["C2"] = "Vendor Name"
    ws["B3"] = "Date:"
    ws.merge_cells("C3:E3")
    ws["C3"].number_format = "m/d/yyyy"
    if alternate:
        ws["B7"] = "Service"
        ws["C7"] = "Unit Fee"
        ws["B8"] = "Card transaction"
        ws["C8"] = 0
        ws["C8"].number_format = '$0.0000'
        ws["B9"] = "ACH transaction"
        ws["C9"] = 0
        ws["C9"].number_format = '$0.0000'
    else:
        ws["B8"] = "Work Item"
        ws["C8"] = "Total Cost"
        ws["B9"] = "Workstream Alpha"
        ws["C9"] = 0
        ws["C9"].number_format = '$#,##0.00'
        ws["B10"] = "Workstream Beta"
        ws["C10"] = 0
        ws["C10"].number_format = '$#,##0.00'
        ws["B11"] = "Total"
        ws["C11"] = "=SUM(C9:C10)"
        ws["C11"].number_format = '$#,##0.00'
        ws["B20"] = "Hidden internal cost"
        ws["C20"] = 0
        ws["C20"].number_format = '$#,##0.00'
        ws.row_dimensions[20].hidden = True
        ws.column_dimensions["F"].hidden = True
    hidden = wb.create_sheet("Buyer source data")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Preserve me"
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _non_matrix_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append(["Requirement ID", "Description", "Mandatory"])
    ws.append(["R-1", "Provide a project plan", "Yes"])
    ws.append(["R-2", "Describe security controls", "Yes"])
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


@pytest.fixture()
def cost_matrix_env(inmemory_db, monkeypatch, tmp_path):
    import app.db.session as db_session
    import app.jobs.intake as intake
    import app.services.cost_matrix as cost_matrix
    import app.services.proposals as proposals

    package_root = tmp_path / "rfp_packages"
    package_root.mkdir()
    monkeypatch.setattr(cost_matrix, "RFP_PACKAGES_DIR", package_root)
    monkeypatch.setattr(proposals, "RFP_PACKAGES_DIR", package_root)
    monkeypatch.setattr(cost_matrix, "session_scope", db_session.session_scope)
    monkeypatch.setattr(intake, "session_scope", db_session.session_scope)
    return {
        "db": db_session,
        "root": package_root,
        "service": cost_matrix,
        "proposals": proposals,
        "intake": intake,
    }


def test_inspector_is_template_specific_and_conservative() -> None:
    from app.services.cost_matrix import inspect_cost_matrix

    line_matrix = inspect_cost_matrix("Attachment F.xlsx", _matrix_bytes())
    assert line_matrix["classification"]["is_cost_matrix"] is True
    assert [
        (target["label"], target["cell"], target["category"])
        for target in line_matrix["targets"]
    ] == [
        ("Vendor Name:", "C2", "metadata"),
        ("Date:", "C3", "metadata"),
        ("Workstream Alpha", "C9", "pricing"),
        ("Workstream Beta", "C10", "pricing"),
    ]
    assert line_matrix["reconciliations"][0]["member_target_ids"] == [
        "target-3", "target-4",
    ]

    fee_matrix = inspect_cost_matrix("Agency Fee Schedule.xlsx", _matrix_bytes(alternate=True))
    assert fee_matrix["classification"]["is_cost_matrix"] is True
    assert {target["label"] for target in fee_matrix["targets"] if target["category"] == "pricing"} == {
        "Card transaction", "ACH transaction",
    }
    assert fee_matrix["reconciliations"] == []

    requirements = inspect_cost_matrix("requirements.xlsx", _non_matrix_bytes())
    assert requirements["classification"]["is_cost_matrix"] is False
    assert requirements["targets"] == []


def test_ooxml_generation_changes_only_approved_parts_and_preserves_structure() -> None:
    from app.models import CostMatrixArtifact
    from app.services.cost_matrix import (
        _patch_workbook,
        _target_map,
        _validate_generated_workbook,
        inspect_cost_matrix,
    )

    source = _matrix_bytes()
    analysis = inspect_cost_matrix("Attachment F.xlsx", source)
    artifact = CostMatrixArtifact(
        id=1,
        proposal_id=1,
        source_document_id=1,
        status="mapping_required",
        template_sha256=analysis["template_sha256"],
        analysis_version=analysis["analysis_version"],
        analysis_json=analysis,
        mapping_json={},
    )
    values = {
        "target-1": "Acme Federal Solutions LLC",
        "target-2": "2026-07-21",
        "target-3": 400.25,
        "target-4": 599.75,
    }
    output, changed = _patch_workbook(
        source,
        targets=_target_map(artifact),
        resolved=values,
    )
    validation = _validate_generated_workbook(
        source,
        output,
        artifact=artifact,
        resolved=values,
        changed_parts=changed,
    )
    assert validation["valid"] is True
    assert set(changed) <= {"xl/workbook.xml", "xl/worksheets/sheet1.xml"}

    with zipfile.ZipFile(io.BytesIO(source)) as original_zip, zipfile.ZipFile(io.BytesIO(output)) as output_zip:
        original_parts = set(original_zip.namelist())
        assert set(output_zip.namelist()) == original_parts
        for part in original_parts - set(changed):
            assert output_zip.read(part) == original_zip.read(part), part

    result = load_workbook(io.BytesIO(output), data_only=False)
    try:
        assert result.sheetnames == ["Buyer Pricing ", "Buyer source data"]
        assert result["Buyer source data"].sheet_state == "hidden"
        assert result["Buyer Pricing "]["C11"].value == "=SUM(C9:C10)"
        assert {str(item) for item in result["Buyer Pricing "].merged_cells.ranges} == {
            "C2:E2", "C3:E3",
        }
    finally:
        result.close()


def _create_proposal_with_matrix(env):
    from app.models import CostMatrixArtifact, Proposal, RfpPackageDocument

    matrix = _matrix_bytes()
    with env["db"].SessionLocal() as db:
        proposal = env["proposals"].create_proposal_with_files(
            db,
            title="Template-specific pricing",
            files=[
                env["proposals"].UploadedFile("solicitation.pdf", b"%PDF fake"),
                env["proposals"].UploadedFile("Attachment F.xlsx", matrix),
            ],
        )
        db.commit()
        proposal_id = proposal.id
    with env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        artifact = db.query(CostMatrixArtifact).filter_by(proposal_id=proposal_id).one()
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        return proposal_id, artifact.id, Path(document.storage_path), matrix


def test_original_package_detects_immediately_preserves_source_and_skips_intake(
    cost_matrix_env,
    monkeypatch,
) -> None:
    from app.models import CostMatrixArtifact, RfpPackageDocument

    proposal_id, artifact_id, source_path, original = _create_proposal_with_matrix(cost_matrix_env)
    assert source_path.read_bytes() == original
    with cost_matrix_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        assert document.document_role == "cost_matrix"
        assert str(document.document_type) in {"form_template", "RfpDocumentType.FORM_TEMPLATE"}
        assert document.extracted_text_md is None

    calls: list[str] = []

    def fake_extract(path: str, filename: str):
        calls.append(filename)
        return "--- Page 1 ---\nA real solicitation requirement", 1

    monkeypatch.setattr(cost_matrix_env["intake"], "_extract_text_for_intake", fake_extract)
    assert cost_matrix_env["intake"]._parse_documents(proposal_id) == 1
    assert calls == ["solicitation.pdf"]
    with cost_matrix_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, artifact_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        review = document.structure_json["requirements_review"]
        assert review["status"] == "not_applicable"
        assert review["requires_manual_review"] is False
        assert "no separate written requirements" in review["reason"]


def test_late_attachment_is_collision_safe_duplicate_safe_and_atomic(
    cost_matrix_env,
    monkeypatch,
) -> None:
    from app.core.enums import ProposalStatus
    from app.models import CostMatrixArtifact, Proposal, RfpPackage, RfpPackageDocument

    with cost_matrix_env["db"].SessionLocal() as db:
        package = RfpPackage(
            uploaded_at=__import__("datetime").datetime.now(__import__("datetime").timezone.utc),
            storage_dir="",
        )
        db.add(package)
        db.flush()
        package.storage_dir = str(cost_matrix_env["root"] / str(package.id))
        proposal = Proposal(
            rfp_package_id=package.id,
            title="Late matrix",
            status=ProposalStatus.DRAFTING,
        )
        db.add(proposal)
        db.commit()
        proposal_id = proposal.id

    content = _matrix_bytes()
    review_calls: list[tuple[int, int]] = []
    monkeypatch.setattr(
        cost_matrix_env["service"],
        "_review_late_attached_cost_matrix",
        lambda pid, did: review_calls.append((pid, did)),
    )
    first_id = cost_matrix_env["service"].attach_cost_matrix(
        proposal_id,
        filename="pricing.xlsx",
        content=content,
    )
    assert first_id
    assert review_calls == []
    with cost_matrix_env["db"].SessionLocal() as db:
        artifact = db.get(CostMatrixArtifact, first_id)
        document = db.get(RfpPackageDocument, artifact.source_document_id)
        review = document.structure_json["requirements_review"]
        assert review["status"] == "not_applicable"
        assert review["requires_manual_review"] is False
        assert "no separate written requirements" in review["reason"]
    with pytest.raises(cost_matrix_env["service"].CostMatrixError, match="already attached"):
        cost_matrix_env["service"].attach_cost_matrix(
            proposal_id,
            filename="pricing-again.xlsx",
            content=content,
        )

    before = sorted(path.name for path in (cost_matrix_env["root"] / "1").glob("*.xlsx"))
    original_register = cost_matrix_env["service"]._register_artifact

    def fail_registration(*args, **kwargs):
        raise RuntimeError("database write failed")

    monkeypatch.setattr(cost_matrix_env["service"], "_register_artifact", fail_registration)
    with pytest.raises(RuntimeError, match="database write failed"):
        cost_matrix_env["service"].attach_cost_matrix(
            proposal_id,
            filename="different.xlsx",
            content=_matrix_bytes(alternate=True),
        )
    after = sorted(path.name for path in (cost_matrix_env["root"] / "1").glob("*.xlsx"))
    assert after == before
    monkeypatch.setattr(cost_matrix_env["service"], "_register_artifact", original_register)

    with cost_matrix_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        proposal.status = ProposalStatus.ARCHIVED
        db.commit()
    with pytest.raises(PermissionError, match="archived"):
        cost_matrix_env["service"].attach_cost_matrix(
            proposal_id,
            filename="archived.xlsx",
            content=_matrix_bytes(alternate=True),
        )


def test_mapping_generation_reconciliation_freshness_and_archive_download(
    cost_matrix_env,
    monkeypatch,
) -> None:
    from app.core.enums import ProposalStatus
    from app.models import PricingPackage, Proposal

    proposal_id, artifact_id, source_path, original = _create_proposal_with_matrix(cost_matrix_env)
    monkeypatch.setattr(
        cost_matrix_env["service"],
        "get_cost_review_freshness",
        lambda db, proposal_id, scenario: {
            "verified": True,
            "legacy": False,
            "review_count": 1,
            "detail": "Current reviewed cost build",
        },
    )
    with cost_matrix_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        proposal.proposed_scenario = "MEDIUM"
        proposal.service_line = "it_services"
        db.add(PricingPackage(
            proposal_id=proposal_id,
            scenario="MEDIUM",
            loaded_labor_cost=700,
            odcs_json=[],
            subcontractor_costs=0,
            indirect_costs_json={},
            total_proposed_price=1000,
            pnl_projection_json={},
            phase_breakdown_json=[],
        ))
        db.commit()

    snapshots = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)
    targets = snapshots[0]["analysis"]["targets"]
    ids = {target["label"]: target["id"] for target in targets}
    cost_matrix_env["service"].save_cost_matrix_mapping(
        proposal_id,
        artifact_id,
        {
            ids["Vendor Name:"]: {"mode": "source", "source_key": "company.legal_name"},
            ids["Date:"]: {"mode": "skip", "reason": "Buyer date meaning not yet specified"},
            ids["Workstream Alpha"]: {"mode": "manual", "value": "400.00"},
            ids["Workstream Beta"]: {"mode": "manual", "value": "600.00"},
        },
    )
    ready = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert ready["readiness"]["ready"] is True
    assert cost_matrix_env["service"].cost_matrix_submission_check(proposal_id)["verified"] is False

    output_id = cost_matrix_env["service"].generate_cost_matrix(proposal_id, artifact_id)
    assert source_path.read_bytes() == original
    data, filename = cost_matrix_env["service"].get_cost_matrix_download(output_id)
    assert filename.endswith("_completed_v1.xlsx")
    wb = load_workbook(io.BytesIO(data), data_only=False)
    try:
        ws = wb["Buyer Pricing "]
        assert ws["C2"].value == "Acme Federal Solutions LLC"
        assert ws["C9"].value == 400
        assert ws["C10"].value == 600
        assert ws["C11"].value == "=SUM(C9:C10)"
        assert wb["Buyer source data"].sheet_state == "hidden"
    finally:
        wb.close()

    current = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert current["latest_output"]["current"] is True
    assert cost_matrix_env["service"].cost_matrix_submission_check(proposal_id)["verified"] is True
    with cost_matrix_env["db"].SessionLocal() as db:
        package = db.query(PricingPackage).filter_by(
            proposal_id=proposal_id,
            scenario="MEDIUM",
        ).one()
        package.total_proposed_price = 1100
        db.commit()
    stale = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert stale["status"] == "stale"
    assert any("reconcile" in blocker for blocker in stale["readiness"]["blockers"])
    assert cost_matrix_env["service"].cost_matrix_submission_check(proposal_id)["verified"] is False

    with cost_matrix_env["db"].SessionLocal() as db:
        proposal = db.get(Proposal, proposal_id)
        proposal.status = ProposalStatus.ARCHIVED
        db.commit()
    with pytest.raises(PermissionError, match="archived"):
        cost_matrix_env["service"].generate_cost_matrix(proposal_id, artifact_id)
    archived_data, _ = cost_matrix_env["service"].get_cost_matrix_download(output_id)
    assert archived_data == data


def test_payment_fee_schedule_uses_its_own_numeric_sources_without_total_assumption(
    cost_matrix_env,
    monkeypatch,
) -> None:
    content = _matrix_bytes(alternate=True)
    with cost_matrix_env["db"].SessionLocal() as db:
        proposal = cost_matrix_env["proposals"].create_proposal_with_files(
            db,
            title="Payment fee schedule",
            files=[
                cost_matrix_env["proposals"].UploadedFile("rfp.pdf", b"%PDF fake"),
                cost_matrix_env["proposals"].UploadedFile("Agency Fee Schedule.xlsx", content),
            ],
        )
        proposal.service_line = "payment_systems"
        proposal.payment_market_scan_json = json.dumps({
            "pricing_structure": {
                "proposed_per_txn_fee_usd": 0.08,
                "proposed_ach_fee_usd": 0.25,
                "proposed_monthly_fee_usd": 90,
            },
            "profit_math": {"annual_net_profit_midpoint_usd": 25000},
        })
        proposal.payment_cost_review_findings_json = json.dumps({"findings": []})
        db.commit()
        proposal_id = proposal.id

    monkeypatch.setattr(
        cost_matrix_env["service"], "payment_market_scan_is_current", lambda scan: True
    )
    monkeypatch.setattr(
        cost_matrix_env["service"],
        "payment_cost_review_is_current",
        lambda proposal_id, review, db=None: True,
    )
    matrix = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert matrix["analysis"]["reconciliations"] == []
    source_keys = {source["key"] for source in matrix["sources"]}
    per_txn_key = "payment.scan.pricing_structure.proposed_per_txn_fee_usd"
    ach_key = "payment.scan.pricing_structure.proposed_ach_fee_usd"
    assert {per_txn_key, ach_key} <= source_keys
    ids = {
        target["label"]: target["id"]
        for target in matrix["analysis"]["targets"]
    }
    cost_matrix_env["service"].save_cost_matrix_mapping(
        proposal_id,
        matrix["id"],
        {
            ids["Vendor Name:"]: {"mode": "source", "source_key": "company.legal_name"},
            ids["Date:"]: {"mode": "skip", "reason": "Date semantics unresolved"},
            ids["Card transaction"]: {"mode": "source", "source_key": per_txn_key},
            ids["ACH transaction"]: {"mode": "source", "source_key": ach_key},
        },
    )
    ready = cost_matrix_env["service"].get_cost_matrix_snapshots(proposal_id)[0]
    assert ready["readiness"]["ready"] is True
    output_id = cost_matrix_env["service"].generate_cost_matrix(
        proposal_id, matrix["id"]
    )
    data, _ = cost_matrix_env["service"].get_cost_matrix_download(output_id)
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    try:
        assert workbook["Fee Schedule"]["C8"].value == 0.08
        assert workbook["Fee Schedule"]["C9"].value == 0.25
    finally:
        workbook.close()


def test_operator_target_rejects_formula_and_hidden_source_cells(
    cost_matrix_env,
) -> None:
    proposal_id, artifact_id, _source_path, _original = _create_proposal_with_matrix(
        cost_matrix_env
    )
    with pytest.raises(cost_matrix_env["service"].CostMatrixError, match="Formula cells"):
        cost_matrix_env["service"].add_cost_matrix_target(
            proposal_id,
            artifact_id,
            sheet="Buyer Pricing ",
            cell_coordinate="C11",
            label="Do not overwrite total",
        )
    with pytest.raises(cost_matrix_env["service"].CostMatrixError, match="Hidden worksheets"):
        cost_matrix_env["service"].add_cost_matrix_target(
            proposal_id,
            artifact_id,
            sheet="Buyer source data",
            cell_coordinate="A1",
            label="Private source",
        )
    with pytest.raises(cost_matrix_env["service"].CostMatrixError, match="hidden rows"):
        cost_matrix_env["service"].add_cost_matrix_target(
            proposal_id,
            artifact_id,
            sheet="Buyer Pricing ",
            cell_coordinate="C20",
            label="Private row",
        )
    with pytest.raises(cost_matrix_env["service"].CostMatrixError, match="hidden rows or columns"):
        cost_matrix_env["service"].add_cost_matrix_target(
            proposal_id,
            artifact_id,
            sheet="Buyer Pricing ",
            cell_coordinate="F9",
            label="Private column",
        )
