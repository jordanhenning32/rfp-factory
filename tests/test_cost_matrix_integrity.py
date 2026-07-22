"""Adversarial workbook-integrity contracts for buyer cost matrices.

These tests deliberately avoid assuming that a buyer workbook was produced by
Excel or that it follows the example template's row/column layout.
"""
from __future__ import annotations

import io
import re
import warnings
import zipfile
from datetime import date
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection

from app.models import CostMatrixArtifact
from app.services.cost_matrix import (
    RECONCILIATION_MAPPING_KEY,
    CostMatrixError,
    _patch_workbook,
    _reconciliation_blockers,
    _resolve_mapping_values,
    _target_map,
    _validate_generated_workbook,
    extract_cost_matrix_instruction_text,
    inspect_cost_matrix,
    try_inspect_cost_matrix,
)

_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _save(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _replace_zip_part(data: bytes, part: str, transform) -> bytes:
    with zipfile.ZipFile(io.BytesIO(data), "r") as source:
        infos = source.infolist()
        parts = {info.filename: source.read(info) for info in infos}
    parts[part] = transform(parts[part])
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as target:
        for info in infos:
            target.writestr(info, parts[info.filename])
    return output.getvalue()


def _artifact(analysis: dict) -> CostMatrixArtifact:
    return CostMatrixArtifact(
        id=1,
        proposal_id=1,
        source_document_id=1,
        status="mapping_required",
        template_sha256=analysis["template_sha256"],
        analysis_version=analysis["analysis_version"],
        analysis_json=analysis,
        mapping_json={},
    )


def _simple_pricing_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet["A1"] = "Item"
    sheet["B1"] = "Cost"
    sheet["A2"] = "Implementation"
    sheet["B2"] = 0
    sheet["B2"].number_format = "$#,##0.00"
    return _save(workbook)


def _prefix_main_namespace(xml: bytes, *, root_tag: str) -> bytes:
    """Serialize the same valid spreadsheet vocabulary with an ``x:`` prefix."""
    text = xml.decode("utf-8")
    text = text.replace(
        f'<{root_tag} xmlns="{_SHEET_NS}">',
        f'<x:{root_tag} xmlns:x="{_SHEET_NS}">',
        1,
    )
    text = re.sub(
        r"<(/?)(?!x:)([A-Za-z][A-Za-z0-9]*)",
        r"<\1x:\2",
        text,
    )
    return text.encode("utf-8")


def test_prefixed_ooxml_worksheet_can_be_completed() -> None:
    source = _replace_zip_part(
        _simple_pricing_workbook(),
        "xl/worksheets/sheet1.xml",
        lambda xml: _prefix_main_namespace(xml, root_tag="worksheet"),
    )
    source = _replace_zip_part(
        source,
        "xl/workbook.xml",
        lambda xml: _prefix_main_namespace(xml, root_tag="workbook"),
    )

    # This is a valid alternate namespace serialization, not a corrupt file.
    workbook = load_workbook(io.BytesIO(source), data_only=False)
    try:
        assert workbook["Pricing"]["B2"].value == 0
    finally:
        workbook.close()

    analysis = inspect_cost_matrix("Buyer Price Schedule.xlsx", source)
    artifact = _artifact(analysis)
    price_target = next(
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    )
    output, changed = _patch_workbook(
        source,
        targets=_target_map(artifact),
        resolved={price_target["id"]: 123.45},
    )
    validation = _validate_generated_workbook(
        source,
        output,
        artifact=artifact,
        resolved={price_target["id"]: 123.45},
        changed_parts=changed,
    )
    assert validation["valid"] is True


def test_non_self_closing_calculation_properties_are_updated_in_place() -> None:
    source = _replace_zip_part(
        _simple_pricing_workbook(),
        "xl/workbook.xml",
        lambda xml: re.sub(
            rb"<calcPr\b([^>]*)/>",
            rb"<calcPr\1></calcPr>",
            xml,
            count=1,
        ),
    )
    analysis = inspect_cost_matrix("Buyer Price Schedule.xlsx", source)
    artifact = _artifact(analysis)
    price_target = next(
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    )
    output, _changed = _patch_workbook(
        source,
        targets=_target_map(artifact),
        resolved={price_target["id"]: 123.45},
    )

    with zipfile.ZipFile(io.BytesIO(output), "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc_nodes = workbook_root.findall(f"{{{_SHEET_NS}}}calcPr")
    assert len(calc_nodes) == 1
    assert calc_nodes[0].attrib["calcMode"] == "auto"
    assert calc_nodes[0].attrib["fullCalcOnLoad"] == "1"
    assert calc_nodes[0].attrib["forceFullCalc"] == "1"


def test_general_format_date_is_written_as_human_readable_text() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Offer"
    sheet["A1"] = "Date:"
    sheet["B1"] = None
    sheet["B1"].number_format = "General"
    sheet["A3"] = "Item"
    sheet["B3"] = "Total Cost"
    sheet["A4"] = "Service"
    sheet["B4"] = 0
    sheet["B4"].number_format = "$#,##0.00"
    source = _save(workbook)

    analysis = inspect_cost_matrix("Buyer Cost Form.xlsx", source)
    artifact = _artifact(analysis)
    date_target = next(
        target for target in analysis["targets"]
        if target.get("semantic") == "matrix_date"
    )
    assert date_target["number_format"] == "General"

    output, changed = _patch_workbook(
        source,
        targets=_target_map(artifact),
        resolved={date_target["id"]: "2026-07-21"},
    )
    validation = _validate_generated_workbook(
        source,
        output,
        artifact=artifact,
        resolved={date_target["id"]: "2026-07-21"},
        changed_parts=changed,
    )
    assert validation["valid"] is True

    result = load_workbook(io.BytesIO(output), data_only=False)
    try:
        assert result["Offer"]["B1"].number_format == "General"
        assert result["Offer"]["B1"].value == "2026-07-21"
    finally:
        result.close()


def test_protected_sheet_only_auto_proposes_unlocked_inputs() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Protected Pricing"
    sheet.append(["Item", "Cost"])
    sheet.append(["Government estimate", 500_000])
    sheet.append(["Offeror input", 0])
    for coordinate in ("B2", "B3"):
        sheet[coordinate].number_format = "$#,##0.00"
    sheet["B3"].protection = Protection(locked=False)
    sheet.protection.sheet = True

    analysis = inspect_cost_matrix("Protected Cost Schedule.xlsx", _save(workbook))
    pricing = [
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    ]
    assert [(target["cell"], target["label"]) for target in pricing] == [
        ("B3", "Offeror input"),
    ]


def test_nonzero_buyer_constants_are_not_assumed_to_be_writable_inputs() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet.append(["Item", "Cost"])
    sheet.append(["Government estimate", 500_000])
    sheet.append(["Not-to-exceed ceiling", 750_000])
    sheet.append(["Offeror proposed price", None])
    for coordinate in ("B2", "B3", "B4"):
        sheet[coordinate].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer Cost Worksheet.xlsx", _save(workbook))
    pricing = [
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    ]
    assert [(target["cell"], target["label"]) for target in pricing] == [
        ("B4", "Offeror proposed price"),
    ]


def test_company_metadata_detection_requires_an_explicit_name_label() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet["A1"] = "Company Address"
    sheet["B1"] = "123 Buyer Street"
    sheet["A2"] = "Vendor UEI"
    sheet["B2"] = "ABC123"
    sheet["A3"] = "Offeror Name:"
    sheet["B3"] = None
    sheet["A5"] = "Item"
    sheet["B5"] = "Total Cost"
    sheet["A6"] = "Service"
    sheet["B6"] = 0
    sheet["B6"].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer Pricing Matrix.xlsx", _save(workbook))
    metadata = [
        target for target in analysis["targets"]
        if target["category"] == "metadata"
    ]
    assert [(target["cell"], target["label"]) for target in metadata] == [
        ("B3", "Offeror Name:"),
    ]


def test_clin_grid_prefers_descriptions_over_unit_codes_for_price_labels() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CLIN Pricing"
    sheet.append([
        "CLIN",
        "Description",
        "Qty",
        "Unit",
        "Unit Price",
        "Extended Price",
    ])
    sheet.append([
        "0001",
        "Program management",
        12,
        "MO",
        None,
        "=C2*E2",
    ])
    sheet.append([
        "0002",
        "Cybersecurity assessment",
        1,
        "LOT",
        None,
        "=C3*E3",
    ])
    for coordinate in ("E2", "E3", "F2", "F3"):
        sheet[coordinate].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer CLIN Pricing Schedule.xlsx", _save(workbook))
    pricing = [
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    ]
    assert [(target["cell"], target["label"]) for target in pricing] == [
        ("E2", "Program management"),
        ("E3", "Cybersecurity assessment"),
    ]


def test_bid_schedule_title_surfaces_general_format_zero_inputs_for_confirmation() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attachment F"
    sheet["A1"] = "Bid Schedule"
    sheet["A3"] = "CLIN 0001"
    sheet["B3"] = 0
    sheet["A4"] = "CLIN 0002"
    sheet["B4"] = 0
    source = _save(workbook)

    analysis = try_inspect_cost_matrix("Attachment F.xlsx", source)
    assert analysis is not None
    assert analysis["classification"]["decision"] in {
        "confirmed",
        "needs_confirmation",
    }
    assert [
        (target["cell"], target["label"])
        for target in analysis["targets"]
        if target["category"] == "pricing"
    ] == [
        ("B3", "CLIN 0001"),
        ("B4", "CLIN 0002"),
    ]


def test_direct_blank_cell_insertion_keeps_rows_sorted_and_expands_dimension() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet["A1"] = "Price schedule"
    sheet["A5"] = "Footer"
    source = _save(workbook)
    target = {
        "id": "manual-1",
        "sheet": "Pricing",
        "cell": "B3",
        "kind": "money",
    }

    output, _changed = _patch_workbook(
        source,
        targets={"manual-1": target},
        resolved={"manual-1": 123.45},
    )
    with zipfile.ZipFile(io.BytesIO(output), "r") as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    rows = [int(value) for value in re.findall(r'<row\b[^>]*\br="(\d+)"', worksheet)]
    assert rows == [1, 3, 5]
    dimension = re.search(r'<dimension\b[^>]*\bref="([^"]+)"', worksheet)
    assert dimension is not None
    assert dimension.group(1) == "A1:B5"

    result = load_workbook(io.BytesIO(output), data_only=False)
    try:
        assert result["Pricing"]["B3"].value == 123.45
    finally:
        result.close()


def test_direct_target_can_be_inserted_into_self_closing_sheet_data() -> None:
    workbook = Workbook()
    source = _save(workbook)
    source = _replace_zip_part(
        source,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(b"<sheetData></sheetData>", b"<sheetData/>", 1),
    )
    target = {
        "id": "manual-1",
        "sheet": "Sheet",
        "cell": "A1",
        "kind": "text",
    }

    output, _changed = _patch_workbook(
        source,
        targets={"manual-1": target},
        resolved={"manual-1": "Reviewed value"},
    )
    result = load_workbook(io.BytesIO(output), data_only=False)
    try:
        assert result["Sheet"]["A1"].value == "Reviewed value"
    finally:
        result.close()


def test_multiple_sheet_totals_are_not_each_forced_to_proposal_total() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title in ("Base Year", "Option Year 1"):
        sheet = workbook.create_sheet(title)
        sheet.append(["Item", "Cost"])
        sheet.append(["Service A", 0])
        sheet.append(["Service B", 0])
        sheet.append(["Total", "=SUM(B2:B3)"])
        for coordinate in ("B2", "B3", "B4"):
            sheet[coordinate].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer Pricing Schedule.xlsx", _save(workbook))
    assert len(analysis["formulas"]) == 2
    assert analysis["reconciliations"] == []


@pytest.mark.parametrize(
    "label",
    ["Total Price", "Grand Total Cost", "Total Evaluated Bid", "Overall Total"],
)
def test_common_overall_total_labels_are_reconciled(label: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bid Pricing"
    sheet.append(["Item", "Cost"])
    sheet.append(["Service A", 0])
    sheet.append(["Service B", 0])
    sheet.append([label, "=SUM(B2:B3)"])
    for coordinate in ("B2", "B3", "B4"):
        sheet[coordinate].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer Pricing Schedule.xlsx", _save(workbook))
    assert len(analysis["reconciliations"]) == 1
    assert analysis["reconciliations"][0]["formula_cell"] == "B4"
    assert analysis["reconciliation_review"]["review_required"] is False


def test_non_money_total_is_never_reconciled_to_proposal_price() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cost Matrix"
    sheet.append(["Labor Category", "Hours"])
    sheet.append(["Project Manager", 0])
    sheet.append(["Engineer", 0])
    sheet.append(["Total", "=SUM(B2:B3)"])

    analysis = inspect_cost_matrix("Buyer Cost Matrix.xlsx", _save(workbook))
    assert [target["kind"] for target in analysis["targets"]] == ["number", "number"]
    assert analysis["reconciliations"] == []
    assert analysis["reconciliation_review"]["review_required"] is False


def test_operator_aggregate_reconciliation_sums_only_money_targets() -> None:
    artifact = CostMatrixArtifact(
        status="mapping_required",
        template_sha256="0" * 64,
        analysis_version="test",
        analysis_json={
            "targets": [
                {
                    "id": "money-1",
                    "sheet": "Pricing",
                    "cell": "B2",
                    "category": "pricing",
                    "kind": "money",
                },
                {
                    "id": "hours-1",
                    "sheet": "Pricing",
                    "cell": "C2",
                    "category": "pricing",
                    "kind": "number",
                },
            ],
            "reconciliation_review": {"review_required": True},
        },
        mapping_json={
            RECONCILIATION_MAPPING_KEY: {
                "mode": "aggregate_to_proposal_total",
            },
        },
    )
    sources = [{
        "key": "pricing.total_proposed_price",
        "value": 100,
    }]
    assert _reconciliation_blockers(
        artifact,
        sources=sources,
        resolved={"money-1": 100, "hours-1": 40},
    ) == []
    blockers = _reconciliation_blockers(
        artifact,
        sources=sources,
        resolved={"money-1": 90, "hours-1": 40},
    )
    assert len(blockers) == 1
    assert "$90.00" in blockers[0]
    assert "$130.00" not in blockers[0]


@pytest.mark.parametrize("review_mode", [False, True])
def test_skipped_blank_financial_cell_contributes_zero_to_reconciliation(
    review_mode: bool,
) -> None:
    targets = [
        {
            "id": "money-1",
            "sheet": "Pricing",
            "cell": "B2",
            "category": "pricing",
            "kind": "money",
            "existing_value": None,
        },
        {
            "id": "money-2",
            "sheet": "Pricing",
            "cell": "B3",
            "category": "pricing",
            "kind": "money",
            "existing_value": None,
        },
    ]
    analysis = {"targets": targets}
    mapping = {
        "money-1": {"mode": "manual", "value": 900},
        "money-2": {"mode": "skip", "reason": "Buyer left blank"},
    }
    if review_mode:
        analysis["reconciliation_review"] = {"review_required": True}
        mapping[RECONCILIATION_MAPPING_KEY] = {
            "mode": "aggregate_to_proposal_total",
        }
    else:
        analysis["reconciliation_review"] = {"review_required": False}
        analysis["reconciliations"] = [{
            "type": "sum_to_proposed_price",
            "sheet": "Pricing",
            "formula_cell": "B4",
            "member_target_ids": ["money-1", "money-2"],
            "tolerance_usd": 0.01,
        }]
    artifact = CostMatrixArtifact(
        status="mapping_required",
        template_sha256="0" * 64,
        analysis_version="test",
        analysis_json=analysis,
        mapping_json=mapping,
    )
    blockers = _reconciliation_blockers(
        artifact,
        sources=[{"key": "pricing.total_proposed_price", "value": 1000}],
        resolved={"money-1": 900},
    )
    assert len(blockers) == 1
    assert "$900.00" in blockers[0]
    assert "$1,000.00" in blockers[0]


def test_metadata_mapping_cannot_bypass_all_financial_inputs_skipped() -> None:
    artifact = CostMatrixArtifact(
        status="mapping_required",
        template_sha256="0" * 64,
        analysis_version="test",
        analysis_json={
            "targets": [
                {
                    "id": "money-1",
                    "sheet": "Pricing",
                    "cell": "B2",
                    "category": "pricing",
                    "kind": "money",
                    "label": "Base price",
                },
                {
                    "id": "money-2",
                    "sheet": "Pricing",
                    "cell": "B3",
                    "category": "pricing",
                    "kind": "money",
                    "label": "Option price",
                },
                {
                    "id": "text-1",
                    "sheet": "Pricing",
                    "cell": "B4",
                    "category": "metadata",
                    "kind": "text",
                    "label": "Company",
                },
            ],
        },
        mapping_json={
            "money-1": {"mode": "skip", "reason": "test"},
            "money-2": {"mode": "skip", "reason": "test"},
            "text-1": {"mode": "manual", "value": "Example LLC"},
        },
    )
    resolved, blockers = _resolve_mapping_values(artifact, [])
    assert resolved == {"text-1": "Example LLC"}
    assert any("All financial inputs are skipped" in blocker for blocker in blockers)


def test_saved_source_mapping_revalidates_current_unit_and_value() -> None:
    artifact = CostMatrixArtifact(
        status="mapping_required",
        template_sha256="0" * 64,
        analysis_version="test",
        analysis_json={
            "targets": [{
                "id": "pct-1",
                "sheet": "Pricing",
                "cell": "B2",
                "category": "pricing",
                "kind": "percentage",
                "label": "Margin",
            }],
        },
        mapping_json={
            "pct-1": {
                "mode": "source",
                "source_key": "pricing.indirect.margin_pct",
            },
        },
    )
    wrong_unit = [{
        "key": "pricing.indirect.margin_pct",
        "kind": "number",
        "value": 24,
    }]
    resolved, blockers = _resolve_mapping_values(artifact, wrong_unit)
    assert resolved == {}
    assert any("source units changed" in blocker for blocker in blockers)

    invalid_percentage = [{
        "key": "pricing.indirect.margin_pct",
        "kind": "percentage",
        "value": 24,
    }]
    resolved, blockers = _resolve_mapping_values(artifact, invalid_percentage)
    assert resolved == {}
    assert any("Current mapped value" in blocker for blocker in blockers)

    valid_percentage = [{
        "key": "pricing.indirect.margin_pct",
        "kind": "percentage",
        "value": 0.24,
    }]
    resolved, blockers = _resolve_mapping_values(artifact, valid_percentage)
    assert resolved == {"pct-1": 0.24}
    assert blockers == []


def test_general_zero_under_cost_header_is_money() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet.append(["Item", "Total Cost"])
    sheet.append(["Zero placeholder", 0])
    sheet.append(["Blank placeholder", None])

    analysis = inspect_cost_matrix("Buyer Pricing Schedule.xlsx", _save(workbook))
    targets = {target["cell"]: target for target in analysis["targets"]}
    assert targets["B2"]["kind"] == "money"
    assert targets["B3"]["kind"] == "money"


def test_rate_only_sla_matrix_requires_confirmation() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Requirements"
    sheet.append(["Requirement", "Target Rate"])
    sheet.append(["System uptime requirement", 0])
    sheet.append(["Response-time compliance", 0])
    for coordinate in ("B2", "B3"):
        sheet[coordinate].number_format = "0%"

    analysis = inspect_cost_matrix("Performance Rate Matrix.xlsx", _save(workbook))
    assert analysis["targets"]
    assert analysis["classification"]["is_cost_matrix"] is False
    assert analysis["classification"]["possible_cost_matrix"] is True
    assert analysis["classification"]["decision"] == "needs_confirmation"


def test_hidden_support_totals_do_not_create_reconciliation_policy() -> None:
    workbook = Workbook()
    pricing = workbook.active
    pricing.title = "Bid Pricing"
    pricing.append(["Item", "Cost"])
    pricing.append(["Service A", 0])
    pricing.append(["Service B", 0])
    pricing.append(["Total", "=SUM(B2:B3)"])
    for coordinate in ("B2", "B3", "B4"):
        pricing[coordinate].number_format = "$#,##0.00"

    support = workbook.create_sheet("Hidden Calc")
    support["A1"] = "Internal value"
    support["B1"] = 100
    support["A2"] = "Total"
    support["B2"] = "=SUM(B1:B1)"
    support.sheet_state = "hidden"

    analysis = inspect_cost_matrix("Buyer Pricing Schedule.xlsx", _save(workbook))
    assert len(analysis["formulas"]) == 2
    assert len(analysis["reconciliations"]) == 1
    assert analysis["reconciliations"][0]["sheet"] == "Bid Pricing"
    assert analysis["reconciliation_review"]["review_required"] is False
    assert len(analysis["reconciliation_review"]["candidates"]) == 1
    assert analysis["reconciliation_review"]["candidates"][0]["sheet"] == "Bid Pricing"


def test_hidden_rows_and_columns_are_not_mapped_or_extracted(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buyer Pricing"
    sheet.append(["Item", "Cost", None, "Notes"])
    sheet.append([
        "Visible service",
        0,
        None,
        "Offeror must disclose hidden column instructions",
    ])
    sheet.append(["Hidden secret service", 0, None, None])
    sheet.append([
        "Offeror shall use a hidden COTS requirement",
        0,
        None,
        None,
    ])
    for coordinate in ("B2", "B3", "B4"):
        sheet[coordinate].number_format = "$#,##0.00"
    sheet.row_dimensions[3].hidden = True
    sheet.row_dimensions[4].hidden = True
    sheet.column_dimensions["D"].hidden = True
    source = _save(workbook)

    analysis = inspect_cost_matrix("Buyer Pricing.xlsx", source)
    assert [target["cell"] for target in analysis["targets"]] == ["B2"]
    assert any("Hidden rows and columns" in warning for warning in analysis["warnings"])

    path = tmp_path / "Buyer Pricing.xlsx"
    path.write_bytes(source)
    strict, _ = extract_cost_matrix_instruction_text(path, analysis)
    visible, _ = extract_cost_matrix_instruction_text(
        path,
        analysis,
        include_visible_context=True,
    )
    for extracted in (strict, visible):
        assert "Hidden secret service" not in extracted
        assert "hidden COTS requirement" not in extracted
        assert "hidden column instructions" not in extracted


def test_instruction_extraction_separates_prices_from_evaluation_numbers(
    tmp_path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluation Criteria"
    sheet.append(["Criterion", "Value"])
    sheet.append(["Offeror shall enter annual charge", 123456])
    sheet.append(["Contractor must state CHF charge", "CHF 123,456"])
    sheet.append(["Vendor shall state remuneration", "JPY 123,456"])
    sheet.append(["Technical factor weight", 30])
    sheet.append(["Management approach points", 20])
    sheet.append(["Price weighted", 25])
    sheet.append(["Price score: 20.", None])
    sheet.append(["Price points: 25.", None])
    sheet.append([
        "Price weighted at 25 points; proposed amount 123456.",
        None,
    ])
    sheet.append(["Evaluation points 20; total cost 500000.", None])
    sheet.append(["Offeror shall hold price for 120 days", None])
    sheet.append(["Price will be evaluated under FAR 15.404-1", None])
    sheet.append(["Offeror shall state the proposed price is 123456.", None])
    sheet.append(["Vendor must enter Amount 123456.00.", None])
    sheet.append(["Contractor shall enter Total cost: 123,456.", None])
    sheet.append([
        "Offeror shall submit price in Volume 3 by July 21, 2026 at 3:00 PM.",
        None,
    ])
    sheet.append(["Offeror shall hold price through", date(2026, 9, 30)])
    source = _save(workbook)
    analysis = inspect_cost_matrix("Evaluation Criteria.xlsx", source)
    path = tmp_path / "Evaluation Criteria.xlsx"
    path.write_bytes(source)

    strict, _ = extract_cost_matrix_instruction_text(path, analysis)
    visible, _ = extract_cost_matrix_instruction_text(
        path,
        analysis,
        include_visible_context=True,
    )
    for extracted in (strict, visible):
        assert "123456" not in extracted
        assert "123,456" not in extracted
        assert "500000" not in extracted
        assert "Technical factor weight | 30" in extracted
        assert "Management approach points | 20" in extracted
        assert "Price weighted | 25" in extracted
        assert "Price score: 20." in extracted
        assert "Price points: 25." in extracted
        assert "weighted at 25 points" in extracted
        assert "Evaluation points 20" in extracted
        assert "120 days" in extracted
        assert "FAR 15.404-1" in extracted
        assert "Volume 3" in extracted
        assert "July 21, 2026" in extracted
        assert "3:00 PM" in extracted
        assert "2026-09-30" in extracted


def test_total_with_fixed_buyer_constants_is_not_partially_reconciled() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pricing"
    sheet.append(["Item", "Cost"])
    sheet.append(["Buyer fixed fee", 100])
    sheet.append(["Offeror labor", 0])
    sheet.append(["Offeror materials", 0])
    sheet.append(["Total", "=SUM(B2:B4)"])
    for coordinate in ("B2", "B3", "B4", "B5"):
        sheet[coordinate].number_format = "$#,##0.00"

    analysis = inspect_cost_matrix("Buyer Pricing Schedule.xlsx", _save(workbook))
    assert [target["cell"] for target in analysis["targets"]] == ["B3", "B4"]
    assert analysis["reconciliations"] == []
    assert analysis["reconciliation_review"]["review_required"] is True


def test_duplicate_zip_members_are_rejected_before_inspection() -> None:
    source = _simple_pricing_workbook()
    with zipfile.ZipFile(io.BytesIO(source), "r") as archive:
        duplicate_part = archive.read("xl/worksheets/sheet1.xml")
    output = io.BytesIO(source)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(output, "a") as archive:
            archive.writestr("xl/worksheets/sheet1.xml", duplicate_part)

    with pytest.raises(CostMatrixError, match="duplicate package member"):
        inspect_cost_matrix("Buyer Pricing.xlsx", output.getvalue())


def test_macro_part_detection_is_case_insensitive() -> None:
    output = io.BytesIO(_simple_pricing_workbook())
    with zipfile.ZipFile(output, "a") as archive:
        archive.writestr("xl/VBAPROJECT.BIN", b"not executable; package probe only")

    with pytest.raises(CostMatrixError, match="Macro-enabled workbooks"):
        inspect_cost_matrix("Buyer Pricing.xlsx", output.getvalue())


def test_mixed_instruction_sheet_is_byte_preserved_when_pricing_is_completed() -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "The offeror shall explain every pricing assumption."
    instructions["A2"] = "Do not alter this sheet."
    pricing = workbook.create_sheet("Pricing")
    pricing.append(["Item", "Cost"])
    pricing.append(["Implementation", 0])
    pricing["B2"].number_format = "$#,##0.00"
    source = _save(workbook)

    analysis = inspect_cost_matrix("Buyer Cost Workbook.xlsx", source)
    artifact = _artifact(analysis)
    price_target = next(
        target for target in analysis["targets"]
        if target["category"] == "pricing"
    )
    output, _changed = _patch_workbook(
        source,
        targets=_target_map(artifact),
        resolved={price_target["id"]: 999.99},
    )

    with zipfile.ZipFile(io.BytesIO(source), "r") as original_zip:
        original_instructions = original_zip.read("xl/worksheets/sheet1.xml")
    with zipfile.ZipFile(io.BytesIO(output), "r") as output_zip:
        assert output_zip.read("xl/worksheets/sheet1.xml") == original_instructions

    result = load_workbook(io.BytesIO(output), data_only=False)
    try:
        assert result["Instructions"]["A1"].value == (
            "The offeror shall explain every pricing assumption."
        )
        assert result["Instructions"]["A2"].value == "Do not alter this sheet."
    finally:
        result.close()
